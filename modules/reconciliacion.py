# modules/reconciliacion.py
# ==============================================================================
# MÓDULO DE RECONCILIACIÓN FINANCIERA - VERSIÓN CORREGIDA (CRUCE DE GUÍAS)
# ==============================================================================

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from datetime import datetime
import io
import tempfile
import re
import unicodedata
import logging
from typing import Optional
from io import BytesIO

from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import grey, whitesmoke, beige, black

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from database.manager import local_db
from utils.ui import add_back_button, show_module_header

logger = logging.getLogger(__name__)

# ------------------------------------------------------------------------------
# CONSTANTES Y DATOS DE TIENDAS
# ------------------------------------------------------------------------------
TIENDAS_DATA = [
    {"Nombre de Tienda": "Aeropostale - (Cuenca) Mall del Rio", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "CUENCA", "Contacto": "Marco Eras", "Dirección": "Av. Felipe II y Autopista Sur - CC Mall del Rio", "Teléfono": "994570933"},
    {"Nombre de Tienda": "Aeropostale - 6 de Diciembre", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "QUITO", "Contacto": "Micaela Yépez", "Dirección": "Av. 6 de Diciembre y Thomas de Berlanga CC Riocentro UIO", "Teléfono": "987883889"},
    {"Nombre de Tienda": "Aeropostale - Paseo Ambato", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "AMBATO", "Contacto": "Franco Torres", "Dirección": "Manuelita Saenz y Pio Baroja, cerca al parque de las Flores CC Paseo Shopping", "Teléfono": "984951515"},
    {"Nombre de Tienda": "Price Club - Ibarra", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "IBARRA", "Contacto": "Silvia Urcuango", "Dirección": "Av. Victor Gómez Jurado y Rodrigo Miño junto a la cancha La Bombonera", "Teléfono": "982649058"},
]

PRICE_CLUBS = ["Price Club - Portoviejo", "Price Club - Machala", "Price Club - Guayaquil", "Price Club - Ibarra", "Price Club - Cuenca"]
TIENDAS_REGULARES = ['AERO CCI', 'AERO DAULE', 'AERO LAGO AGRIO', 'AERO MALL DEL RIO GYE', 'AERO PLAYAS', 'AEROPOSTALE 6 DE DICIEMBRE', 'AEROPOSTALE BOMBOLI', 'AEROPOSTALE CAYAMBE', 'AEROPOSTALE EL COCA', 'AEROPOSTALE PASAJE', 'AEROPOSTALE PEDERNALES', 'AMBATO', 'BABAHOYO', 'BAHIA DE CARAQUEZ', 'CARAPUNGO', 'CEIBOS', 'CONDADO SHOPPING', 'CUENCA', 'DURAN', 'LA PLAZA SHOPPING', 'MACHALA', 'MAL DEL SUR', 'MALL DEL PACIFICO', 'MALL DEL SOL', 'MANTA', 'MILAGRO', 'MULTIPLAZA RIOBAMBA', 'PASEO AMBATO', 'PENINSULA', 'PORTOVIEJO', 'QUEVEDO', 'RIOBAMBA', 'RIOCENTRO EL DORADO', 'RIOCENTRO NORTE', 'SAN LUIS', 'SANTO DOMINGO']
VENTAS_POR_MAYOR = ["VENTAS POR MAYOR", "MAYORISTA"]
TIENDA_WEB = ["TIENDA WEB", "WEB", "TIENDA MOVIL", "MOVIL"]
FALLAS = ["FALLAS"]
COLORS = {
    'PRICE CLUB': '#0033A0',
    'TIENDAS AEROPOSTALE': '#E4002B',
    'VENTAS POR MAYOR': '#10B981',
    'TIENDA WEB': '#8B5CF6',
    'FALLAS': '#F59E0B',
    'FUNDAS': '#EC4899'
}

# ------------------------------------------------------------------------------
# FUNCIONES AUXILIARES (CORREGIDAS)
# ------------------------------------------------------------------------------
def normalizar_texto(texto) -> str:
    if pd.isna(texto) or texto == "": return ""
    texto = str(texto)
    try:
        texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    except Exception:
        texto = texto.upper()
    texto = re.sub(r"[^A-Za-z0-9\s]", " ", texto.upper())
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

def procesar_subtotal(valor) -> float:
    """Convierte a float cualquier valor de subtotal, manejando comas y puntos."""
    if pd.isna(valor): return 0.0
    try:
        if isinstance(valor, (int, float, np.number)): return float(valor)
        valor_str = str(valor).strip()
        # Reemplazar coma decimal por punto (si es el último separador)
        if ',' in valor_str and '.' in valor_str:
            if valor_str.rfind(',') > valor_str.rfind('.'):
                valor_str = valor_str.replace('.', '').replace(',', '.')
            else:
                valor_str = valor_str.replace(',', '')
        elif ',' in valor_str:
            valor_str = valor_str.replace(',', '.')
        # Eliminar cualquier caracter no numérico excepto punto y signo menos
        valor_str = re.sub(r"[^\d.-]", "", valor_str)
        return float(valor_str) if valor_str else 0.0
    except Exception:
        return 0.0

def limpiar_guia(valor) -> str:
    """
    Limpia el número de guía:
    - Convierte a string sin decimales (elimina .0 final)
    - Elimina espacios y caracteres no alfanuméricos
    - Convierte a mayúsculas
    """
    if pd.isna(valor):
        return ""
    # Convertir a string y eliminar posibles .0 de números enteros leídos como float
    s = str(valor).strip()
    if s.endswith('.0'):
        s = s[:-2]
    # Eliminar caracteres no alfanuméricos (incluyendo espacios, guiones, puntos, etc.)
    s = re.sub(r"[^A-Za-z0-9]", "", s)
    return s.upper()

def obtener_columna_piezas(df: pd.DataFrame) -> Optional[str]:
    posibles = ["PIEZAS", "CANTIDAD", "UNIDADES", "QTY", "CANT", "PZS", "BULTOS", "PIEZA"]
    for col in df.columns:
        if any(p in str(col).upper() for p in posibles):
            return col
    return None

def obtener_columna_fecha(df: pd.DataFrame) -> Optional[str]:
    posibles = ["FECHA", "FECHA ING", "FECHA INGRESO", "FECHA CREACION", "FECHA_ING", "FECHA_CREACION"]
    for col in df.columns:
        if any(p in str(col).upper() for p in posibles):
            return col
    return None

def identificar_tipo_tienda(nombre) -> str:
    if pd.isna(nombre) or nombre == "": return "DESCONOCIDO"
    nombre_upper = normalizar_texto(nombre)
    if "JOFRE" in nombre_upper and "SANTANA" in nombre_upper:
        return "VENTAS AL POR MAYOR"
    nombres_personales = ["ROCIO","ALEJANDRA","ANGELICA","DELGADO","CRUZ","LILIANA",
                          "SALAZAR","RICARDO","SANCHEZ","JAZMIN","ALVARADO","MELISSA",
                          "CHAVEZ","KARLA","SORIANO","ESTEFANIA","GUALPA","MARIA",
                          "JESSICA","PEREZ","LOYO"]
    palabras = nombre_upper.split()
    for p in palabras:
        if len(p) > 2 and p in nombres_personales:
            return "VENTA WEB"
    patrones_fisicas = ["LOCAL","AEROPOSTALE","MALL","PLAZA","SHOPPING","CENTRO COMERCIAL",
                        "CC","C.C","TIENDA","SUCURSAL","PRICE","CLUB","DORADO","CIUDAD",
                        "RIOCENTRO","PASEO","PORTAL","SOL","CONDADO","CITY","CEIBOS",
                        "IBARRA","MATRIZ","BODEGA","FASHION","GYE","QUITO","MACHALA",
                        "PORTOVIEJO","BABAHOYO","MANTA","AMBATO","CUENCA","ALMACEN","PRATI"]
    for patron in patrones_fisicas:
        if patron in nombre_upper:
            return "TIENDA FÍSICA"
    if len(palabras) >= 3 or any(len(p) > 3 for p in palabras):
        return "TIENDA FÍSICA"
    return "VENTA WEB"

def cargar_archivo_local(uploaded_file, nombre):
    if uploaded_file is None:
        return None
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, encoding='utf-8')
        else:
            df = pd.read_excel(uploaded_file, engine='openpyxl')
        return df
    except Exception as e:
        st.error(f"Error al cargar {nombre}: {str(e)}")
        return None

# ------------------------------------------------------------------------------
# PROCESAMIENTO PRINCIPAL CON DEPURACIÓN
# ------------------------------------------------------------------------------
def procesar_gastos_reconciliacion(manifesto, facturas, config):
    """
    Procesa manifiesto y facturas, cruza por guía después de limpieza mejorada.
    """
    # 1. Preparar manifiesto
    st.info("📦 Procesando manifiesto...")
    col_guia_m = config["guia_m"]
    col_subtotal_m = config["subtotal_m"]
    col_ciudad_m = config.get("ciudad_destino", None)
    col_piezas_m = obtener_columna_piezas(manifesto)
    col_fecha_m = obtener_columna_fecha(manifesto)

    df_m = manifesto.copy()
    # Limpieza mejorada de guías
    df_m["GUIA_ORIGINAL"] = df_m[col_guia_m].astype(str).str.strip()
    df_m["GUIA_LIMPIA"] = df_m["GUIA_ORIGINAL"].apply(limpiar_guia)
    df_m["SUBTOTAL_MANIFIESTO"] = df_m[col_subtotal_m].apply(procesar_subtotal)
    
    if col_ciudad_m and col_ciudad_m in df_m.columns:
        df_m["CIUDAD"] = df_m[col_ciudad_m].fillna("DESCONOCIDA").astype(str)
    else:
        df_m["CIUDAD"] = "DESCONOCIDA"
    
    if col_piezas_m:
        df_m["PIEZAS"] = pd.to_numeric(df_m[col_piezas_m], errors="coerce").fillna(1)
    else:
        df_m["PIEZAS"] = 1
    
    if col_fecha_m:
        try:
            df_m["FECHA"] = pd.to_datetime(df_m[col_fecha_m], errors="coerce")
        except:
            df_m["FECHA"] = None
    
    # Destinatario
    col_dest_m = None
    for posible in ["DESTINATARIO", "CONSIGNATARIO", "CLIENTE", "NOMBRE", "RAZON SOCIAL", "DESTINO"]:
        if posible in df_m.columns:
            col_dest_m = posible
            break
    if not col_dest_m:
        for col in df_m.columns:
            if any(p in str(col).upper() for p in ["DEST", "CONSIG", "CLIEN", "NOMB", "RAZON"]):
                col_dest_m = col
                break
    if col_dest_m:
        df_m["DESTINATARIO"] = df_m[col_dest_m].fillna("DESCONOCIDO").astype(str)
    else:
        df_m["DESTINATARIO"] = "TIENDA " + df_m["CIUDAD"]
    
    total_manifiesto_calc = df_m["SUBTOTAL_MANIFIESTO"].sum()
    st.success(f"Manifiesto: {len(df_m)} registros, piezas: {df_m['PIEZAS'].sum():.0f}, total manifiesto: ${total_manifiesto_calc:,.2f}")
    
    # 2. Procesar facturas
    st.info("🧾 Procesando facturas...")
    col_guia_f = config["guia_f"]
    col_subtotal_f = config["subtotal"]
    df_f = facturas.copy()
    df_f["GUIA_ORIGINAL"] = df_f[col_guia_f].astype(str).str.strip()
    df_f["GUIA_LIMPIA"] = df_f["GUIA_ORIGINAL"].apply(limpiar_guia)
    df_f["SUBTOTAL_FACTURA"] = df_f[col_subtotal_f].apply(procesar_subtotal)
    # Agrupar por guía (por si hay varias líneas)
    df_f = df_f.groupby("GUIA_LIMPIA", as_index=False)["SUBTOTAL_FACTURA"].sum()
    total_facturas_calc = df_f["SUBTOTAL_FACTURA"].sum()
    st.success(f"Facturas: {len(df_f)} registros, total facturado: ${total_facturas_calc:,.2f}")
    
    # --- DEPURACIÓN DETALLADA ---
    with st.expander("🔍 Verificar coincidencia de guías (muestras)", expanded=False):
        st.write("**Primeras 10 guías del manifiesto (original vs limpia):**")
        muestras_m = df_m[["GUIA_ORIGINAL", "GUIA_LIMPIA"]].drop_duplicates().head(10)
        st.dataframe(muestras_m)
        
        st.write("**Primeras 10 guías de facturas (original vs limpia):**")
        # Facturas: necesitamos mantener original, pero agrupamos, así que recuperamos original de facturas sin agrupar
        df_f_original = facturas.copy()
        df_f_original["GUIA_ORIGINAL"] = df_f_original[col_guia_f].astype(str).str.strip()
        df_f_original["GUIA_LIMPIA"] = df_f_original["GUIA_ORIGINAL"].apply(limpiar_guia)
        muestras_f = df_f_original[["GUIA_ORIGINAL", "GUIA_LIMPIA"]].drop_duplicates().head(10)
        st.dataframe(muestras_f)
        
        # Contar guías vacías después de limpiar
        guias_m_vacias = df_m[df_m["GUIA_LIMPIA"] == ""].shape[0]
        guias_f_vacias = df_f_original[df_f_original["GUIA_LIMPIA"] == ""].shape[0]
        st.write(f"**Guías vacías en manifiesto después de limpiar:** {guias_m_vacias}")
        st.write(f"**Guías vacías en facturas después de limpiar:** {guias_f_vacias}")
        
        guias_m_set = set(df_m["GUIA_LIMPIA"].dropna().unique())
        guias_f_set = set(df_f["GUIA_LIMPIA"].dropna().unique())
        comunes = guias_m_set.intersection(guias_f_set)
        st.write(f"**Guías únicas en manifiesto (limpias):** {len(guias_m_set)}")
        st.write(f"**Guías únicas en facturas (limpias):** {len(guias_f_set)}")
        st.write(f"**Guías coincidentes:** {len(comunes)}")
        
        if len(comunes) == 0:
            st.error("❌ No hay ninguna guía en común. Revisa las muestras para ver si las guías limpias coinciden.")
            st.stop()
    
    # 3. Unir datos
    st.info("🔗 Uniendo datos por guía limpia...")
    df_completo = pd.merge(df_m, df_f, on="GUIA_LIMPIA", how="left")
    df_completo["ESTADO"] = df_completo["SUBTOTAL_FACTURA"].apply(
        lambda x: "FACTURADA" if pd.notna(x) and x > 0 else "ANULADA"
    )
    df_completo["SUBTOTAL"] = df_completo["SUBTOTAL_FACTURA"].fillna(0)
    df_completo["DIFERENCIA"] = df_completo["SUBTOTAL_MANIFIESTO"] - df_completo["SUBTOTAL"]
    df_completo["TIPO"] = df_completo["DESTINATARIO"].apply(identificar_tipo_tienda)
    df_completo["NOMBRE_NORMALIZADO"] = df_completo["DESTINATARIO"].apply(normalizar_texto)
    
    # Crear grupo para agregación
    def crear_grupo(fila):
        tipo = fila["TIPO"]
        nombre = fila["NOMBRE_NORMALIZADO"]
        ciudad = normalizar_texto(fila["CIUDAD"])
        if tipo == "VENTA WEB":
            palabras = nombre.split()
            if len(palabras) >= 2:
                return f"VENTA WEB - {palabras[0]} {palabras[1]}"
            return f"VENTA WEB - {nombre}"
        elif tipo == "VENTAS AL POR MAYOR":
            return "VENTAS AL POR MAYOR - JOFRE SANTANA"
        elif tipo == "TIENDA FÍSICA":
            grupo_ciudad = f"{ciudad} - " if ciudad != "DESCONOCIDA" else ""
            palabras = nombre.split()
            if palabras:
                return f"{grupo_ciudad}{' '.join(palabras[:3])}"
            return f"{grupo_ciudad}TIENDA"
        else:
            return f"DESCONOCIDO - {nombre[:20]}"
    df_completo["GRUPO"] = df_completo.apply(crear_grupo, axis=1)
    
    guias_facturadas = df_completo[df_completo["ESTADO"] == "FACTURADA"].shape[0]
    guias_anuladas = df_completo[df_completo["ESTADO"] == "ANULADA"].shape[0]
    st.success(f"Unión completada: {len(df_completo)} registros (Facturadas: {guias_facturadas}, Anuladas: {guias_anuladas})")
    
    # 4. Métricas por grupo (solo facturadas)
    st.info("📊 Calculando métricas por grupo...")
    df_facturadas = df_completo[df_completo["ESTADO"] == "FACTURADA"]
    if df_facturadas.empty:
        st.warning("No hay guías facturadas. Verifica la depuración de coincidencia de guías.")
        metricas = pd.DataFrame(columns=["GRUPO","GUIAS","PIEZAS","SUBTOTAL","SUBTOTAL_MANIFIESTO","DIFERENCIA","DESTINATARIOS","CIUDADES","TIPO","PORCENTAJE","PROMEDIO_POR_PIEZA","PIEZAS_POR_GUIA"])
        resumen = pd.DataFrame(columns=["TIPO","TIENDAS","GUIAS","PIEZAS","SUBTOTAL","PORCENTAJE"])
    else:
        metricas = df_facturadas.groupby("GRUPO").agg(
            GUIAS=("GUIA_LIMPIA","count"),
            PIEZAS=("PIEZAS","sum"),
            SUBTOTAL=("SUBTOTAL","sum"),
            SUBTOTAL_MANIFIESTO=("SUBTOTAL_MANIFIESTO","sum"),
            DIFERENCIA=("DIFERENCIA","sum"),
            DESTINATARIOS=("DESTINATARIO", lambda x: ", ".join(sorted(set(str(d) for d in x if pd.notna(d)))[:5])),
            CIUDADES=("CIUDAD", lambda x: ", ".join(sorted(set(str(c) for c in x if pd.notna(c)))[:3])),
            TIPO=("TIPO", lambda x: x.mode()[0] if not x.mode().empty else "DESCONOCIDO")
        ).reset_index()
        total_general = metricas["SUBTOTAL"].sum()
        if total_general > 0:
            metricas["PORCENTAJE"] = (metricas["SUBTOTAL"] / total_general * 100).round(2)
            metricas["PROMEDIO_POR_PIEZA"] = (metricas["SUBTOTAL"] / metricas["PIEZAS"]).round(2)
        else:
            metricas["PORCENTAJE"] = 0.0
            metricas["PROMEDIO_POR_PIEZA"] = 0.0
        metricas["PIEZAS_POR_GUIA"] = (metricas["PIEZAS"] / metricas["GUIAS"]).round(2)
        metricas = metricas.sort_values("SUBTOTAL", ascending=False)
        
        # Resumen por tipo
        resumen = df_facturadas.groupby("TIPO").agg(
            TIENDAS=("GRUPO","nunique"),
            GUIAS=("GUIA_LIMPIA","count"),
            PIEZAS=("PIEZAS","sum"),
            SUBTOTAL=("SUBTOTAL","sum")
        ).reset_index()
        if total_general > 0:
            resumen["PORCENTAJE"] = (resumen["SUBTOTAL"] / total_general * 100).round(2)
        else:
            resumen["PORCENTAJE"] = 0.0
        resumen = resumen.sort_values("SUBTOTAL", ascending=False)
    
    # 5. Validación
    total_manifiesto = df_completo["SUBTOTAL_MANIFIESTO"].sum()
    total_facturas = df_completo["SUBTOTAL"].sum()
    validacion = {
        "total_manifiesto": total_manifiesto,
        "total_facturas": total_facturas,
        "diferencia": abs(total_manifiesto - total_facturas),
        "porcentaje": (abs(total_manifiesto - total_facturas)/total_manifiesto*100) if total_manifiesto > 0 else 0,
        "coincide": abs(total_manifiesto - total_facturas) < 0.01,
        "guias_procesadas": len(df_completo),
        "guias_facturadas": guias_facturadas,
        "guias_anuladas": guias_anuladas,
        "piezas_totales": df_completo["PIEZAS"].sum(),
        "grupos_identificados": len(metricas) if not metricas.empty else 0,
        "porcentaje_facturadas": (guias_facturadas/len(df_completo)*100) if len(df_completo)>0 else 0,
        "porcentaje_anuladas": (guias_anuladas/len(df_completo)*100) if len(df_completo)>0 else 0,
    }
    
    guias_anuladas_df = df_completo[df_completo["ESTADO"] == "ANULADA"].copy()
    return df_completo, metricas, resumen, validacion, guias_anuladas_df

# ------------------------------------------------------------------------------
# GENERACIÓN DE EXCEL Y PDF (sin cambios, pero los incluyo por completitud)
# ------------------------------------------------------------------------------
def generar_excel_con_formato_exacto(metricas_filt, resultado, guias_anuladas, manifesto_original, filtros_aplicados=None):
    try:
        output = BytesIO()
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Reporte"
        hoja1_data = metricas_filt[['GRUPO', 'SUBTOTAL']].copy().sort_values('GRUPO') if not metricas_filt.empty else pd.DataFrame()
        ws1.append(["", ""])
        ws1.append(["", ""])
        ws1.append(["Etiquetas de fila", "Suma de SUBTOTAL"])
        for _, row in hoja1_data.iterrows():
            ws1.append([row['GRUPO'], row['SUBTOTAL']])
        ws1.append(["Total general", hoja1_data['SUBTOTAL'].sum() if not hoja1_data.empty else 0])
        for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws1[3]:
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in range(4, ws1.max_row + 1):
            ws1.cell(row=row, column=2).number_format = '#,##0.00'
        ws1.column_dimensions['A'].width = 50
        ws1.column_dimensions['B'].width = 20

        ws2 = wb.create_sheet(title="Tiendas")
        columnas = ["GRUPO", "GUIAS", "PIEZAS", "SUBTOTAL", "DESTINATARIOS", "CIUDADES", "TIPO", "PORCENTAJE", "PROMEDIO_POR_PIEZA", "PIEZAS_POR_GUIA"]
        ws2.append(columnas)
        if not metricas_filt.empty:
            for _, row in metricas_filt.iterrows():
                ws2.append([row['GRUPO'], int(row['GUIAS']), int(row['PIEZAS']), row['SUBTOTAL'],
                            row['DESTINATARIOS'], row['CIUDADES'], row['TIPO'], row['PORCENTAJE'],
                            row['PROMEDIO_POR_PIEZA'], row['PIEZAS_POR_GUIA']])
        ws2.append(["" for _ in range(len(columnas))])
        ult_fila = ws2.max_row - 1
        total_row = ["" for _ in range(len(columnas))]
        total_row[0] = "Total general"
        total_row[3] = f"=SUBTOTAL(109,D2:D{ult_fila})"
        ws2.append(total_row)
        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=len(columnas)):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws2[1]:
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in range(2, ws2.max_row + 1):
            ws2.cell(row=row, column=4).number_format = '#,##0.00'
            ws2.cell(row=row, column=8).number_format = '0.00'
            ws2.cell(row=row, column=9).number_format = '0.00'
            ws2.cell(row=row, column=10).number_format = '0.00'
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True)
        anchos = [40,10,10,15,50,20,20,15,20,20]
        for i, ancho in enumerate(anchos,1):
            ws2.column_dimensions[get_column_letter(i)].width = ancho

        if not guias_anuladas.empty:
            ws3 = wb.create_sheet(title="Guias Anuladas")
            cols_mostrar = ['GUIA_LIMPIA', 'DESTINATARIO', 'CIUDAD', 'SUBTOTAL_MANIFIESTO', 'PIEZAS']
            cols_existentes = [c for c in cols_mostrar if c in guias_anuladas.columns]
            ws3.append(cols_existentes)
            for _, row in guias_anuladas.iterrows():
                ws3.append([row.get(c, '') for c in cols_existentes])
            for cell in ws3[1]:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

        ws4 = wb.create_sheet(title="Detalle")
        cols_detalle = ['GUIA_LIMPIA', 'ESTADO', 'GRUPO', 'DESTINATARIO', 'CIUDAD', 'PIEZAS', 'SUBTOTAL_MANIFIESTO', 'SUBTOTAL', 'DIFERENCIA', 'TIPO']
        cols_detalle = [c for c in cols_detalle if c in resultado.columns]
        ws4.append(cols_detalle)
        for _, row in resultado.iterrows():
            ws4.append([row.get(c, '') for c in cols_detalle])
        for cell in ws4[1]:
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        for row in range(2, ws4.max_row + 1):
            for i, col in enumerate(cols_detalle, 1):
                if col in ['SUBTOTAL_MANIFIESTO', 'SUBTOTAL', 'DIFERENCIA']:
                    ws4.cell(row=row, column=i).number_format = '#,##0.00'
        wb.save(output)
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"Error generando Excel: {e}")
        return None

def generar_pdf_reporte(metricas, resumen, validacion):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            pdf_path = tmp.name
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=12, alignment=1)
        subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Heading2'], fontSize=12, spaceAfter=6)
        normal_style = styles['Normal']
        elements.append(Paragraph("REPORTE EJECUTIVO - GESTIÓN DE GASTOS POR TIENDA", title_style))
        elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
        elements.append(Spacer(1,12))
        elements.append(Paragraph("MÉTRICAS PRINCIPALES", subtitle_style))
        metricas_data = [
            ["Total Facturado", f"${validacion['total_facturas']:,.2f}"],
            ["Total Manifiesto", f"${validacion['total_manifiesto']:,.2f}"],
            ["Diferencia", f"${validacion['diferencia']:,.2f} ({validacion['porcentaje']:.2f}%)"],
            ["Guías Procesadas", f"{validacion['guias_procesadas']}"],
            ["Guías Facturadas", f"{validacion['guias_facturadas']} ({validacion['porcentaje_facturadas']:.1f}%)"],
            ["Guías Anuladas", f"{validacion['guias_anuladas']} ({validacion['porcentaje_anuladas']:.1f}%)"],
            ["Piezas Totales", f"{validacion['piezas_totales']}"],
            ["Grupos Identificados", f"{validacion['grupos_identificados']}"]
        ]
        metricas_table = Table(metricas_data, colWidths=[200,150])
        metricas_table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),grey), ("TEXTCOLOR",(0,0),(-1,0),whitesmoke),
            ("ALIGN",(0,0),(-1,-1),"CENTER"), ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),10), ("BOTTOMPADDING",(0,0),(-1,0),12),
            ("BACKGROUND",(0,1),(-1,-1),beige), ("GRID",(0,0),(-1,-1),1,black)
        ]))
        elements.append(metricas_table)
        elements.append(Spacer(1,20))

        if not resumen.empty:
            elements.append(Paragraph("RESUMEN POR TIPO DE TIENDA", subtitle_style))
            resumen_data = [["TIPO","TIENDAS","GUÍAS","PIEZAS","SUBTOTAL","%"]]
            for _,row in resumen.iterrows():
                resumen_data.append([row['TIPO'], str(int(row['TIENDAS'])), str(int(row['GUIAS'])), str(int(row['PIEZAS'])), f"${row['SUBTOTAL']:,.2f}", f"{row['PORCENTAJE']:.2f}%"])
            resumen_table = Table(resumen_data, colWidths=[120,80,80,80,100,80])
            resumen_table.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),grey), ("TEXTCOLOR",(0,0),(-1,0),whitesmoke),
                ("ALIGN",(0,0),(-1,-1),"CENTER"), ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,-1),9), ("BOTTOMPADDING",(0,0),(-1,0),12),
                ("BACKGROUND",(0,1),(-1,-1),beige), ("GRID",(0,0),(-1,-1),1,black)
            ]))
            elements.append(resumen_table)
            elements.append(Spacer(1,20))

        elements.append(Paragraph("ANÁLISIS EJECUTIVO", subtitle_style))
        analisis = f"""
        <b>Validación:</b> {"✅ COINCIDENCIA EXACTA" if validacion["coincide"] else "⚠ CON DIFERENCIAS"}<br/>
        <b>Facturación:</b> {validacion["porcentaje_facturadas"]:.1f}% de guías facturadas<br/>
        <b>Anulaciones:</b> {validacion["porcentaje_anuladas"]:.1f}%<br/>
        <b>Recomendación:</b> {"Revisar guías anuladas" if validacion["guias_anuladas"]>0 else "Proceso eficiente"}
        """
        elements.append(Paragraph(analisis, normal_style))
        doc.build(elements)
        return pdf_path
    except Exception as e:
        st.error(f"Error PDF: {e}")
        return None

# ------------------------------------------------------------------------------
# INTERFAZ PRINCIPAL DE STREAMLIT
# ------------------------------------------------------------------------------
# INTERFAZ PRINCIPAL DE STREAMLIT
# ------------------------------------------------------------------------------
def show_reconciliacion_v8():
    add_back_button(key="back_reconciliacion")
    show_module_header("💰 Gestión de Gastos y Transporte", "Conciliación financiera de transporte CD Ibarra y gastos por tienda")
    st.markdown('<div class="module-content">', unsafe_allow_html=True)

    tab_transporte, tab_gastos_tiendas = st.tabs([
        "🚚 Costos de Transporte (CD Ibarra / Courier)",
        "🏪 Conciliación de Gastos por Tienda"
    ])

    # ==========================================================================
    # PESTAÑA 1: ANALISTA DE COSTOS DE TRANSPORTE (CD IBARRA / AEROPOSTALE)
    # ==========================================================================
    with tab_transporte:
        from services.costos_transporte_service import (
            cargar_y_limpiar_manifiesto,
            cargar_y_limpiar_factura,
            leer_distribucion_original,
            procesar_costos_transporte,
            generar_excel_costos_transporte
        )
        from utils.ui import inject_acumatica_css, acu_metric

        inject_acumatica_css()

        st.markdown("""
        <style>
        .transport-hero-card {
            background: linear-gradient(135deg, #002D62 0%, #001737 100%);
            border: 1px solid rgba(207, 10, 44, 0.4);
            border-radius: 12px;
            padding: 18px 24px;
            color: #FFFFFF;
            margin-bottom: 20px;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.2);
        }
        .transport-hero-title {
            font-size: 1.3rem;
            font-weight: 700;
            color: #FFFFFF;
            margin-bottom: 4px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .transport-hero-sub {
            font-size: 0.9rem;
            color: #94A3B8;
        }
        .transport-table-header {
            background-color: #CF0A2C;
            color: #FFFFFF;
            font-weight: bold;
            padding: 8px 12px;
            border-radius: 4px 4px 0 0;
            font-size: 0.95rem;
            margin-top: 15px;
        }
        .kpi-title-box {
            font-size: 1.05rem;
            font-weight: bold;
            color: #002D62;
            margin-top: 15px;
            margin-bottom: 8px;
            border-bottom: 2px solid #002D62;
            padding-bottom: 4px;
        }
        </style>
        <div class="transport-hero-card">
            <div class="transport-hero-title">🚚 Analista de Costos de Transporte + Distribución por Sucursal</div>
            <div class="transport-hero-sub">Fashion Club (Aeropostale) Ecuador • Conciliación mensual de Facturación Courier vs. Manifiesto e Integración en Ventas</div>
        </div>
        """, unsafe_allow_html=True)

        if "transporte_datos" not in st.session_state:
            st.session_state.transporte_datos = {
                "procesado": False,
                "cruce": None,
                "mes": "AGOSTO",
                "anio": "2026"
            }

        # ── Formulario de Carga de Archivos ──
        col_up1, col_up2, col_up3 = st.columns(3)
        with col_up1:
            st.markdown("##### 📦 1. Manifiesto (`.xlsx`)")
            file_manifiesto = st.file_uploader(
                "Manifiesto de Recolección (Hoja Guias)",
                type=["xlsx", "xls"],
                key="uploader_manifiesto_transporte"
            )
        with col_up2:
            st.markdown("##### 📑 2. Factura Courier (`.xlsx`)")
            file_factura = st.file_uploader(
                "Factura Courier (Carga, Documentos, Seguro)",
                type=["xlsx", "xls"],
                key="uploader_factura_transporte"
            )
        with col_up3:
            st.markdown("##### 📊 3. Ventas x Sucursal (`.xlsx`) *(Opcional)*")
            file_distribucion = st.file_uploader(
                "Reporte de Ventas por Sucursal (Distribucción_*.xlsx)",
                type=["xlsx", "xls"],
                key="uploader_distribucion_transporte"
            )

        col_m1, col_m2, col_m3 = st.columns([1, 1, 2])
        meses_list = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        mes_actual_idx = max(0, datetime.now().month - 1)
        with col_m1:
            mes_sel = st.selectbox("Mes del Análisis:", meses_list, index=mes_actual_idx, key="sel_mes_transporte")
        with col_m2:
            anio_sel = st.text_input("Año:", value=str(datetime.now().year), key="inp_anio_transporte")

        btn_procesar = st.button("🚀 Procesar Conciliación e Integración de Costos", type="primary", use_container_width=True)

        if btn_procesar:
            if not file_manifiesto or not file_factura:
                st.error("⚠️ Debes cargar obligatoriamente al menos los 2 primeros archivos: 1) Manifiesto de Recolección y 2) Factura del Courier.")
            else:
                with st.spinner("Ejecutando limpieza, cruce de guías, costeo e integración por sucursal..."):
                    try:
                        # 1. Cargar y limpiar
                        df_manif_limpio = cargar_y_limpiar_manifiesto(file_manifiesto)
                        df_car, df_doc, df_seg = cargar_y_limpiar_factura(file_factura)
                        
                        df_dist_orig = None
                        if file_distribucion:
                            df_dist_orig, _, _ = leer_distribucion_original(file_distribucion)

                        # 2. Cruce y costeo
                        cruce_resultado = procesar_costos_transporte(df_manif_limpio, df_car, df_doc, df_seg, df_distribucion=df_dist_orig)

                        st.session_state.transporte_datos["procesado"] = True
                        st.session_state.transporte_datos["cruce"] = cruce_resultado
                        st.session_state.transporte_datos["mes"] = mes_sel
                        st.session_state.transporte_datos["anio"] = anio_sel
                        st.session_state.transporte_datos["tiene_ventas"] = (df_dist_orig is not None)
                        st.success(f"✅ ¡Conciliación e integración completada exitosamente para {mes_sel} {anio_sel}!")
                    except Exception as e:
                        st.error(f"❌ Error al procesar los archivos: {str(e)}")
                        logger.exception(e)

        # ── Visualización de Resultados ──
        if st.session_state.transporte_datos.get("procesado") and st.session_state.transporte_datos.get("cruce"):
            cruce = st.session_state.transporte_datos["cruce"]
            kpis = cruce["kpis"]
            mes_n = st.session_state.transporte_datos["mes"]
            anio_n = st.session_state.transporte_datos["anio"]
            tiene_ventas = st.session_state.transporte_datos.get("tiene_ventas", False)

            st.divider()

            # Botón de Descarga Excel Oficial al inicio
            excel_bytes = generar_excel_costos_transporte(cruce, mes_nombre=mes_n, anio=anio_n)
            if tiene_ventas:
                nombre_archivo_excel = f"Reporte_Ventas_y_Costos_Transporte_Fashion_Club_{mes_n}_{anio_n}.xlsx"
            else:
                nombre_archivo_excel = f"Analisis_Costos_Transporte_Fashion_Club_{mes_n}_{anio_n}.xlsx"

            st.download_button(
                label=f"📥 Descargar Libro Oficial Excel ({nombre_archivo_excel})",
                data=excel_bytes,
                file_name=nombre_archivo_excel,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

            st.write("")

            # TARJETAS DE KPIS SUPERIORES ESTILO ACUMATICA
            c_k1, c_k2, c_k3, c_k4 = st.columns(4)
            c_k1.markdown(acu_metric("COSTO TOTAL PERÍODO", f"${kpis['costo_total_periodo']:,.2f}", color="blue", icon="💰"), unsafe_allow_html=True)
            c_k2.markdown(acu_metric("PROMEDIO X GUÍA", f"${kpis['costo_promedio_guia']:,.2f}", color="yellow", icon="⚡"), unsafe_allow_html=True)
            c_k3.markdown(acu_metric("GUÍAS FACTURADAS", f"{kpis['guias_carga'] + kpis['guias_doc']:,}", color="green", icon="📦"), unsafe_allow_html=True)
            c_k4.markdown(acu_metric("ANULADAS / NO FACT.", f"{kpis['guias_anuladas']} ({kpis['pct_anulacion']:.1f}%)", color="red", icon="🚫"), unsafe_allow_html=True)

            st.write("")

            # ESTRUCTURA EXACTA DE LA IMAGEN DEL USUARIO
            col_izq, col_der = st.columns(2)

            with col_izq:
                # 1. INDICADORES DE GESTIÓN (KPI)
                st.markdown("<div class='kpi-title-box'>1. INDICADORES DE GESTIÓN (KPI)</div>", unsafe_allow_html=True)
                df_kpis_gest = pd.DataFrame([
                    {"INDICADOR": "Guías en el Manifiesto (total del mes)", "VALOR": f"{kpis['guias_manifiesto']:,}"},
                    {"INDICADOR": "Guías facturadas - Carga", "VALOR": f"{kpis['guias_carga']:,}"},
                    {"INDICADOR": "Guías facturadas - Documentos", "VALOR": f"{kpis['guias_doc']:,}"},
                    {"INDICADOR": "Guías con Seguro contratado", "VALOR": f"{kpis['guias_seguro']:,}"},
                    {"INDICADOR": "Guías ANULADAS / no facturadas", "VALOR": f"{kpis['guias_anuladas']:,}"},
                    {"INDICADOR": "% de Anulación sobre el Manifiesto", "VALOR": f"{kpis['pct_anulacion']:.1f}%"},
                ])
                st.dataframe(df_kpis_gest, hide_index=True, use_container_width=True)

                # 3. CIUDAD-CIUDAD vs. DESDE CD IBARRA
                st.markdown("<div class='kpi-title-box'>3. CIUDAD-CIUDAD vs. DESDE CD IBARRA</div>", unsafe_allow_html=True)
                st.caption("Ver detalle completo en la pestaña 'Analisis Ciudad-Ciudad'")
                df_mov_show = cruce["df_resumen_mov"].copy()
                if not df_mov_show.empty:
                    df_mov_disp = pd.DataFrame({
                        "TIPO DE MOVIMIENTO": df_mov_show["TIPO DE MOVIMIENTO"],
                        "N° GUÍAS": df_mov_show["N° GUIAS"],
                        "COSTO TOTAL (USD)": df_mov_show["COSTO TOTAL (USD)"].apply(lambda x: f"${x:,.2f}"),
                        "% DEL COSTO TOTAL": df_mov_show["% DEL COSTO TOTAL"].apply(lambda x: f"{x:.1f}%")
                    })
                    st.dataframe(df_mov_disp, hide_index=True, use_container_width=True)

            with col_der:
                # 2. COSTOS DEL PERIODO (USD)
                st.markdown("<div class='kpi-title-box'>2. COSTOS DEL PERIODO (USD)</div>", unsafe_allow_html=True)
                df_costos_gest = pd.DataFrame([
                    {"CONCEPTO": "Costo Flete - Carga", "VALOR": f"${kpis['costo_flete_carga']:,.2f}"},
                    {"CONCEPTO": "Costo Flete - Documentos", "VALOR": f"${kpis['costo_flete_doc']:,.2f}"},
                    {"CONCEPTO": "Costo Seguro (todas las guías)", "VALOR": f"${kpis['costo_seguro_total']:,.2f}"},
                    {"CONCEPTO": "COSTO TOTAL DEL PERIODO", "VALOR": f"${kpis['costo_total_periodo']:,.2f}"},
                    {"CONCEPTO": "Costo promedio por guía facturada", "VALOR": f"${kpis['costo_promedio_guia']:,.2f}"},
                ])
                st.dataframe(df_costos_gest, hide_index=True, use_container_width=True)

                # 4. TOP 15 CIUDADES DESTINO POR COSTO
                st.markdown("<div class='kpi-title-box'>4. TOP 15 CIUDADES DESTINO POR COSTO</div>", unsafe_allow_html=True)
                st.caption("Ver detalle completo por ciudad-ciudad y por contenido")
                df_top_show = cruce["df_top_ciudades"].copy()
                if not df_top_show.empty:
                    df_top_disp = pd.DataFrame({
                        "CIUDAD DESTINO": df_top_show["CIUDAD DESTINO"],
                        "N° GUÍAS": df_top_show["N_GUIAS"],
                        "COSTO TOTAL (USD)": df_top_show["COSTO_TOTAL"].apply(lambda x: f"${x:,.2f}"),
                        "% DEL COSTO TOTAL": df_top_show["PCT_COSTO_TOTAL"].apply(lambda x: f"{x:.1f}%")
                    })
                    st.dataframe(df_top_disp.head(8), hide_index=True, use_container_width=True)

            st.write("")

            # Pestañas de Detalle y Análisis Adicionales
            if tiene_ventas and not cruce["df_dist_enriquecida"].empty:
                subtabs_det = st.tabs([
                    "🏬 Distribución con Costos",
                    "📊 Análisis por Contenido",
                    "🏙️ Rutas Ciudad-Ciudad",
                    "🚫 Guías Anuladas / Alertas",
                    "📦 Detalle Carga",
                    "📑 Detalle Documentos",
                    "🛡️ Detalle Seguro",
                    "💬 Resumen Ejecutivo"
                ])
                tab_idx_dist = 0
                tab_idx_cont = 1
                tab_idx_rutas = 2
                tab_idx_anul = 3
                tab_idx_car = 4
                tab_idx_doc = 5
                tab_idx_seg = 6
                tab_idx_res = 7

                with subtabs_det[tab_idx_dist]:
                    st.subheader("Reporte de Ventas por Sucursal con Costos de Transporte")
                    st.caption("Estructura original enriquecida con Costo Flete, Costo Seguro, Costo Total y % Distribución Logística.")
                    
                    df_dist_disp = cruce["df_dist_enriquecida"].copy()
                    st.dataframe(df_dist_disp, hide_index=True, use_container_width=True)
                    
                    st.markdown("---")
                    st.markdown("#### 🏆 Top 5 Sucursales con Mayor Costo de Transporte")
                    stats_s = cruce.get("stats_sucursales", {})
                    if "df_top_5" in stats_s and not stats_s["df_top_5"].empty:
                        col_t5_1, col_t5_2 = st.columns([0.55, 0.45])
                        with col_t5_1:
                            df_t5 = stats_s["df_top_5"][[stats_s["col_codigo"], stats_s["col_sucursal"], "Costo Flete", "Costo Seguro", "Costo Total Transporte", "% Distribución Logística"]].copy()
                            st.dataframe(df_t5, hide_index=True, use_container_width=True)
                        with col_t5_2:
                            fig_t5 = px.bar(stats_s["df_top_5"], x=stats_s["col_sucursal"], y="Costo Total Transporte", title="Top 5 Sucursales por Costo", text_auto="$.2s")
                            fig_t5.update_layout(template="plotly_dark", height=280)
                            st.plotly_chart(fig_t5, use_container_width=True)
            else:
                subtabs_det = st.tabs([
                    "📊 Análisis por Contenido",
                    "🏙️ Rutas Ciudad-Ciudad",
                    "🚫 Guías Anuladas / Alertas",
                    "📦 Detalle Carga",
                    "📑 Detalle Documentos",
                    "🛡️ Detalle Seguro",
                    "💬 Resumen Ejecutivo"
                ])
                tab_idx_cont = 0
                tab_idx_rutas = 1
                tab_idx_anul = 2
                tab_idx_car = 3
                tab_idx_doc = 4
                tab_idx_seg = 5
                tab_idx_res = 6

            with subtabs_det[tab_idx_cont]:
                st.subheader("Análisis de Costos por Categoría de Contenido")
                df_cont_show = cruce["df_resumen_contenido"].copy()
                if not df_cont_show.empty:
                    col_t_c, col_g_c = st.columns([0.6, 0.4])
                    with col_t_c:
                        df_cont_disp = pd.DataFrame({
                            "CATEGORÍA": df_cont_show["CATEGORIA"],
                            "N° GUÍAS": df_cont_show["N_GUIAS"],
                            "COSTO TOTAL (USD)": df_cont_show["COSTO_TOTAL"].apply(lambda x: f"${x:,.2f}"),
                            "% GUÍAS": df_cont_show["PCT_GUIAS"].apply(lambda x: f"{x:.1f}%"),
                            "% COSTO": df_cont_show["PCT_COSTO"].apply(lambda x: f"{x:.1f}%")
                        })
                        st.dataframe(df_cont_disp, hide_index=True, use_container_width=True)
                    with col_g_c:
                        fig_cont = px.pie(df_cont_show, values="COSTO_TOTAL", names="CATEGORIA", title="Distribución de Costo por Contenido", hole=0.35)
                        fig_cont.update_layout(template="plotly_dark", height=320)
                        st.plotly_chart(fig_cont, use_container_width=True)

            with subtabs_det[tab_idx_rutas]:
                st.subheader("Análisis de Costos: Desde CD Ibarra vs. Movimientos Ciudad-Ciudad")
                st.caption("Ciudad-Ciudad = guías cuya CIUDAD ORIGEN no es Ibarra (no salieron del Centro de Distribución).")
                
                df_mov_res = cruce["df_resumen_mov"].copy()
                if not df_mov_res.empty:
                    df_mov_res_disp = pd.DataFrame({
                        "TIPO DE MOVIMIENTO": df_mov_res["TIPO DE MOVIMIENTO"],
                        "N° GUÍAS": df_mov_res["N° GUIAS"],
                        "COSTO TOTAL (USD)": df_mov_res["COSTO TOTAL (USD)"].apply(lambda x: f"${x:,.2f}"),
                        "% GUÍAS": df_mov_res.get("% GUIAS", pd.Series([0,0])).apply(lambda x: f"{x:.1f}%"),
                        "% COSTO": df_mov_res["% DEL COSTO TOTAL"].apply(lambda x: f"{x:.1f}%")
                    })
                    st.dataframe(df_mov_res_disp, hide_index=True, use_container_width=True)

                st.markdown("#### Detalle por Par Ciudad Origen ➔ Ciudad Destino (Solo Ciudad-Ciudad)")
                df_rutas_show = cruce["df_pares_rutas"].copy()
                if not df_rutas_show.empty:
                    df_rutas_disp = pd.DataFrame({
                        "CIUDAD ORIGEN": df_rutas_show["CIUDAD ORIGEN"],
                        "CIUDAD DESTINO": df_rutas_show["CIUDAD DESTINO"],
                        "N° GUÍAS": df_rutas_show["N_GUIAS"],
                        "COSTO TOTAL (USD)": df_rutas_show["COSTO_TOTAL"].apply(lambda x: f"${x:,.2f}"),
                        "COSTO PROMEDIO / GUÍA (USD)": df_rutas_show["COSTO PROMEDIO"].apply(lambda x: f"${x:,.2f}")
                    })
                    st.dataframe(df_rutas_disp, hide_index=True, use_container_width=True)

            with subtabs_det[tab_idx_anul]:
                st.subheader("Guías Anuladas o No Facturadas (Auditoría)")
                df_anul_show = cruce["df_guias_anuladas"]
                if df_anul_show.empty:
                    st.success("✅ Todas las guías del manifiesto fueron facturadas (0 anuladas).")
                else:
                    urgentes = df_anul_show[df_anul_show["NIVEL_ALERTA"] == "REVISION_URGENTE"]
                    if not urgentes.empty:
                        st.error(f"🚨 ATENCIÓN: Se encontraron {len(urgentes)} guías con estado ENTREGADO que NO fueron incluidas en la factura. Requieren reclamo/revisión con el courier.")
                        st.dataframe(urgentes, hide_index=True, use_container_width=True)
                    st.dataframe(df_anul_show, hide_index=True, use_container_width=True)

            with subtabs_det[tab_idx_car]:
                st.subheader("Detalle Facturado Carga")
                st.dataframe(cruce["df_det_carga"], hide_index=True, use_container_width=True)

            with subtabs_det[tab_idx_doc]:
                st.subheader("Detalle Facturado Documentos")
                st.dataframe(cruce["df_det_doc"], hide_index=True, use_container_width=True)

            with subtabs_det[tab_idx_seg]:
                st.subheader("Detalle Seguro Contratado")
                st.dataframe(cruce["df_det_seg"], hide_index=True, use_container_width=True)

            with subtabs_det[tab_idx_res]:
                st.subheader("💬 Resumen Ejecutivo para Chat / Reporte")
                cat_top = cruce["df_resumen_contenido"].iloc[0]["CATEGORIA"] if not cruce["df_resumen_contenido"].empty else "N/A"
                cat_top_val = cruce["df_resumen_contenido"].iloc[0]["COSTO_TOTAL"] if not cruce["df_resumen_contenido"].empty else 0.0
                pct_cc_val = cruce["df_resumen_mov"][cruce["df_resumen_mov"]["TIPO DE MOVIMIENTO"] == "CIUDAD-CIUDAD"]["% DEL COSTO TOTAL"].values[0] if not cruce["df_resumen_mov"].empty else 0.0

                lineas_extra = ""
                stats_s = cruce.get("stats_sucursales", {})
                if "df_top_5" in stats_s and not stats_s["df_top_5"].empty:
                    top5_str = ", ".join([f"{r[stats_s['col_sucursal']]}: ${r['Costo Total Transporte']:,.2f}" for _, r in stats_s["df_top_5"].iterrows()])
                    lineas_extra = f"""
• **Top 5 Sucursales por Costo:** {top5_str}.
• **Guías no asignadas a tiendas (OTROS):** {stats_s.get('otros_guias', 0)} guías (${stats_s.get('otros_costo', 0.0):,.2f} USD)."""

                resumen_texto = f"""*Resumen de Costos de Transporte e Integración — {mes_n} {anio_n} (Fashion Club / Aeropostale)*
• **Costo Total Facturado:** ${kpis['costo_total_periodo']:,.2f} USD ({kpis['guias_carga'] + kpis['guias_doc']:,} guías facturadas a un promedio de ${kpis['costo_promedio_guia']:,.2f}/guía).
• **Fletes y Seguros:** Flete Carga ${kpis['costo_flete_carga']:,.2f} | Flete Documentos ${kpis['costo_flete_doc']:,.2f} | Seguro ${kpis['costo_seguro_total']:,.2f}.
• **Flujo Operativo:** {100 - pct_cc_val:.1f}% Despachos desde CD Ibarra vs. {pct_cc_val:.1f}% Traslados Ciudad-Ciudad.
• **Categoría Principal:** '{cat_top}' representa ${cat_top_val:,.2f} USD del costo total del mes.
• **Auditoría de Manifiesto:** {kpis['guias_manifiesto']:,} guías generadas, {kpis['guias_anuladas']} anuladas/no facturadas ({kpis['pct_anulacion']:.1f}% de anulación).{lineas_extra}"""

                st.code(resumen_texto, language="markdown")

    # ==========================================================================
    # PESTAÑA 2: CONCILIACIÓN DE GASTOS POR TIENDA (EXISTENTE)
    # ==========================================================================
    with tab_gastos_tiendas:
        if 'gastos_datos' not in st.session_state:
            st.session_state.gastos_datos = {
                'manifesto': None, 'facturas': None, 'resultado': None,
                'metricas': None, 'resumen': None, 'validacion': None,
                'guias_anuladas': None, 'procesado': False
            }

        with st.sidebar:
            st.header("📁 Carga Conciliación Tiendas")
            st.markdown("**Formatos:** Excel (.xlsx, .xls) y CSV")
            
            tipo_carga_rec = st.radio("Método:", ["Google Drive", "Manual"], horizontal=True, key="radio_tipo_carga_tiendas")
            
            if tipo_carga_rec == "Google Drive":
                from services.drive_service import _obtener_servicio_drive, listar_archivos_excel_recientes, descargar_archivo_drive
                try:
                    drive_service = _obtener_servicio_drive()
                    archivos_recientes = listar_archivos_excel_recientes(drive_service, limit=15)
                    if not archivos_recientes:
                        st.warning("No se encontraron archivos en Drive.")
                    else:
                        opciones_arch = {f"{a['name']} ({a['createdTime'][:10]})": a['id'] for a in archivos_recientes}
                        
                        idx_m, idx_f = 0, 0
                        for i, name in enumerate(opciones_arch.keys()):
                            if "manifiesto" in name.lower() or "guia" in name.lower(): idx_m = i
                            if "factura" in name.lower(): idx_f = i
                            
                        sel_m = st.selectbox("Manifiesto:", list(opciones_arch.keys()), index=idx_m, key="sel_drive_manif")
                        sel_f = st.selectbox("Facturas:", list(opciones_arch.keys()), index=idx_f, key="sel_drive_fact")
                        
                        if st.button("📥 Importar de Drive", type="primary", use_container_width=True, key="btn_drive_import"):
                            with st.spinner("Descargando..."):
                                f_m = descargar_archivo_drive(drive_service, opciones_arch[sel_m])
                                f_f = descargar_archivo_drive(drive_service, opciones_arch[sel_f])
                                f_m.name = sel_m
                                f_f.name = sel_f
                                
                                manifesto = cargar_archivo_local(f_m, "Manifiesto")
                                facturas = cargar_archivo_local(f_f, "Facturas")
                                
                                if manifesto is not None and facturas is not None:
                                    st.session_state.gastos_datos["manifesto"] = manifesto
                                    st.session_state.gastos_datos["facturas"] = facturas
                                    st.rerun()
                except Exception as e:
                    st.error(f"Error con Drive: {e}")
            else:
                uploaded_manifesto = st.file_uploader("Manifiesto (GUIA y SUBTOTAL)", type=["csv", "xlsx", "xls"], key="manifesto_upload")
                uploaded_facturas = st.file_uploader("Facturas (GUIA y VALOR)", type=["csv", "xlsx", "xls"], key="facturas_upload")
                if uploaded_manifesto and uploaded_facturas:
                    if st.button("📥 Cargar Manual", type="primary", use_container_width=True, key="btn_manual_tiendas"):
                        with st.spinner("Cargando..."):
                            manifesto = cargar_archivo_local(uploaded_manifesto, "Manifiesto")
                            facturas = cargar_archivo_local(uploaded_facturas, "Facturas")
                            if manifesto is not None and facturas is not None:
                                st.session_state.gastos_datos["manifesto"] = manifesto
                                st.session_state.gastos_datos["facturas"] = facturas
                                st.rerun()

        if st.session_state.gastos_datos["manifesto"] is not None:
            manifesto = st.session_state.gastos_datos["manifesto"]
            facturas = st.session_state.gastos_datos["facturas"]
            st.header("⚙️ Configuración de Procesamiento")
            st.subheader("🔍 Selecciona las columnas correctas")
            col1, col2 = st.columns(2)
            with col1:
                st.write("**Manifiesto**")
                guia_candidates = [c for c in manifesto.columns if "GUIA" in str(c).upper() or "GUÍA" in str(c).upper()]
                if not guia_candidates:
                    guia_candidates = manifesto.columns.tolist()
                guia_m = st.selectbox("Columna Guía (Manifiesto)", guia_candidates, index=0)
                subtotal_candidates = [c for c in manifesto.columns if any(p in str(c).upper() for p in ["SUBTOTAL", "TOTAL", "VALOR", "FLETE"])]
                if not subtotal_candidates:
                    subtotal_candidates = manifesto.columns.tolist()
                subtotal_m = st.selectbox("Columna Subtotal/Valor (Manifiesto)", subtotal_candidates, index=0)
                ciudad_candidates = [c for c in manifesto.columns if "CIUDAD" in str(c).upper() or "DESTINO" in str(c).upper()]
                ciudad_destino = st.selectbox("Columna Ciudad (opcional)", ["(No usar)"] + ciudad_candidates, index=0)
                if ciudad_destino == "(No usar)":
                    ciudad_destino = None
            with col2:
                st.write("**Facturas**")
                guia_f_candidates = [c for c in facturas.columns if "GUIA" in str(c).upper() or "GUÍA" in str(c).upper()]
                if not guia_f_candidates:
                    guia_f_candidates = facturas.columns.tolist()
                guia_f = st.selectbox("Columna Guía (Facturas)", guia_f_candidates, index=0)
                subtotal_f_candidates = [c for c in facturas.columns if any(p in str(c).upper() for p in ["SUBTOTAL", "TOTAL", "IMPORTE", "VALOR"])]
                if not subtotal_f_candidates:
                    subtotal_f_candidates = facturas.columns.tolist()
                subtotal_f = st.selectbox("Columna Subtotal/Valor (Facturas)", subtotal_f_candidates, index=0)

            config = {
                "guia_m": guia_m, "subtotal_m": subtotal_m,
                "ciudad_destino": ciudad_destino if ciudad_destino else None,
                "guia_f": guia_f, "subtotal": subtotal_f
            }
            if st.button("🚀 Procesar Conciliación Tiendas", type="primary", use_container_width=True, key="btn_proc_tiendas"):
                with st.spinner("Procesando..."):
                    try:
                        resultado, metricas, resumen, validacion, guias_anuladas = procesar_gastos_reconciliacion(manifesto, facturas, config)
                        st.session_state.gastos_datos["resultado"] = resultado
                        st.session_state.gastos_datos["metricas"] = metricas
                        st.session_state.gastos_datos["resumen"] = resumen
                        st.session_state.gastos_datos["validacion"] = validacion
                        st.session_state.gastos_datos["guias_anuladas"] = guias_anuladas
                        st.session_state.gastos_datos["procesado"] = True
                        st.success("✅ Procesamiento completado")
                    except Exception as e:
                        st.error(f"Error en el procesamiento: {str(e)}")
                        logger.exception(e)

        if st.session_state.gastos_datos["procesado"]:
            resultado = st.session_state.gastos_datos["resultado"]
            metricas = st.session_state.gastos_datos["metricas"]
            resumen = st.session_state.gastos_datos["resumen"]
            validacion = st.session_state.gastos_datos["validacion"]
            guias_anuladas = st.session_state.gastos_datos["guias_anuladas"]
            manifesto_original = st.session_state.gastos_datos["manifesto"]

            tabs_t = st.tabs(["📊 Resumen", "✅ Validación", "🏪 Todas las Tiendas", "🚫 Guías Anuladas", "🌎 Geografía", "📦 Peso Volumétrico", "📋 Datos", "💾 Exportar"])

            with tabs_t[0]:
                st.header("📊 Resumen Ejecutivo")
                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("Grupos de Tiendas", f"{validacion['grupos_identificados']}")
                col2.metric("Total Guías", f"{validacion['guias_procesadas']}")
                col3.metric("Guías Facturadas", f"{validacion['guias_facturadas']}")
                col4.metric("Guías Anuladas", f"{validacion['guias_anuladas']}")
                col5.metric("Total Facturado", f"${validacion['total_facturas']:,.2f}")
                st.subheader("Distribución por Tipo de Tienda")
                if not resumen.empty:
                    fig = px.pie(resumen, values="SUBTOTAL", names="TIPO", title="Distribución de Gastos por Tipo", hole=0.4)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No hay datos facturados para mostrar.")
                st.subheader("Resumen por Tipo de Tienda")
                if not resumen.empty:
                    st.dataframe(resumen.style.format({"SUBTOTAL": "${:,.2f}", "PORCENTAJE": "{:.2f}%"}), use_container_width=True)

            with tabs_t[1]:
                st.header("✅ Validación de Totales")
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Total Manifiesto", f"${validacion['total_manifiesto']:,.2f}")
                col2.metric("Total Facturas", f"${validacion['total_facturas']:,.2f}")
                col3.metric("Diferencia", f"${validacion['diferencia']:,.2f}")
                col4.metric("% Diferencia", f"{validacion['porcentaje']:.2f}%")
                if validacion['coincide']:
                    st.success("✅ Los totales coinciden dentro del margen aceptable.")
                else:
                    st.warning(f"⚠️ Diferencia de ${validacion['diferencia']:,.2f} ({validacion['porcentaje']:.2f}%). Revisar guías anuladas.")

            with tabs_t[2]:
                st.header("🏪 Gastos por Tienda/Grupo")
                if not metricas.empty:
                    st.dataframe(metricas.style.format({
                        "SUBTOTAL": "${:,.2f}", "PORCENTAJE": "{:.2f}%",
                        "PROMEDIO_POR_PIEZA": "${:,.2f}", "PIEZAS_POR_GUIA": "{:.2f}"
                    }), use_container_width=True)
                else:
                    st.warning("No hay métricas para mostrar (todas las guías están anuladas).")

            with tabs_t[3]:
                st.header("🚫 Guías Anuladas")
                if not guias_anuladas.empty:
                    st.dataframe(guias_anuladas[["GUIA_LIMPIA","DESTINATARIO","CIUDAD","SUBTOTAL_MANIFIESTO","PIEZAS"]], use_container_width=True)
                    st.download_button("Descargar anuladas CSV", data=guias_anuladas.to_csv(index=False), file_name="anuladas.csv", mime="text/csv")
                else:
                    st.success("✅ No hay guías anuladas.")

            with tabs_t[4]:
                st.header("🌎 Distribución Geográfica")
                if 'CIUDAD' in resultado.columns and not resultado.empty:
                    ciudad_data = resultado[resultado["ESTADO"]=="FACTURADA"].groupby("CIUDAD")["SUBTOTAL"].sum().reset_index().sort_values("SUBTOTAL", ascending=False)
                    if not ciudad_data.empty:
                        fig = px.bar(ciudad_data.head(15), x="SUBTOTAL", y="CIUDAD", orientation='h', title="Top Ciudades por Gasto")
                        st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No hay datos geográficos disponibles.")

            with tabs_t[5]:
                st.header("📦 Peso Volumétrico por Tienda")
                if 'DESTINATARIO' in resultado.columns and 'PIEZAS' in resultado.columns:
                    vol_data = resultado[resultado["ESTADO"]=="FACTURADA"].groupby("DESTINATARIO")["PIEZAS"].sum().reset_index().sort_values("PIEZAS", ascending=False)
                    if not vol_data.empty:
                        fig = px.bar(vol_data.head(20), x='PIEZAS', y='DESTINATARIO', orientation='h', title="Peso Volumétrico (Piezas) por Destinatario", color='PIEZAS', color_continuous_scale='Sunset')
                        st.plotly_chart(fig, use_container_width=True)
                        st.dataframe(vol_data.rename(columns={'DESTINATARIO': 'Tienda', 'PIEZAS': 'Total Piezas'}), use_container_width=True)
                else:
                    st.warning("Las columnas DESTINATARIO o PIEZAS no están mapeadas.")

            with tabs_t[6]:
                st.header("📋 Datos Detallados")
                st.dataframe(resultado.head(100), use_container_width=True)

            with tabs_t[7]:
                st.header("💾 Exportar Resultados")
                excel_data = generar_excel_con_formato_exacto(metricas, resultado, guias_anuladas, manifesto_original)
                if excel_data:
                    st.download_button("📥 Descargar Excel", data=excel_data, file_name=f"reconciliacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                pdf_path = generar_pdf_reporte(metricas, resumen, validacion)
                if pdf_path:
                    with open(pdf_path, "rb") as f:
                        st.download_button("📄 Descargar PDF", data=f, file_name=f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf", mime="application/pdf")

        else:
            st.info("👆 Carga los archivos desde el panel lateral y selecciona las columnas correctas.")

    st.markdown('</div>', unsafe_allow_html=True)

