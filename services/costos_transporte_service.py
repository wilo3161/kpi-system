import os
import sys
import io
import re
import unicodedata
from datetime import datetime
from typing import Dict, Any, Tuple, Optional, List
import pandas as pd
import numpy as np

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# 1. NORMALIZACIÓN Y HELPERS
# ==============================================================================

def normalizar_texto_transporte(val) -> str:
    """Normaliza texto para cruce seguro: mayúsculas, sin tildes ni espacios extras."""
    if pd.isna(val) or val is None:
        return ""
    texto = str(val).strip()
    try:
        texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    except Exception:
        pass
    texto = re.sub(r"\s+", " ", texto.upper()).strip()
    return texto

def limpiar_numero_guia(val) -> str:
    """Limpia el número de guía eliminando .0 y caracteres invisibles."""
    if pd.isna(val) or val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    # Quitar comillas o espacios
    s = re.sub(r"[\s\'\"]", "", s).upper()
    return s

def es_guia_valida(guia: str) -> bool:
    """Valida que la guía tenga formato de guía real y no sea fila de cierre ni conteo."""
    if not guia:
        return False
    g = str(guia).strip().upper()
    palabras_invalidas = [
        "TOTAL", "ENTREGADO", "FIRMA", "DESPACHADO", "RECOLECCION", 
        "FECHA", "RECIBE", "MANIFIESTO", "SUBTOTAL", "OBSERVACION",
        "SUMA", "PIEZAS", "GUIAS", "GUIA", "FLETE", "VALOR", "NAN", "NONE"
    ]
    if any(p in g for p in palabras_invalidas):
        return False
    if g.endswith(".0"):
        g = g[:-2]
    # Descartar conteos simples de filas (ej. 1440, 169)
    if g.isdigit() and len(g) < 6:
        return False
    # Guías válidas: LC + dígitos, G + dígitos, o números de 6 o más dígitos
    if re.match(r"^(LC\d+|G\d+|\d{6,})$", g):
        return True
    if len(g) >= 6 and re.search(r"\d", g) and len(g) <= 25:
        return True
    return False

def parse_float_seguro(val) -> float:
    """Convierte valor monetario a float."""
    if pd.isna(val) or val is None:
        return 0.0
    if isinstance(val, (int, float, np.number)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(re.sub(r"[^\d.-]", "", s))
    except Exception:
        return 0.0

# ==============================================================================
# 2. CLASIFICADOR DE CONTENIDO
# ==============================================================================

def clasificar_contenido_transporte(contenido_raw: str, producto_raw: str = "") -> str:
    """
    Clasificador según las 10 reglas estrictas del centro de distribución:
    1. DOCUMENTOS
    2. SIN ESPECIFICAR
    3. PUBLICIDAD
    4. VENTA WEB
    5. DEVOLUCIÓN
    6. MERCADERÍA
    7. ACCESORIOS / PRENDAS
    8. MOBILIARIO / EQUIPOS
    9. INSUMOS / ADMINISTRATIVO
    10. OTROS
    """
    cont = normalizar_texto_transporte(contenido_raw)
    prod = normalizar_texto_transporte(producto_raw)
    
    # 1. DOCUMENTOS
    if prod == "DOCUMENTOS_SERV" or any(w in cont for w in ["DOCUMENTO", "DOCUMENTOS", "DOCS", "DOC ", " SOBRE", "SOBRE "]) or cont == "DOC":
        return "DOCUMENTOS"
    
    # 2. SIN ESPECIFICAR
    if not cont or cont in ["", "NAN", "NONE", "NULL", "-", "SD", "S/D"]:
        return "SIN ESPECIFICAR"
    
    # 3. PUBLICIDAD
    if any(p in cont for p in ["PUBLICID", "BANNER", "POP", "AFICHE", "DISPLAY"]):
        return "PUBLICIDAD"
    
    # 4. VENTA WEB
    if any(w in cont for w in ["VENTA WEB", "WEB", "ECOMMERCE", "E-COMMERCE"]):
        return "VENTA WEB"
    
    # 5. DEVOLUCIÓN
    if any(d in cont for d in ["DEVOLU", "CAMBIO"]):
        return "DEVOLUCIÓN"
    
    # 6. MERCADERÍA
    if any(m in cont for m in ["MERCADER", "CARGA", "FALLAS", "PACK"]):
        return "MERCADERÍA"
    
    # 7. ACCESORIOS / PRENDAS
    accesorios_keywords = [
        "GAFA", "GORRA", "FUNDA", "MANIQUI", "PERFUME", "SANDALIA", 
        "ZANDALIA", "PRENDA", "UNIFORME", "ZAPATO", "CAMISETA", "JEAN", 
        "ROPA", "POLO", "ZAPATILLA", "KID"
    ]
    if any(k in cont for k in accesorios_keywords):
        return "ACCESORIOS / PRENDAS"
    
    # 8. MOBILIARIO / EQUIPOS
    mobiliario_keywords = ["MUEBLE", "PLANCHA", "EQUIPO", "COMPUTADORA", "IMPRESORA", "SILLA", "MESA", "INTERNET", "CANASTILLA"]
    if any(k in cont for k in mobiliario_keywords):
        return "MOBILIARIO / EQUIPOS"
    
    # 9. INSUMOS / ADMINISTRATIVO
    insumos_keywords = [
        "CINTA", "ROLLO", "CARTON", "TABLA", "FACTURA", "REPORTE", 
        "FICHA", "MEDICINA", "TRABAJO SOCIAL", "UPC", "CABLE", "LONA", "PAPEL", "ETIQUETA"
    ]
    if any(k in cont for k in insumos_keywords):
        return "INSUMOS / ADMINISTRATIVO"
    
    # 10. OTROS
    return "OTROS"

# ==============================================================================
# 3. PARSERS DE ARCHIVOS
# ==============================================================================

def _es_fila_total(row) -> bool:
    """Detecta si una fila es de resumen, total o firma."""
    row_str = " ".join([str(v).upper() for v in row.values if pd.notna(v)])
    return any(k in row_str for k in ["TOTAL GENERAL", "TOTAL A ENTREGAR", "TOTAL CARGA", "TOTAL DOC", "TOTAL:", "FIRMA", "CANTIDAD DE GUIAS", "CANTIDAD GUIAS"])

def _extraer_serie_limpia(df: pd.DataFrame, col_name: str, default_val: Any = "") -> pd.Series:
    """Extrae una columna como Series de forma segura, incluso si hay columnas duplicadas."""
    if col_name not in df.columns:
        return pd.Series([default_val] * len(df), index=df.index)
    col_data = df[col_name]
    if isinstance(col_data, pd.DataFrame):
        col_data = col_data.iloc[:, 0]
    return col_data

def cargar_y_limpiar_manifiesto(file_or_bytes) -> pd.DataFrame:
    """Carga el archivo de manifiesto detectando encabezado y filtrando filas basura."""
    if hasattr(file_or_bytes, "seek"):
        file_or_bytes.seek(0)
        
    excel_file = pd.ExcelFile(file_or_bytes)
    sheet_name = excel_file.sheet_names[0]
    for s in excel_file.sheet_names:
        if "GUIA" in s.upper() or "MANIF" in s.upper():
            sheet_name = s
            break
            
    df_raw = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
    
    header_idx = 1
    for idx, row in df_raw.head(10).iterrows():
        row_str = " ".join([str(c).upper() for c in row.values if pd.notna(c)])
        if "GUIA" in row_str and ("DESTINO" in row_str or "DESTINATARIO" in row_str or "ORIGEN" in row_str):
            header_idx = idx
            break
            
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_idx)
    
    # Filtrar filas de totales del pie
    df = df[~df.apply(_es_fila_total, axis=1)].copy()
    
    col_map = {}
    for c in df.columns:
        c_norm = normalizar_texto_transporte(c)
        if "GUIA" in c_norm: col_map[c] = "GUIA"
        elif "CIUDAD ORIGEN" in c_norm or ("CIUDAD" in c_norm and "ORI" in c_norm) or c_norm == "ORIGEN": col_map[c] = "CIUDAD ORIGEN"
        elif "CIUDAD DESTINO" in c_norm or ("CIUDAD" in c_norm and "DES" in c_norm) or c_norm == "DESTINO": col_map[c] = "CIUDAD DESTINO"
        elif "DESTINATARIO" in c_norm or "CLIENTE" in c_norm: col_map[c] = "DESTINATARIO"
        elif "TELEFONO" in c_norm or "TEL" in c_norm: col_map[c] = "TELEFONO"
        elif "DIRECCION" in c_norm or "DIR" in c_norm: col_map[c] = "DIRECCION DESTINATARIO"
        elif "CONTENIDO" in c_norm or "DETALLE" in c_norm: col_map[c] = "CONTENIDO"
        elif "PRODUCTO" in c_norm or "SERVICIO" in c_norm: col_map[c] = "PRODUCTO"
        elif "VALOR DECLARADO" in c_norm or "VAL. DEC" in c_norm: col_map[c] = "VALOR DECLARADO"
        elif "PESO" in c_norm: col_map[c] = "PESO"
        elif "ESTADO" in c_norm: col_map[c] = "ESTADO"
        elif "FECHA ENTREGA" in c_norm: col_map[c] = "FECHA ENTREGA"
        elif "RECIBE" in c_norm: col_map[c] = "RECIBE"
        elif "FECHA CREACION" in c_norm or "FECHA EMISION" in c_norm or "FECHA" in c_norm: col_map[c] = "FECHA CREACION"
        elif "COSTO FLETE" in c_norm or "FLETE" in c_norm: col_map[c] = "COSTO FLETE"
        
    df = df.rename(columns=col_map)
    df = df.loc[:, ~df.columns.duplicated(keep='first')].copy()
    
    if "GUIA" not in df.columns:
        raise ValueError("El manifiesto no contiene una columna 'GUIA' identificable.")
        
    guia_series = _extraer_serie_limpia(df, "GUIA")
    df["GUIA"] = guia_series.apply(limpiar_numero_guia)
    df = df[df["GUIA"].apply(es_guia_valida)].copy()
    
    for col in ["CIUDAD ORIGEN", "CIUDAD DESTINO", "DESTINATARIO", "TELEFONO", "DIRECCION DESTINATARIO", "CONTENIDO", "PRODUCTO", "ESTADO", "RECIBE", "FECHA CREACION", "FECHA ENTREGA"]:
        s_col = _extraer_serie_limpia(df, col, default_val="")
        df[col] = s_col.fillna("").astype(str).str.strip()
        
    df["CIUDAD ORIGEN"] = df["CIUDAD ORIGEN"].apply(lambda x: normalizar_texto_transporte(x) if x else "IBARRA")
    df["CIUDAD DESTINO"] = df["CIUDAD DESTINO"].apply(normalizar_texto_transporte)
    
    df["VALOR DECLARADO"] = _extraer_serie_limpia(df, "VALOR DECLARADO", default_val=0.0).apply(parse_float_seguro)
    df["PESO"] = _extraer_serie_limpia(df, "PESO", default_val=0.0).apply(parse_float_seguro)

    return df

def cargar_y_limpiar_factura(file_or_bytes) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Carga y clasifica las 3 pestañas de la factura del Courier:
    - Documentos (DOC)
    - Carga (CAR)
    - Seguro
    Elimina la última fila de total de cada pestaña.
    """
    if hasattr(file_or_bytes, "seek"):
        file_or_bytes.seek(0)
        
    excel_file = pd.ExcelFile(file_or_bytes)
    sheet_names = excel_file.sheet_names
    
    df_doc = pd.DataFrame()
    df_car = pd.DataFrame()
    df_seg = pd.DataFrame()
    
    for sheet in sheet_names:
        df_raw = pd.read_excel(excel_file, sheet_name=sheet, header=None)
        header_idx = 1
        for idx, row in df_raw.head(8).iterrows():
            row_str = " ".join([str(c).upper() for c in row.values if pd.notna(c)])
            if "GUIA" in row_str and ("FLETE" in row_str or "SUBT" in row_str or "SEGURO" in row_str or "SER" in row_str):
                header_idx = idx
                break
                
        df_sheet = pd.read_excel(excel_file, sheet_name=sheet, header=header_idx)
        df_sheet = df_sheet[~df_sheet.apply(_es_fila_total, axis=1)].copy()
        
        col_map = {}
        for c in df_sheet.columns:
            cn = normalizar_texto_transporte(c)
            if "GUIA" in cn: col_map[c] = "GUIA"
            elif "FECH" in cn and "REM" in cn: col_map[c] = "FECHA REM"
            elif "FECHA" in cn: col_map[c] = "FECHA"
            elif "CIU" in cn and "ORI" in cn: col_map[c] = "CIUDAD ORIGEN"
            elif "CIU" in cn and "DES" in cn: col_map[c] = "CIUDAD DESTINO"
            elif cn in ["ORIGEN", "CIUDAD ORIGEN"]: col_map[c] = "CIUDAD ORIGEN"
            elif cn in ["DESTINO", "CIUDAD DESTINO"]: col_map[c] = "CIUDAD DESTINO"
            elif cn in ["SER.", "SER", "SERVICIO"]: col_map[c] = "SERVICIO"
            elif cn in ["TRA.", "TRA", "TRAYECTO"]: col_map[c] = "TRAYECTO"
            elif cn in ["PIE.", "PIE", "PIEZAS"]: col_map[c] = "PIEZAS"
            elif cn in ["PES.", "PES", "PESO"]: col_map[c] = "PESO"
            elif cn in ["FLETE", "COSTO FLETE"]: col_map[c] = "FLETE"
            elif cn in ["DESC.", "DESC", "DESCUENTO"]: col_map[c] = "DESCUENTO"
            elif cn in ["SUBT.", "SUBT", "SUBTOTAL"]: col_map[c] = "SUBTOTAL"
            elif "VAL" in cn and "DEC" in cn: col_map[c] = "VALOR DECLARADO"
            elif "SEGURO" in cn: col_map[c] = "SEGURO_TASA" if "TASA" in cn or "%" in cn else "SEGURO"
            
        df_sheet = df_sheet.rename(columns=col_map)
        df_sheet = df_sheet.loc[:, ~df_sheet.columns.duplicated(keep='first')].copy()
        
        if "GUIA" not in df_sheet.columns:
            continue
            
        guia_s = _extraer_serie_limpia(df_sheet, "GUIA")
        df_sheet["GUIA"] = guia_s.apply(limpiar_numero_guia)
        df_sheet = df_sheet[df_sheet["GUIA"].apply(es_guia_valida)].copy()
        
        sheet_upper = sheet.upper()
        cols_upper = [str(c).upper() for c in df_sheet.columns]
        
        serv_series = _extraer_serie_limpia(df_sheet, "SERVICIO", default_val="")
        has_doc_serv = serv_series.fillna("").astype(str).str.upper().str.contains("DOC").any()
        has_car_serv = serv_series.fillna("").astype(str).str.upper().str.contains("CAR").any()
        
        es_seguro = "SEGURO" in sheet_upper or "SEGURO" in cols_upper or "VALOR DECLARADO" in df_sheet.columns
        es_doc = "DOC" in sheet_upper or has_doc_serv
        es_car = "CAR" in sheet_upper or "CARGA" in sheet_upper or has_car_serv
        
        if es_seguro and df_seg.empty:
            df_seg = df_sheet
        elif es_doc and df_doc.empty:
            df_doc = df_sheet
        elif es_car and df_car.empty:
            df_car = df_sheet
        else:
            if "SEGURO" in cols_upper and df_seg.empty:
                df_seg = df_sheet
            elif df_car.empty:
                df_car = df_sheet
            elif df_doc.empty:
                df_doc = df_sheet
                
    for df_curr in [df_car, df_doc, df_seg]:
        if not df_curr.empty:
            if "FLETE" in df_curr.columns:
                df_curr["FLETE"] = _extraer_serie_limpia(df_curr, "FLETE", default_val=0.0).apply(parse_float_seguro)
            if "SUBTOTAL" in df_curr.columns:
                df_curr["SUBTOTAL"] = _extraer_serie_limpia(df_curr, "SUBTOTAL", default_val=0.0).apply(parse_float_seguro)
            if "PIEZAS" in df_curr.columns:
                df_curr["PIEZAS"] = pd.to_numeric(_extraer_serie_limpia(df_curr, "PIEZAS", default_val=1), errors="coerce").fillna(1).astype(int)
            if "PESO" in df_curr.columns:
                df_curr["PESO"] = _extraer_serie_limpia(df_curr, "PESO", default_val=0.0).apply(parse_float_seguro)
            if "CIUDAD ORIGEN" in df_curr.columns:
                df_curr["CIUDAD ORIGEN"] = _extraer_serie_limpia(df_curr, "CIUDAD ORIGEN", default_val="IBARRA").apply(normalizar_texto_transporte)
            if "CIUDAD DESTINO" in df_curr.columns:
                df_curr["CIUDAD DESTINO"] = _extraer_serie_limpia(df_curr, "CIUDAD DESTINO", default_val="").apply(normalizar_texto_transporte)
                
    return df_car, df_doc, df_seg

# ==============================================================================
# 4. DISTRIBUCIÓN LOGÍSTICA POR SUCURSAL (GENERADA AUTOMÁTICAMENTE)
# ==============================================================================

CATALOGO_SUCURSALES_OFICIAL = [
    {
        "codigo": "1.001", 
        "sucursal": "MATRIZ / CD IBARRA", 
        "keywords": ["MATRIZ", "CD IBARRA", "BODEGA IBARRA", "CENTRO DE DISTRIBUCION", "CENTRO DISTRIBUCION IBARRA", "CENTRO LOGISTICO IBARRA"]
    },
    {
        "codigo": "2001", 
        "sucursal": "MALL DEL RIO CUENCA", 
        "keywords": ["MALL DEL RIO CUENCA", "AEROPOSTALE (CUENCA) MALL DEL RIO", "AEROPOSTALE ALTOS DEL RIO", "ALTOS DEL RIO", "CUENCA MALL", "MALL DEL RIO", "AERO CUENCA", "AEROPOSTALE CUENCA", "FELIPE II", "MARCO ERAS", "ADRIAN"]
    },
    {
        "codigo": "2002", 
        "sucursal": "RIOBAMBA", 
        "keywords": ["MULTIPLAZA RIOBAMBA", "PASEO SHOPPING RIOBAMBA", "PESO SHOPPING RIOBAMBA", "PASEO RIOBAMBA", "AERO RIOBAMBA", "LIZARZABURU", "MARIA FERNANDA IBARRA", "JENNIFER JIMENEZ", "RIOBAMBA"]
    },
    {
        "codigo": "2003", 
        "sucursal": "PASEO AMBATO", 
        "keywords": ["PASEO SHOPPING AMBATO", "PASEO AMBATO", "AMBATO PASEO", "MALL DE LOS ANDES", "AERO AMBATO", "PIO BAROJA", "MANUELITA SAENS", "FRANCO TORRES", "GABRIELA URRUTIA", "AMBATO"]
    },
    {
        "codigo": "2004", 
        "sucursal": "MALL DEL PACIFICO MANTA", 
        "keywords": ["MALL DEL PACIFICO", "PACIFICO MANTA", "MANTA PACIFICO", "MANTA SHOPPING", "PASEO SHOPPING MANTA", "AERO MANTA", "KARINA FIGUEROA", "YENNY ALVIA", "4 DE NOVIEMBRE"]
    },
    {
        "codigo": "2005", 
        "sucursal": "CONDADO SHOPPING", 
        "keywords": ["CONDADO SHOPPING", "EL CONDADO", "CONDADO", "AERO CONDADO", "MARISCAL SUCRE", "MATEO RECALDE"]
    },
    {
        "codigo": "2006", 
        "sucursal": "SAN LUIS SHOPPING", 
        "keywords": ["SAN LUIS SHOPPING", "SAN LUIS", "AERO SAN LUIS", "SANGOLQUI", "GENERAL RUMINAHUI", "KARINA PROANO"]
    },
    {
        "codigo": "2007", 
        "sucursal": "SANTO DOMINGO", 
        "keywords": ["AEROPOSTALE BOMBOLI", "BOMBOLI SHOPPING", "BOMBOLI", "SANTO DOMINGO", "AERO SANTO DOMINGO", "JOSSELYN NAVARRETE", "MATEO FRUTO", "ABRAHAM CALAZACON"]
    },
    {
        "codigo": "2008", 
        "sucursal": "MALL DEL SUR GUAYAQUIL", 
        "keywords": ["MALL DEL SUR", "MAL DEL SUR", "SUR GUAYAQUIL", "AERO SUR", "JUDITH ASUNCION", "HOSPITAL DE IESS"]
    },
    {
        "codigo": "2009", 
        "sucursal": "MALL DEL SOL GUAYAQUIL", 
        "keywords": ["MALL DEL SOL", "SOL GUAYAQUIL", "AERO SOL", "KIARA DAVALOS", "JUAN TANCA MARENGO"]
    },
    {
        "codigo": "2010", 
        "sucursal": "RIOCENTRO EL DORADO", 
        "keywords": ["RIOCENTRO EL DORADO", "RIO CENTRO DORADO", "EL DORADO", "DORADO GUAYAQUIL", "DORADO DAULE", "OSCAR ALVARADO", "FEBRES CORDERO", "AURORA"]
    },
    {
        "codigo": "2011", 
        "sucursal": "RIOCENTRO NORTE", 
        "keywords": ["RIOCENTRO NORTE", "RIO CENTRO NORTE", "NORTE GUAYAQUIL", "AERO NORTE", "DORIS ZAMBRANO", "URB. ALCANCE"]
    },
    {
        "codigo": "2012", 
        "sucursal": "RIOCENTRO CEIBOS", 
        "keywords": ["RIOCENTRO CEIBOS", "RIO CENTRO CEIBOS", "LOS CEIBOS", "CEIBOS", "AERO CEIBOS", "DEL BOMBERO", "SAN EDUARDO"]
    },
    {
        "codigo": "2013", 
        "sucursal": "PASEO SHOPPING PORTOVIEJO", 
        "keywords": ["PASEO SHOPPING PORTOVIEJO", "PORTOVIEJO SHOPPING", "PASEO PORTOVIEJO", "AERO PORTOVIEJO", "GISSEL LOOR", "JORGE WASHINGTON"]
    },
    {
        "codigo": "2014", 
        "sucursal": "PASEO SHOPPING MACHALA", 
        "keywords": ["PASEO SHOPPING MACHALA", "PASEO MACHALA", "AERO MACHALA", "IRIS CARPIO", "PAQUISHA", "MACHALA"]
    },
    {
        "codigo": "2015", 
        "sucursal": "PASEO SHOPPING DURAN", 
        "keywords": ["PASEO SHOPPING DURAN", "PASEO DURAN", "AERO DURAN", "YARITZA CORDOVA", "BOLICHE", "DURAN"]
    },
    {
        "codigo": "2016", 
        "sucursal": "PASEO SHOPPING QUEVEDO", 
        "keywords": ["PASEO SHOPPING QUEVEDO", "QUEVEDO SHOPPING", "PASEO QUEVEDO", "AERO QUEVEDO", "DAYANA LEON", "QUEVEDO"]
    },
    {
        "codigo": "2017", 
        "sucursal": "PASEO SHOPPING BABAHOYO", 
        "keywords": ["PASEO SHOPPING BABAHOYO", "PASEO BABAHOYO", "AERO BABAHOYO", "YOMAIRA SELLAN", "PONCE LUQUE", "BABAHOYO"]
    },
    {
        "codigo": "2018", 
        "sucursal": "CCI IÑAQUITO", 
        "keywords": ["AERO CCI", "CCI INAQUITO", "MALL CCI", "INAQUITO", "CAROLINA PROCEL", "NACIONES UNIDAS"]
    },
    {
        "codigo": "2019", 
        "sucursal": "6 DE DICIEMBRE", 
        "keywords": ["AEROPOSTALE 6 DE DICIEMBRE", "RIOCENTRO 6 DE DICIEMBRE", "RIOCENTRO UIO", "6 DE DICIEMBRE", "SEIS DE DICIEMBRE", "MICAELA YEPEZ", "THOMAS DE BERLANGA"]
    },
    {
        "codigo": "2020", 
        "sucursal": "CARAPUNGO", 
        "keywords": ["PORTAL SHOPPING", "CC CARAPUNGO", "EL PORTAL", "CARAPUNGO", "PORTAL QUITO", "MARIA JOSE BENALCAZAR", "GIOVANNI CALLES"]
    },
    {
        "codigo": "2021", 
        "sucursal": "LA PLAZA SHOPPING MANTA", 
        "keywords": ["LA PLAZA SHOPPING", "LA PLAZA MANTA", "PLAZA MANTA"]
    },
    {
        "codigo": "2022", 
        "sucursal": "CAYAMBE", 
        "keywords": ["AEROPOSTALE CAYAMBE", "ALTOS DE CAYAMBE", "CAYAMBE", "AERO CAYAMBE", "CELESTE CONTRERAS"]
    },
    {
        "codigo": "2023", 
        "sucursal": "EL COCA", 
        "keywords": ["AEROPOSTALE EL COCA", "EL COCA", "PUERTO FRANCISCO DE ORELLANA", "ORELLANA", "ADRIANA ZURITA", "RIO CURARAY"]
    },
    {
        "codigo": "2024", 
        "sucursal": "LAGO AGRIO", 
        "keywords": ["AERO LAGO AGRIO", "LAGO AGRIO", "NUEVA LOJA", "ANGIE MALDONADO", "PASAJE BRAZIL"]
    },
    {
        "codigo": "2025", 
        "sucursal": "PEDERNALES", 
        "keywords": ["AEROPOSTALE PEDERNALES", "PEDERNALES", "AERO PEDERNALES", "MONICA MUNOZ", "GARCIA MORENO"]
    },
    {
        "codigo": "2026", 
        "sucursal": "PASAJE", 
        "keywords": ["AEROPOSTALE PASAJE", "PASAJE", "AERO PASAJE", "JHONNY CUN", "REDONDEL DEL LEON"]
    },
    {
        "codigo": "2027", 
        "sucursal": "DAULE", 
        "keywords": ["AERO DAULE", "DAULE", "PASEO DAULE", "ALISSON RAMIREZ", "PIEDRAHITA"]
    },
    {
        "codigo": "2028", 
        "sucursal": "PLAYAS", 
        "keywords": ["AERO PLAYAS", "PASEO SHOPPING PLAYAS", "GENERAL VILLAMIL", "VILLAMIL PLAYAS", "PLAYAS", "STEVEN ORTIZ"]
    },
    {
        "codigo": "2029", 
        "sucursal": "PENINSULA", 
        "keywords": ["PASEO SHOPPING LA PENINSULA", "PASEO PENINSULA", "PENINSULA", "SANTA ELENA", "SALINAS", "LA LIBERTAD", "KENNY BOHORQUEZ", "CARLOS ESPINOSA"]
    },
    {
        "codigo": "2030", 
        "sucursal": "BAHIA DE CARAQUEZ", 
        "keywords": ["PASEO SHOPPING BAHIA", "BAHIA DE CARAQUEZ", "BAHIA DE CARAQUE", "BAHIA", "AERO BAHIA", "NAYELY OREJUELA", "3 DE NOVIEMBRE"]
    },
    {
        "codigo": "2031", 
        "sucursal": "MILAGRO", 
        "keywords": ["PASEO SHOPPING MILAGRO", "PASEO MILAGRO", "MILAGRO", "AERO MILAGRO", "LADY SILVA", "12 DE OCTUBRE"]
    },
    {
        "codigo": "3001", 
        "sucursal": "PRICE CLUB IBARRA", 
        "keywords": ["PRICE CLUB IBARRA", "PRICE IBARRA", "PRICE CLUB CD IBARRA", "SILVIA URCUANGO", "LA BOMBONERA"]
    },
    {
        "codigo": "3002", 
        "sucursal": "PRICE CLUB PORTOVIEJO", 
        "keywords": ["PRICE CLUB PORTOVIEJO", "PRICE PORTOVIEJO", "PRICE CLUB MANABI", "DAYANA MERCHAN"]
    },
    {
        "codigo": "3003", 
        "sucursal": "PRICE CLUB MACHALA", 
        "keywords": ["PRICE CLUB MACHALA", "PRICE MACHALA", "PRICE CLUB EL ORO", "YULEYSI DELGADO", "ORO PLAZA"]
    },
    {
        "codigo": "3004", 
        "sucursal": "PRICE CLUB GUAYAQUIL", 
        "keywords": ["PRICE CLUB GUAYAQUIL", "PRICE GUAYAQUIL", "PRICE CLUB GYE", "PRICE CLUB CITY MALL", "CITY MALL", "ANGIE DELGADO", "JORDAN GUALE", "PEDRO CARBO", "ESTUARDO SANCHEZ"]
    },
    {
        "codigo": "3005", 
        "sucursal": "PRICE CLUB CUENCA", 
        "keywords": ["PRICE CLUB CUENCA", "PRICE CUENCA", "PRICE CLUB AZUAY"]
    },
]

def obtener_catalogo_tiendas_dinamico() -> List[Dict[str, Any]]:
    """
    Combina el catálogo contable oficial base con cualquier tienda
    registrada dinámicamente en el módulo central de configuración (config.stores_data.TIENDAS_DATA).
    Esto garantiza que si el usuario añade o edita tiendas en el sistema, se reconozcan de inmediato.
    """
    import copy
    catalogo = copy.deepcopy(CATALOGO_SUCURSALES_OFICIAL)
    try:
        from config.stores_data import TIENDAS_DATA, reload_stores_data
        reload_stores_data()
        
        for tienda in TIENDAS_DATA:
            nom_tienda = tienda.get("Nombre de Tienda", "")
            empresa = tienda.get("Empresa", "")
            contacto = tienda.get("Contacto", "")
            direccion = tienda.get("Dirección", "")
            destino = tienda.get("Destino", "")
            ciudad = tienda.get("Ciudad", "")
            
            encontrado = False
            nom_norm = normalizar_texto_transporte(nom_tienda)
            for item in catalogo:
                kw_item = [normalizar_texto_transporte(k) for k in item["keywords"]]
                if nom_norm in kw_item or any(kw in nom_norm for kw in kw_item if len(kw) > 5):
                    for extra in [contacto, direccion, destino, ciudad, nom_tienda]:
                        if extra:
                            ex_n = normalizar_texto_transporte(extra)
                            if ex_n and len(ex_n) >= 4 and ex_n not in item["keywords"]:
                                item["keywords"].append(ex_n)
                    encontrado = True
                    break
                    
            if not encontrado and nom_tienda:
                codigo_sug = "3099" if "PRICE" in empresa.upper() else "2099"
                kw_extra = [normalizar_texto_transporte(x) for x in [nom_tienda, contacto, direccion, destino, ciudad] if x]
                catalogo.append({
                    "codigo": codigo_sug,
                    "sucursal": normalizar_texto_transporte(nom_tienda),
                    "keywords": [k for k in kw_extra if len(k) >= 3]
                })
    except Exception:
        pass
        
    return catalogo

def clasificar_sucursal_guia(destinatario: str, direccion: str = "", ciudad_destino: str = "", catalogo: Optional[List[Dict[str, Any]]] = None) -> Tuple[str, str]:
    """Determina con precisión contable el código y nombre de la sucursal a partir de los datos de la guía y la BD del sistema."""
    dest_norm = normalizar_texto_transporte(destinatario)
    dir_norm = normalizar_texto_transporte(direccion)
    ciu_norm = normalizar_texto_transporte(ciudad_destino)
    full_text = f"{dest_norm} {dir_norm} {ciu_norm}".strip()
    
    cat_uso = catalogo if catalogo is not None else obtener_catalogo_tiendas_dinamico()
    
    # 1. Regla Especial Price Club
    if "PRICE" in dest_norm or "PRICE" in dir_norm:
        for item in [s for s in cat_uso if str(s["codigo"]).startswith("300")]:
            for kw in sorted(item["keywords"], key=len, reverse=True):
                if kw in full_text:
                    return item["codigo"], item["sucursal"]
        if ciu_norm:
            for item in [s for s in cat_uso if str(s["codigo"]).startswith("300")]:
                if any(c in item["sucursal"] for c in ciu_norm.split()):
                    return item["codigo"], item["sucursal"]

    # 2. Búsqueda por palabras clave compuestas ordenadas de mayor a menor longitud
    lista_reglas = []
    for item in cat_uso:
        for kw in item["keywords"]:
            lista_reglas.append((kw, len(kw), item["codigo"], item["sucursal"]))
    lista_reglas.sort(key=lambda x: x[1], reverse=True)
    
    # Evaluar destinatario y dirección
    for kw, _, cod, suc in lista_reglas:
        if kw in dest_norm or kw in dir_norm:
            return cod, suc
            
    # Evaluar texto completo si la palabra clave es larga (>5 letras)
    for kw, l_kw, cod, suc in lista_reglas:
        if l_kw > 5 and kw in full_text:
            return cod, suc

    # 3. Fallback: Si el destinatario indica tienda física / Aeropostale / Fashion Club
    if any(tag in dest_norm for tag in ["AERO", "TIENDA", "LOCAL", "FASHION", "SUCURSAL", "MODA"]):
        for kw, _, cod, suc in lista_reglas:
            if kw in ciu_norm:
                return cod, suc

    return "9999", "OTROS (VENTA WEB / CLIENTES / PROVEEDORES)"

def construir_resumen_sucursal_especifico(df_items: pd.DataFrame, tipo: str = "CONSOLIDADO") -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Construye la tabla de distribución por sucursal para un tipo específico:
    - CONSOLIDADO (Carga + Documentos + Seguro)
    - CARGA (Detalle Carga)
    - DOCUMENTOS (Detalle Documentos)
    - SEGURO (Detalle Seguro)
    """
    if df_items is None or df_items.empty:
        return pd.DataFrame(), {}
        
    df_work = df_items.copy()
    
    cat_dinamico = obtener_catalogo_tiendas_dinamico()
    codigos = []
    sucursales = []
    for _, row in df_work.iterrows():
        cod, suc = clasificar_sucursal_guia(
            destinatario=row.get("DESTINATARIO", ""),
            direccion=row.get("DIRECCION", ""),
            ciudad_destino=row.get("CIUDAD DESTINO", row.get("DESTINO", "")),
            catalogo=cat_dinamico
        )
        codigos.append(cod)
        sucursales.append(suc)
        
    df_work["COD_SUCURSAL"] = codigos
    df_work["NOM_SUCURSAL"] = sucursales
    
    if tipo == "SEGURO":
        tot_general = df_work["COSTO SEGURO"].sum()
        tot_guias_general = len(df_work)
        tot_val_dec_general = df_work["VALOR DECLARADO"].sum() if "VALOR DECLARADO" in df_work.columns else 0.0
        
        agrup = df_work.groupby(["COD_SUCURSAL", "NOM_SUCURSAL"]).agg(
            GUIAS=("GUIA", "count"),
            VAL_DEC=("VALOR DECLARADO", "sum"),
            COSTO_SEG=("COSTO SEGURO", "sum")
        ).reset_index()
        
        df_tiendas = agrup[agrup["COD_SUCURSAL"] != "9999"].sort_values("COSTO_SEG", ascending=False).copy()
        df_otros = agrup[agrup["COD_SUCURSAL"] == "9999"].copy()
        
        filas_dist = []
        n_idx = 1
        for _, r in df_tiendas.iterrows():
            pct = (r["COSTO_SEG"] / tot_general * 100) if tot_general > 0 else 0.0
            filas_dist.append({
                "N.-": str(n_idx),
                "Código": r["COD_SUCURSAL"],
                "SUCURSAL": r["NOM_SUCURSAL"],
                "N° Guías": int(r["GUIAS"]),
                "Valor Declarado": r["VAL_DEC"],
                "Costo Seguro": r["COSTO_SEG"],
                "% Distribución Seguro": pct
            })
            n_idx += 1
            
        if not df_otros.empty:
            r_o = df_otros.iloc[0]
            pct_o = (r_o["COSTO_SEG"] / tot_general * 100) if tot_general > 0 else 0.0
            filas_dist.append({
                "N.-": str(n_idx),
                "Código": "9999",
                "SUCURSAL": "OTROS (VENTA WEB / CLIENTES / PROVEEDORES)",
                "N° Guías": int(r_o["GUIAS"]),
                "Valor Declarado": r_o["VAL_DEC"],
                "Costo Seguro": r_o["COSTO_SEG"],
                "% Distribución Seguro": pct_o
            })
            
        df_tabla = pd.DataFrame(filas_dist)
        fila_total = {
            "N.-": "",
            "Código": "TOTAL",
            "SUCURSAL": "TOTAL GENERAL",
            "N° Guías": int(tot_guias_general),
            "Valor Declarado": tot_val_dec_general,
            "Costo Seguro": tot_general,
            "% Distribución Seguro": 100.0
        }
        df_final = pd.concat([df_tabla, pd.DataFrame([fila_total])], ignore_index=True)
        top5 = df_tabla[df_tabla["Código"] != "9999"].head(5).copy()
        stats = {
            "total_costo": tot_general,
            "total_guias": tot_guias_general,
            "df_top_5": top5,
            "otros_costo": df_otros.iloc[0]["COSTO_SEG"] if not df_otros.empty else 0.0,
            "otros_guias": int(df_otros.iloc[0]["GUIAS"]) if not df_otros.empty else 0,
            "col_codigo": "Código",
            "col_sucursal": "SUCURSAL",
            "col_total": "Costo Seguro"
        }
        return df_final, stats

    elif tipo == "DOCUMENTOS":
        tot_general = df_work["COSTO TOTAL"].sum()
        tot_guias_general = len(df_work)
        
        agrup = df_work.groupby(["COD_SUCURSAL", "NOM_SUCURSAL"]).agg(
            GUIAS=("GUIA", "count"),
            FLETE=("FLETE", "sum"),
            TOTAL=("COSTO TOTAL", "sum")
        ).reset_index()
        
        df_tiendas = agrup[agrup["COD_SUCURSAL"] != "9999"].sort_values("TOTAL", ascending=False).copy()
        df_otros = agrup[agrup["COD_SUCURSAL"] == "9999"].copy()
        
        filas_dist = []
        n_idx = 1
        for _, r in df_tiendas.iterrows():
            pct = (r["TOTAL"] / tot_general * 100) if tot_general > 0 else 0.0
            filas_dist.append({
                "N.-": str(n_idx),
                "Código": r["COD_SUCURSAL"],
                "SUCURSAL": r["NOM_SUCURSAL"],
                "N° Guías": int(r["GUIAS"]),
                "Costo Flete": r["FLETE"],
                "Costo Total Transporte": r["TOTAL"],
                "% Distribución Documentos": pct
            })
            n_idx += 1
            
        if not df_otros.empty:
            r_o = df_otros.iloc[0]
            pct_o = (r_o["TOTAL"] / tot_general * 100) if tot_general > 0 else 0.0
            filas_dist.append({
                "N.-": str(n_idx),
                "Código": "9999",
                "SUCURSAL": "OTROS (VENTA WEB / CLIENTES / PROVEEDORES)",
                "N° Guías": int(r_o["GUIAS"]),
                "Costo Flete": r_o["FLETE"],
                "Costo Total Transporte": r_o["TOTAL"],
                "% Distribución Documentos": pct_o
            })
            
        df_tabla = pd.DataFrame(filas_dist)
        fila_total = {
            "N.-": "",
            "Código": "TOTAL",
            "SUCURSAL": "TOTAL GENERAL",
            "N° Guías": int(tot_guias_general),
            "Costo Flete": df_tabla["Costo Flete"].sum() if not df_tabla.empty else 0.0,
            "Costo Total Transporte": tot_general,
            "% Distribución Documentos": 100.0
        }
        df_final = pd.concat([df_tabla, pd.DataFrame([fila_total])], ignore_index=True)
        top5 = df_tabla[df_tabla["Código"] != "9999"].head(5).copy()
        stats = {
            "total_costo": tot_general,
            "total_guias": tot_guias_general,
            "df_top_5": top5,
            "otros_costo": df_otros.iloc[0]["TOTAL"] if not df_otros.empty else 0.0,
            "otros_guias": int(df_otros.iloc[0]["GUIAS"]) if not df_otros.empty else 0,
            "col_codigo": "Código",
            "col_sucursal": "SUCURSAL",
            "col_total": "Costo Total Transporte"
        }
        return df_final, stats

    else: # CARGA o CONSOLIDADO
        tot_general = df_work["COSTO TOTAL"].sum()
        tot_guias_general = len(df_work)
        
        agrup = df_work.groupby(["COD_SUCURSAL", "NOM_SUCURSAL"]).agg(
            GUIAS=("GUIA", "count"),
            FLETE=("FLETE", "sum"),
            SEGURO=("SEGURO", "sum"),
            TOTAL=("COSTO TOTAL", "sum")
        ).reset_index()
        
        df_tiendas = agrup[agrup["COD_SUCURSAL"] != "9999"].sort_values("TOTAL", ascending=False).copy()
        df_otros = agrup[agrup["COD_SUCURSAL"] == "9999"].copy()
        
        filas_dist = []
        n_idx = 1
        pct_col_name = "% Distribución Logística" if tipo == "CONSOLIDADO" else "% Distribución Carga"
        for _, r in df_tiendas.iterrows():
            pct = (r["TOTAL"] / tot_general * 100) if tot_general > 0 else 0.0
            filas_dist.append({
                "N.-": str(n_idx),
                "Código": r["COD_SUCURSAL"],
                "SUCURSAL": r["NOM_SUCURSAL"],
                "N° Guías": int(r["GUIAS"]),
                "Costo Flete": r["FLETE"],
                "Costo Seguro": r["SEGURO"],
                "Costo Total Transporte": r["TOTAL"],
                pct_col_name: pct
            })
            n_idx += 1
            
        if not df_otros.empty:
            r_o = df_otros.iloc[0]
            pct_o = (r_o["TOTAL"] / tot_general * 100) if tot_general > 0 else 0.0
            filas_dist.append({
                "N.-": str(n_idx),
                "Código": "9999",
                "SUCURSAL": "OTROS (VENTA WEB / CLIENTES / PROVEEDORES)",
                "N° Guías": int(r_o["GUIAS"]),
                "Costo Flete": r_o["FLETE"],
                "Costo Seguro": r_o["SEGURO"],
                "Costo Total Transporte": r_o["TOTAL"],
                pct_col_name: pct_o
            })
            
        df_tabla = pd.DataFrame(filas_dist)
        fila_total = {
            "N.-": "",
            "Código": "TOTAL",
            "SUCURSAL": "TOTAL GENERAL",
            "N° Guías": int(tot_guias_general),
            "Costo Flete": df_tabla["Costo Flete"].sum() if not df_tabla.empty else 0.0,
            "Costo Seguro": df_tabla["Costo Seguro"].sum() if not df_tabla.empty else 0.0,
            "Costo Total Transporte": df_tabla["Costo Total Transporte"].sum() if not df_tabla.empty else 0.0,
            pct_col_name: 100.0
        }
        df_final = pd.concat([df_tabla, pd.DataFrame([fila_total])], ignore_index=True)
        top5 = df_tabla[df_tabla["Código"] != "9999"].head(5).copy()
        stats = {
            "total_costo": tot_general,
            "total_guias": tot_guias_general,
            "df_top_5": top5,
            "otros_costo": df_otros.iloc[0]["TOTAL"] if not df_otros.empty else 0.0,
            "otros_guias": int(df_otros.iloc[0]["GUIAS"]) if not df_otros.empty else 0,
            "col_codigo": "Código",
            "col_sucursal": "SUCURSAL",
            "col_total": "Costo Total Transporte"
        }
        return df_final, stats

def generar_distribucion_logistica(df_det_carga: pd.DataFrame, df_det_doc: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """Alias para compatibilidad."""
    df_fact = pd.concat([df_det_carga, df_det_doc], ignore_index=True) if (not df_det_carga.empty or not df_det_doc.empty) else pd.DataFrame()
    return construir_resumen_sucursal_especifico(df_fact, tipo="CONSOLIDADO")

def procesar_costos_transporte(df_manifiesto: pd.DataFrame, df_carga: pd.DataFrame, df_doc: pd.DataFrame, df_seguro: pd.DataFrame) -> Dict[str, Any]:
    """
    Ejecuta el cruce completo de punta a punta a partir de los 2 archivos de entrada:
    - Cruce Factura Carga + Manifiesto + Seguro
    - Cruce Factura Documentos + Manifiesto
    - Identificación de Guías Anuladas / No Facturadas
    - Cálculo de métricas ejecutivas, ciudad-ciudad y contenido
    - Generación de las 4 Distribuciones Logísticas por Sucursal (Consolidado, Carga, Documentos, Seguro)
    """
    mapa_seguro = {}
    if not df_seguro.empty and "GUIA" in df_seguro.columns:
        col_costo_seg = "SUBTOTAL" if "SUBTOTAL" in df_seguro.columns else ("SEGURO" if "SEGURO" in df_seguro.columns else df_seguro.columns[-1])
        for _, row in df_seguro.iterrows():
            g = row["GUIA"]
            val_seg = parse_float_seguro(row.get(col_costo_seg, 0.0))
            mapa_seguro[g] = val_seg

    mapa_manifiesto = {}
    if not df_manifiesto.empty and "GUIA" in df_manifiesto.columns:
        for _, row in df_manifiesto.iterrows():
            g = row["GUIA"]
            mapa_manifiesto[g] = row.to_dict()

    guias_manifiesto_set = set(df_manifiesto["GUIA"].unique()) if not df_manifiesto.empty else set()
    guias_facturadas_set = set()

    # 3. Procesar Carga Facturada
    detalle_carga_list = []
    if not df_carga.empty:
        for _, row in df_carga.iterrows():
            guia = row["GUIA"]
            guias_facturadas_set.add(guia)
            manif_data = mapa_manifiesto.get(guia, {})
            
            ciu_ori = normalizar_texto_transporte(row.get("CIUDAD ORIGEN", ""))
            if not ciu_ori:
                ciu_ori = normalizar_texto_transporte(manif_data.get("CIUDAD ORIGEN", "IBARRA"))
            ciu_des = normalizar_texto_transporte(row.get("CIUDAD DESTINO", ""))
            if not ciu_des:
                ciu_des = normalizar_texto_transporte(manif_data.get("CIUDAD DESTINO", ""))
                
            tipo_mov = "DESDE CD IBARRA" if ciu_ori == "IBARRA" or "IBARRA" in ciu_ori else "CIUDAD-CIUDAD"
            
            flete = parse_float_seguro(row.get("FLETE", row.get("SUBTOTAL", 0.0)))
            seguro = mapa_seguro.get(guia, 0.0)
            costo_total = flete + seguro
            
            contenido = manif_data.get("CONTENIDO", "")
            producto = manif_data.get("PRODUCTO", "")
            categoria = clasificar_contenido_transporte(contenido, producto)
            
            detalle_carga_list.append({
                "GUIA": guia,
                "FECHA ENVIO": row.get("FECHA REM", row.get("FECHA", manif_data.get("FECHA CREACION", ""))),
                "CIUDAD ORIGEN": ciu_ori,
                "CIUDAD DESTINO": ciu_des,
                "TIPO MOVIMIENTO": tipo_mov,
                "DESTINATARIO": manif_data.get("DESTINATARIO", ""),
                "TELEFONO": manif_data.get("TELEFONO", ""),
                "DIRECCION": manif_data.get("DIRECCION DESTINATARIO", ""),
                "CONTENIDO": contenido,
                "CATEGORIA": categoria,
                "TRAYECTO": row.get("TRAYECTO", ""),
                "PIEZAS": row.get("PIEZAS", 1),
                "PESO": row.get("PESO", manif_data.get("PESO", 0.0)),
                "FLETE": flete,
                "SEGURO": seguro,
                "COSTO TOTAL": costo_total,
                "VALOR DECLARADO": manif_data.get("VALOR DECLARADO", 0.0),
                "ESTADO ENTREGA": manif_data.get("ESTADO", "FACTURADO"),
                "FECHA ENTREGA": manif_data.get("FECHA ENTREGA", ""),
                "RECIBE": manif_data.get("RECIBE", ""),
                "TIPO_SERVICIO": "CARGA"
            })
            
    df_det_carga = pd.DataFrame(detalle_carga_list)

    # 4. Procesar Documentos Facturados
    detalle_doc_list = []
    if not df_doc.empty:
        for _, row in df_doc.iterrows():
            guia = row["GUIA"]
            guias_facturadas_set.add(guia)
            manif_data = mapa_manifiesto.get(guia, {})
            
            ciu_ori = normalizar_texto_transporte(row.get("CIUDAD ORIGEN", ""))
            if not ciu_ori:
                ciu_ori = normalizar_texto_transporte(manif_data.get("CIUDAD ORIGEN", "IBARRA"))
            ciu_des = normalizar_texto_transporte(row.get("CIUDAD DESTINO", ""))
            if not ciu_des:
                ciu_des = normalizar_texto_transporte(manif_data.get("CIUDAD DESTINO", ""))
                
            tipo_mov = "DESDE CD IBARRA" if ciu_ori == "IBARRA" or "IBARRA" in ciu_ori else "CIUDAD-CIUDAD"
            
            flete = parse_float_seguro(row.get("FLETE", row.get("SUBTOTAL", 0.0)))
            seguro = 0.0
            costo_total = flete
            
            contenido = manif_data.get("CONTENIDO", "DOCUMENTOS")
            producto = manif_data.get("PRODUCTO", "DOCUMENTOS_SERV")
            categoria = "DOCUMENTOS"
            
            detalle_doc_list.append({
                "GUIA": guia,
                "FECHA ENVIO": row.get("FECHA REM", row.get("FECHA", manif_data.get("FECHA CREACION", ""))),
                "CIUDAD ORIGEN": ciu_ori,
                "CIUDAD DESTINO": ciu_des,
                "TIPO MOVIMIENTO": tipo_mov,
                "DESTINATARIO": manif_data.get("DESTINATARIO", ""),
                "TELEFONO": manif_data.get("TELEFONO", ""),
                "DIRECCION": manif_data.get("DIRECCION DESTINATARIO", ""),
                "CONTENIDO": contenido,
                "CATEGORIA": categoria,
                "TRAYECTO": row.get("TRAYECTO", ""),
                "PIEZAS": row.get("PIEZAS", 1),
                "PESO": row.get("PESO", 0.0),
                "FLETE": flete,
                "SEGURO": seguro,
                "COSTO TOTAL": costo_total,
                "VALOR DECLARADO": 0.0,
                "ESTADO ENTREGA": manif_data.get("ESTADO", "FACTURADO"),
                "FECHA ENTREGA": manif_data.get("FECHA ENTREGA", ""),
                "RECIBE": manif_data.get("RECIBE", ""),
                "TIPO_SERVICIO": "DOCUMENTOS"
            })
            
    df_det_doc = pd.DataFrame(detalle_doc_list)

    # 5. Procesar Seguro Facturado
    detalle_seg_list = []
    if not df_seguro.empty:
        col_costo_seg = "SUBTOTAL" if "SUBTOTAL" in df_seguro.columns else ("SEGURO" if "SEGURO" in df_seguro.columns else df_seguro.columns[-1])
        for _, row in df_seguro.iterrows():
            guia = row["GUIA"]
            manif_data = mapa_manifiesto.get(guia, {})
            costo_seg = parse_float_seguro(row.get(col_costo_seg, 0.0))
            
            flete_asoc = 0.0
            if not df_det_carga.empty:
                match_c = df_det_carga[df_det_carga["GUIA"] == guia]
                if not match_c.empty:
                    flete_asoc = match_c.iloc[0]["FLETE"]
                    
            detalle_seg_list.append({
                "GUIA": guia,
                "FECHA": row.get("FECHA", manif_data.get("FECHA CREACION", "")),
                "CIUDAD ORIGEN": row.get("ORIGEN", manif_data.get("CIUDAD ORIGEN", "IBARRA")),
                "CIUDAD DESTINO": row.get("DESTINO", manif_data.get("CIUDAD DESTINO", "")),
                "DESTINATARIO": manif_data.get("DESTINATARIO", ""),
                "DIRECCION": manif_data.get("DIRECCION DESTINATARIO", ""),
                "CONTENIDO": manif_data.get("CONTENIDO", ""),
                "CATEGORIA": clasificar_contenido_transporte(manif_data.get("CONTENIDO", ""), manif_data.get("PRODUCTO", "")),
                "ESTADO": manif_data.get("ESTADO", "FACTURADO"),
                "TRAYECTO": row.get("TRAYECTO", ""),
                "VALOR DECLARADO": parse_float_seguro(row.get("VALOR DECLARADO", row.get("VAL. DEC.", manif_data.get("VALOR DECLARADO", 0.0)))),
                "TASA SEGURO": row.get("SEGURO_TASA", row.get("SEGURO", "1%")),
                "COSTO SEGURO": costo_seg,
                "FLETE ASOCIADO": flete_asoc
            })
    df_det_seg = pd.DataFrame(detalle_seg_list)

    # 6. Identificar Guías Anuladas / No Facturadas
    guias_anuladas_set = guias_manifiesto_set - guias_facturadas_set
    guias_anuladas_list = []
    for guia in guias_anuladas_set:
        manif_data = mapa_manifiesto.get(guia, {})
        estado = normalizar_texto_transporte(manif_data.get("ESTADO", ""))
        
        if "NO MOVILIZADO" in estado or "NO RECOLECTADO" in estado or "ANULADO" in estado:
            motivo = "No movilizado por el courier (no se generó recolección) - correctamente excluido de factura"
            nivel_alerta = "NORMAL"
        elif any(e in estado for e in ["ENTREGADO", "CON NOVEDAD", "DEVOLUCION", "ENTREGA"]):
            motivo = "Entregada pero NO incluida en esta factura - revisar con courier (posible corte de periodo o guía omitida)"
            nivel_alerta = "REVISION_URGENTE"
        else:
            motivo = "Requiere revisión manual con el courier"
            nivel_alerta = "ATENCION"
            
        ciu_ori = manif_data.get("CIUDAD ORIGEN", "IBARRA")
        tipo_mov = "DESDE CD IBARRA" if ciu_ori == "IBARRA" else "CIUDAD-CIUDAD"
        
        guias_anuladas_list.append({
            "GUIA": guia,
            "TIPO GUIA": "CARGA" if manif_data.get("PRODUCTO") != "DOCUMENTOS_SERV" else "DOCUMENTO",
            "TIPO MOVIMIENTO": tipo_mov,
            "CIUDAD ORIGEN": ciu_ori,
            "CIUDAD DESTINO": manif_data.get("CIUDAD DESTINO", ""),
            "DESTINATARIO": manif_data.get("DESTINATARIO", ""),
            "CONTENIDO": manif_data.get("CONTENIDO", ""),
            "CATEGORIA": clasificar_contenido_transporte(manif_data.get("CONTENIDO", ""), manif_data.get("PRODUCTO", "")),
            "VALOR DECLARADO": manif_data.get("VALOR DECLARADO", 0.0),
            "ESTADO EN MANIFIESTO": manif_data.get("ESTADO", "NO FACTURADO"),
            "FECHA CREACION": manif_data.get("FECHA CREACION", ""),
            "FECHA ENTREGA": manif_data.get("FECHA ENTREGA", ""),
            "MOTIVO": motivo,
            "NIVEL_ALERTA": nivel_alerta
        })
    df_guias_anuladas = pd.DataFrame(guias_anuladas_list)

    df_todas_facturadas = pd.concat([df_det_carga, df_det_doc], ignore_index=True) if (not df_det_carga.empty or not df_det_doc.empty) else pd.DataFrame()

    total_manifiesto = len(df_manifiesto)
    total_carga = len(df_det_carga)
    total_doc = len(df_det_doc)
    total_seg = len(df_det_seg)
    total_anuladas = len(df_guias_anuladas)
    pct_anulacion = (total_anuladas / total_manifiesto * 100) if total_manifiesto > 0 else 0.0
    
    costo_flete_carga = df_det_carga["FLETE"].sum() if not df_det_carga.empty else 0.0
    costo_flete_doc = df_det_doc["FLETE"].sum() if not df_det_doc.empty else 0.0
    costo_seguro_total = df_det_carga["SEGURO"].sum() if not df_det_carga.empty else (df_det_seg["COSTO SEGURO"].sum() if not df_det_seg.empty else 0.0)
    costo_total_periodo = costo_flete_carga + costo_flete_doc + costo_seguro_total
    costo_promedio_guia = (costo_total_periodo / (total_carga + total_doc)) if (total_carga + total_doc) > 0 else 0.0

    resumen_movimiento = []
    if not df_todas_facturadas.empty:
        for tipo in ["DESDE CD IBARRA", "CIUDAD-CIUDAD"]:
            sub = df_todas_facturadas[df_todas_facturadas["TIPO MOVIMIENTO"] == tipo]
            n_g = len(sub)
            c_tot = sub["COSTO TOTAL"].sum()
            pct_c = (c_tot / costo_total_periodo * 100) if costo_total_periodo > 0 else 0.0
            pct_g = (n_g / len(df_todas_facturadas) * 100) if len(df_todas_facturadas) > 0 else 0.0
            resumen_movimiento.append({
                "TIPO DE MOVIMIENTO": tipo,
                "N° GUIAS": n_g,
                "COSTO TOTAL (USD)": c_tot,
                "% GUIAS": pct_g,
                "% DEL COSTO TOTAL": pct_c
            })
    df_resumen_mov = pd.DataFrame(resumen_movimiento)

    df_top_ciudades = pd.DataFrame()
    if not df_todas_facturadas.empty and "CIUDAD DESTINO" in df_todas_facturadas.columns:
        top_c = df_todas_facturadas.groupby("CIUDAD DESTINO").agg(
            N_GUIAS=("GUIA", "count"),
            COSTO_TOTAL=("COSTO TOTAL", "sum")
        ).reset_index()
        top_c["PCT_COSTO_TOTAL"] = (top_c["COSTO_TOTAL"] / costo_total_periodo * 100) if costo_total_periodo > 0 else 0.0
        df_top_ciudades = top_c.sort_values("COSTO_TOTAL", ascending=False).head(15).copy()

    df_resumen_contenido = pd.DataFrame()
    if not df_todas_facturadas.empty and "CATEGORIA" in df_todas_facturadas.columns:
        cont_c = df_todas_facturadas.groupby("CATEGORIA").agg(
            N_GUIAS=("GUIA", "count"),
            COSTO_TOTAL=("COSTO TOTAL", "sum")
        ).reset_index()
        tot_g = len(df_todas_facturadas)
        cont_c["PCT_GUIAS"] = (cont_c["N_GUIAS"] / tot_g * 100) if tot_g > 0 else 0.0
        cont_c["PCT_COSTO"] = (cont_c["COSTO_TOTAL"] / costo_total_periodo * 100) if costo_total_periodo > 0 else 0.0
        orden_oficial = [
            "DOCUMENTOS", "PUBLICIDAD", "VENTA WEB", "MERCADERÍA", 
            "DEVOLUCIÓN", "ACCESORIOS / PRENDAS", "MOBILIARIO / EQUIPOS", 
            "INSUMOS / ADMINISTRATIVO", "OTROS", "SIN ESPECIFICAR"
        ]
        cont_c["ORDEN"] = cont_c["CATEGORIA"].apply(lambda x: orden_oficial.index(x) if x in orden_oficial else 99)
        df_resumen_contenido = cont_c.sort_values("ORDEN").drop(columns=["ORDEN"]).copy()

    df_pares_rutas = pd.DataFrame()
    if not df_todas_facturadas.empty:
        df_cc_only = df_todas_facturadas[df_todas_facturadas["TIPO MOVIMIENTO"] == "CIUDAD-CIUDAD"]
        if not df_cc_only.empty:
            pares = df_cc_only.groupby(["CIUDAD ORIGEN", "CIUDAD DESTINO"]).agg(
                N_GUIAS=("GUIA", "count"),
                COSTO_TOTAL=("COSTO TOTAL", "sum")
            ).reset_index()
            pares["COSTO PROMEDIO"] = pares["COSTO_TOTAL"] / pares["N_GUIAS"]
            df_pares_rutas = pares.sort_values("COSTO_TOTAL", ascending=False).copy()

    # Generación de las 4 distribuciones logísticas por sucursal
    df_dist_consolidada, stats_consolidada = construir_resumen_sucursal_especifico(df_todas_facturadas, tipo="CONSOLIDADO")
    df_dist_carga, stats_carga = construir_resumen_sucursal_especifico(df_det_carga, tipo="CARGA")
    df_dist_doc, stats_doc = construir_resumen_sucursal_especifico(df_det_doc, tipo="DOCUMENTOS")
    df_dist_seguro, stats_seguro = construir_resumen_sucursal_especifico(df_det_seg, tipo="SEGURO")

    return {
        "kpis": {
            "guias_manifiesto": total_manifiesto,
            "guias_carga": total_carga,
            "guias_doc": total_doc,
            "guias_seguro": total_seg,
            "guias_anuladas": total_anuladas,
            "pct_anulacion": pct_anulacion,
            "costo_flete_carga": costo_flete_carga,
            "costo_flete_doc": costo_flete_doc,
            "costo_seguro_total": costo_seguro_total,
            "costo_total_periodo": costo_total_periodo,
            "costo_promedio_guia": costo_promedio_guia,
        },
        "df_det_carga": df_det_carga,
        "df_det_doc": df_det_doc,
        "df_det_seg": df_det_seg,
        "df_guias_anuladas": df_guias_anuladas,
        "df_resumen_mov": df_resumen_mov,
        "df_top_ciudades": df_top_ciudades,
        "df_resumen_contenido": df_resumen_contenido,
        "df_pares_rutas": df_pares_rutas,
        "df_dist_enriquecida": df_dist_consolidada,
        "df_dist_carga": df_dist_carga,
        "df_dist_doc": df_dist_doc,
        "df_dist_seguro": df_dist_seguro,
        "stats_sucursales": stats_consolidada,
        "stats_carga": stats_carga,
        "stats_doc": stats_doc,
        "stats_seguro": stats_seguro
    }

# ==============================================================================
# 5. GENERADOR OFICIAL DE EXCEL (OPENPYXL CON FÓRMULAS VIVAS Y ESTILOS)
# ==============================================================================

def generar_excel_costos_transporte(datos_cruce: Dict[str, Any], mes_nombre: str = "PERIODO", anio: str = "2026") -> bytes:
    """
    Genera el libro oficial .xlsx con 11 pestañas corporativas:
    1. Distribución Consolidada
    2. Distribución - Carga
    3. Distribución - Documentos
    4. Distribución - Seguro
    5. Resumen Ejecutivo
    6. Analisis Ciudad-Ciudad
    7. Analisis Contenido
    8. Detalle Carga
    9. Detalle Documentos
    10. Detalle Seguro
    11. Guias Anuladas
    """
    wb = openpyxl.Workbook()
    
    NAVY_FILL = PatternFill(start_color="002D62", end_color="002D62", fill_type="solid")
    RED_FILL = PatternFill(start_color="CF0A2C", end_color="CF0A2C", fill_type="solid")
    LIGHT_GRAY_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    ALERT_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    FONT_TITLE = Font(name="Arial", size=13, bold=True, color="002D62")
    FONT_SUBTITLE = Font(name="Arial", size=10, italic=True, color="555555")
    FONT_BOLD = Font(name="Arial", size=10, bold=True)
    FONT_TOTAL = Font(name="Arial", size=11, bold=True, color="991B1B")
    
    BORDER_THIN = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )
    BORDER_TOTAL = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000")
    )

    df_carga = datos_cruce["df_det_carga"]
    df_doc = datos_cruce["df_det_doc"]
    df_seg = datos_cruce["df_det_seg"]
    df_anul = datos_cruce["df_guias_anuladas"]

    def _escribir_hoja_distribucion(sheet_title: str, df_dist_data: pd.DataFrame, header_title_text: str):
        if df_dist_data is None or df_dist_data.empty:
            return
        ws_d = wb.create_sheet(title=sheet_title)
        ws_d["A1"] = f"{header_title_text} - {mes_nombre} {anio}"
        ws_d["A1"].font = FONT_TITLE
        
        headers_d = list(df_dist_data.columns)
        ws_d.append([])
        ws_d.append(headers_d)
        r_d_h = 3
        for c_i in range(1, len(headers_d) + 1):
            cell = ws_d.cell(row=r_d_h, column=c_i)
            cell.fill = NAVY_FILL
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center")
            
        num_d_rows = len(df_dist_data)
        for idx, row in df_dist_data.iterrows():
            r = 4 + idx
            row_vals = []
            for col_name in headers_d:
                val = row.get(col_name, "")
                row_vals.append(val)
            ws_d.append(row_vals)
            
            es_tot_row = (idx == num_d_rows - 1)
            for c_i, col_name in enumerate(headers_d, 1):
                cell = ws_d.cell(row=r, column=c_i)
                if es_tot_row:
                    cell.font = FONT_TOTAL
                    cell.border = BORDER_TOTAL
                    cell.fill = TOTAL_FILL
                else:
                    cell.border = BORDER_THIN
                    if r % 2 == 0:
                        cell.fill = LIGHT_GRAY_FILL
                
                cn = str(col_name).upper()
                if "FLETE" in cn or "SEGURO" in cn or "TOTAL TRANSPORTE" in cn or "VALOR DECLARADO" in cn or "TOTAL" in cn:
                    if "%" not in cn and "GUIAS" not in cn and "PIEZAS" not in cn and "N.-" not in cn:
                        cell.number_format = '$#,##0.00'
                if "%" in cn or "DISTRIBUCION" in cn:
                    if isinstance(cell.value, (int, float)) and cell.value > 1.0:
                        cell.value = cell.value / 100.0
                    cell.number_format = '0.00%'

    # 1. Distribución Consolidada
    _escribir_hoja_distribucion("Distribución Consolidada", datos_cruce.get("df_dist_enriquecida"), "DISTRIBUCIÓN LOGÍSTICA DE COSTOS DE TRANSPORTE POR SUCURSAL")
    # 2. Distribución Carga
    _escribir_hoja_distribucion("Distribución Carga", datos_cruce.get("df_dist_carga"), "DISTRIBUCIÓN LOGÍSTICA - GUÍAS DE CARGA POR SUCURSAL")
    # 3. Distribución Documentos
    _escribir_hoja_distribucion("Distribución Documentos", datos_cruce.get("df_dist_doc"), "DISTRIBUCIÓN LOGÍSTICA - GUÍAS DE DOCUMENTOS POR SUCURSAL")
    # 4. Distribución Seguro
    _escribir_hoja_distribucion("Distribución Seguro", datos_cruce.get("df_dist_seguro"), "DISTRIBUCIÓN LOGÍSTICA - SEGURO CONTRATADO POR SUCURSAL")

    # --------------------------------------------------------------------------
    # PESTAÑA 4: Detalle Carga
    # --------------------------------------------------------------------------
    ws_carga = wb.create_sheet(title="Detalle Carga")
    headers_carga = [
        "GUIA", "FECHA ENVIO", "CIUDAD ORIGEN", "CIUDAD DESTINO", "TIPO MOVIMIENTO", 
        "DESTINATARIO", "TELEFONO", "DIRECCION", "CONTENIDO", "CATEGORIA", 
        "TRAYECTO", "PIEZAS", "PESO", "FLETE", "SEGURO", "COSTO TOTAL", 
        "VALOR DECLARADO", "ESTADO ENTREGA", "FECHA ENTREGA", "RECIBE"
    ]
    ws_carga.append(headers_carga)
    for col_idx in range(1, len(headers_carga) + 1):
        cell = ws_carga.cell(row=1, column=col_idx)
        cell.fill = NAVY_FILL
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    num_rows_carga = len(df_carga)
    for r_idx, row in df_carga.iterrows():
        r = r_idx + 2
        ws_carga.append([
            row["GUIA"], str(row["FECHA ENVIO"])[:10], row["CIUDAD ORIGEN"], row["CIUDAD DESTINO"],
            f'=IF(C{r}="IBARRA","DESDE CD IBARRA","CIUDAD-CIUDAD")',
            row["DESTINATARIO"], str(row["TELEFONO"]), row["DIRECCION"], row["CONTENIDO"], row["CATEGORIA"],
            row["TRAYECTO"], row["PIEZAS"], row["PESO"], row["FLETE"],
            f'=IFERROR(INDEX(\'Detalle Seguro\'!$H:$H,MATCH(A{r},\'Detalle Seguro\'!$A:$A,0)),0)',
            f'=N{r}+O{r}',
            row["VALOR DECLARADO"], row["ESTADO ENTREGA"], str(row["FECHA ENTREGA"])[:10], row["RECIBE"]
        ])
        
        ws_carga.cell(row=r, column=12).number_format = '#,##0'
        ws_carga.cell(row=r, column=13).number_format = '#,##0.00'
        ws_carga.cell(row=r, column=14).number_format = '$#,##0.00'
        ws_carga.cell(row=r, column=15).number_format = '$#,##0.00'
        ws_carga.cell(row=r, column=16).number_format = '$#,##0.00'
        ws_carga.cell(row=r, column=17).number_format = '$#,##0.00'
        
        if r % 2 == 0:
            for c_i in range(1, len(headers_carga) + 1):
                ws_carga.cell(row=r, column=c_i).fill = LIGHT_GRAY_FILL
                
    if num_rows_carga > 0:
        tot_r_c = num_rows_carga + 2
        ws_carga.cell(row=tot_r_c, column=1, value="TOTAL").font = FONT_TOTAL
        ws_carga.cell(row=tot_r_c, column=12, value=f'=SUM(L2:L{tot_r_c-1})').number_format = '#,##0'
        ws_carga.cell(row=tot_r_c, column=13, value=f'=SUM(M2:M{tot_r_c-1})').number_format = '#,##0.00'
        ws_carga.cell(row=tot_r_c, column=14, value=f'=SUM(N2:N{tot_r_c-1})').number_format = '$#,##0.00'
        ws_carga.cell(row=tot_r_c, column=15, value=f'=SUM(O2:O{tot_r_c-1})').number_format = '$#,##0.00'
        ws_carga.cell(row=tot_r_c, column=16, value=f'=SUM(P2:P{tot_r_c-1})').number_format = '$#,##0.00'
        ws_carga.cell(row=tot_r_c, column=17, value=f'=SUM(Q2:Q{tot_r_c-1})').number_format = '$#,##0.00'
        for c_i in range(1, len(headers_carga) + 1):
            cell = ws_carga.cell(row=tot_r_c, column=c_i)
            cell.font = FONT_TOTAL
            cell.border = BORDER_TOTAL
            cell.fill = TOTAL_FILL
            
    ws_carga.freeze_panes = "A2"

    # --------------------------------------------------------------------------
    # PESTAÑA 5: Detalle Documentos
    # --------------------------------------------------------------------------
    ws_doc = wb.create_sheet(title="Detalle Documentos")
    ws_doc.append(headers_carga)
    for col_idx in range(1, len(headers_carga) + 1):
        cell = ws_doc.cell(row=1, column=col_idx)
        cell.fill = NAVY_FILL
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    num_rows_doc = len(df_doc)
    for r_idx, row in df_doc.iterrows():
        r = r_idx + 2
        ws_doc.append([
            row["GUIA"], str(row["FECHA ENVIO"])[:10], row["CIUDAD ORIGEN"], row["CIUDAD DESTINO"],
            f'=IF(C{r}="IBARRA","DESDE CD IBARRA","CIUDAD-CIUDAD")',
            row["DESTINATARIO"], str(row["TELEFONO"]), row["DIRECCION"], row["CONTENIDO"], row["CATEGORIA"],
            row["TRAYECTO"], row["PIEZAS"], row["PESO"], row["FLETE"],
            0.0, f'=N{r}+O{r}',
            0.0, row["ESTADO ENTREGA"], str(row["FECHA ENTREGA"])[:10], row["RECIBE"]
        ])
        ws_doc.cell(row=r, column=12).number_format = '#,##0'
        ws_doc.cell(row=r, column=13).number_format = '#,##0.00'
        ws_doc.cell(row=r, column=14).number_format = '$#,##0.00'
        ws_doc.cell(row=r, column=15).number_format = '$#,##0.00'
        ws_doc.cell(row=r, column=16).number_format = '$#,##0.00'
        ws_doc.cell(row=r, column=17).number_format = '$#,##0.00'
        if r % 2 == 0:
            for c_i in range(1, len(headers_carga) + 1):
                ws_doc.cell(row=r, column=c_i).fill = LIGHT_GRAY_FILL
                
    if num_rows_doc > 0:
        tot_r_d = num_rows_doc + 2
        ws_doc.cell(row=tot_r_d, column=1, value="TOTAL").font = FONT_TOTAL
        ws_doc.cell(row=tot_r_d, column=12, value=f'=SUM(L2:L{tot_r_d-1})').number_format = '#,##0'
        ws_doc.cell(row=tot_r_d, column=14, value=f'=SUM(N2:N{tot_r_d-1})').number_format = '$#,##0.00'
        ws_doc.cell(row=tot_r_d, column=16, value=f'=SUM(P2:P{tot_r_d-1})').number_format = '$#,##0.00'
        for c_i in range(1, len(headers_carga) + 1):
            cell = ws_doc.cell(row=tot_r_d, column=c_i)
            cell.font = FONT_TOTAL
            cell.border = BORDER_TOTAL
            cell.fill = TOTAL_FILL
            
    ws_doc.freeze_panes = "A2"

    # --------------------------------------------------------------------------
    # PESTAÑA 6: Detalle Seguro
    # --------------------------------------------------------------------------
    ws_seg = wb.create_sheet(title="Detalle Seguro")
    headers_seg = [
        "GUIA", "FECHA/HORA POLIZA", "ORIGEN", "DESTINO", "TRAYECTO", 
        "VALOR DECLARADO", "TASA SEGURO", "COSTO SEGURO", "FLETE ASOCIADO", 
        "COSTO TOTAL GUIA", "DESTINATARIO", "CONTENIDO", "CATEGORIA", "ESTADO"
    ]
    ws_seg.append(headers_seg)
    for col_idx in range(1, len(headers_seg) + 1):
        cell = ws_seg.cell(row=1, column=col_idx)
        cell.fill = RED_FILL
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    num_rows_seg = len(df_seg)
    for r_idx, row in df_seg.iterrows():
        r = r_idx + 2
        ws_seg.append([
            row.get("GUIA", ""), str(row.get("FECHA", row.get("FECHA/HORA POLIZA", "")))[:19],
            row.get("ORIGEN", row.get("CIUDAD ORIGEN", "")), row.get("DESTINO", row.get("CIUDAD DESTINO", "")),
            row.get("TRAYECTO", ""), row.get("VALOR DECLARADO", 0.0), row.get("TASA SEGURO", "1%"),
            row.get("COSTO SEGURO", 0.0),
            f'=IFERROR(INDEX(\'Detalle Carga\'!$N:$N,MATCH(A{r},\'Detalle Carga\'!$A:$A,0)),0)',
            f'=H{r}+I{r}',
            row.get("DESTINATARIO", ""), row.get("CONTENIDO", ""), row.get("CATEGORIA", ""), row.get("ESTADO", "")
        ])
        ws_seg.cell(row=r, column=6).number_format = '$#,##0.00'
        ws_seg.cell(row=r, column=8).number_format = '$#,##0.00'
        ws_seg.cell(row=r, column=9).number_format = '$#,##0.00'
        ws_seg.cell(row=r, column=10).number_format = '$#,##0.00'
        if r % 2 == 0:
            for c_i in range(1, len(headers_seg) + 1):
                ws_seg.cell(row=r, column=c_i).fill = LIGHT_GRAY_FILL
                
    if num_rows_seg > 0:
        tot_r_s = num_rows_seg + 2
        ws_seg.cell(row=tot_r_s, column=1, value="TOTAL").font = FONT_TOTAL
        ws_seg.cell(row=tot_r_s, column=6, value=f'=SUM(F2:F{tot_r_s-1})').number_format = '$#,##0.00'
        ws_seg.cell(row=tot_r_s, column=8, value=f'=SUM(H2:H{tot_r_s-1})').number_format = '$#,##0.00'
        ws_seg.cell(row=tot_r_s, column=9, value=f'=SUM(I2:I{tot_r_s-1})').number_format = '$#,##0.00'
        ws_seg.cell(row=tot_r_s, column=10, value=f'=SUM(J2:J{tot_r_s-1})').number_format = '$#,##0.00'
        for c_i in range(1, len(headers_seg) + 1):
            cell = ws_seg.cell(row=tot_r_s, column=c_i)
            cell.font = FONT_TOTAL
            cell.border = BORDER_TOTAL
            cell.fill = TOTAL_FILL
            
    ws_seg.freeze_panes = "A2"

    # --------------------------------------------------------------------------
    # PESTAÑA 7: Guias Anuladas
    # --------------------------------------------------------------------------
    ws_anul = wb.create_sheet(title="Guias Anuladas")
    headers_anul = [
        "GUIA", "TIPO GUIA", "TIPO MOVIMIENTO", "CIUDAD ORIGEN", "CIUDAD DESTINO", 
        "DESTINATARIO", "CONTENIDO", "CATEGORIA", "VALOR DECLARADO", 
        "ESTADO EN MANIFIESTO", "FECHA CREACION", "FECHA ENTREGA", "MOTIVO"
    ]
    ws_anul.append(headers_anul)
    for col_idx in range(1, len(headers_anul) + 1):
        cell = ws_anul.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for r_idx, row in df_anul.iterrows():
        r = r_idx + 2
        ws_anul.append([
            row.get("GUIA", ""), row.get("TIPO GUIA", ""), row.get("TIPO MOVIMIENTO", ""),
            row.get("CIUDAD ORIGEN", ""), row.get("CIUDAD DESTINO", ""),
            row.get("DESTINATARIO", ""), row.get("CONTENIDO", ""), row.get("CATEGORIA", ""), row.get("VALOR DECLARADO", 0.0),
            row.get("ESTADO EN MANIFIESTO", ""), str(row.get("FECHA CREACION", ""))[:10],
            str(row.get("FECHA ENTREGA", ""))[:10], row.get("MOTIVO", "")
        ])
        ws_anul.cell(row=r, column=9).number_format = '$#,##0.00'
        if row.get("NIVEL_ALERTA") == "REVISION_URGENTE":
            for c_i in range(1, len(headers_anul) + 1):
                ws_anul.cell(row=r, column=c_i).fill = ALERT_FILL
                ws_anul.cell(row=r, column=c_i).font = Font(name="Arial", size=10, bold=True, color="991B1B")
        elif r % 2 == 0:
            for c_i in range(1, len(headers_anul) + 1):
                ws_anul.cell(row=r, column=c_i).fill = LIGHT_GRAY_FILL
                
    ws_anul.freeze_panes = "A2"

    # --------------------------------------------------------------------------
    # PESTAÑA 2: Analisis Ciudad-Ciudad
    # --------------------------------------------------------------------------
    ws_cc = wb.create_sheet(title="Analisis Ciudad-Ciudad")
    ws_cc["A1"] = "ANÁLISIS CIUDAD-CIUDAD vs. DESDE CD IBARRA"
    ws_cc["A1"].font = FONT_TITLE
    
    headers_cc1 = ["TIPO DE MOVIMIENTO", "N° GUÍAS", "COSTO TOTAL (USD)", "% DEL COSTO TOTAL"]
    ws_cc.append([])
    ws_cc.append(headers_cc1)
    row_h_cc1 = 3
    for c_i in range(1, 5):
        cell = ws_cc.cell(row=row_h_cc1, column=c_i)
        cell.fill = RED_FILL
        cell.font = FONT_HEADER
        
    max_c = max(2, num_rows_carga + 1)
    max_d = max(2, num_rows_doc + 1)
    
    ws_cc.append([
        "DESDE CD IBARRA",
        f'=COUNTIFS(\'Detalle Carga\'!$E$2:$E${max_c},"DESDE CD IBARRA")+COUNTIFS(\'Detalle Documentos\'!$E$2:$E${max_d},"DESDE CD IBARRA")',
        f'=SUMIFS(\'Detalle Carga\'!$P$2:$P${max_c},\'Detalle Carga\'!$E$2:$E${max_c},"DESDE CD IBARRA")+SUMIFS(\'Detalle Documentos\'!$P$2:$P${max_d},\'Detalle Documentos\'!$E$2:$E${max_d},"DESDE CD IBARRA")',
        '=C4/$C$6'
    ])
    ws_cc.append([
        "CIUDAD-CIUDAD",
        f'=COUNTIFS(\'Detalle Carga\'!$E$2:$E${max_c},"CIUDAD-CIUDAD")+COUNTIFS(\'Detalle Documentos\'!$E$2:$E${max_d},"CIUDAD-CIUDAD")',
        f'=SUMIFS(\'Detalle Carga\'!$P$2:$P${max_c},\'Detalle Carga\'!$E$2:$E${max_c},"CIUDAD-CIUDAD")+SUMIFS(\'Detalle Documentos\'!$P$2:$P${max_d},\'Detalle Documentos\'!$E$2:$E${max_d},"CIUDAD-CIUDAD")',
        '=C5/$C$6'
    ])
    ws_cc.append(["TOTAL", '=SUM(B4:B5)', '=SUM(C4:C5)', '=SUM(D4:D5)'])
    
    for r in range(4, 7):
        ws_cc.cell(row=r, column=2).number_format = '#,##0'
        ws_cc.cell(row=r, column=3).number_format = '$#,##0.00'
        ws_cc.cell(row=r, column=4).number_format = '0.0%'
        if r == 6:
            for c_i in range(1, 5):
                ws_cc.cell(row=r, column=c_i).font = FONT_TOTAL
                ws_cc.cell(row=r, column=c_i).fill = TOTAL_FILL
                ws_cc.cell(row=r, column=c_i).border = BORDER_TOTAL

    ws_cc.append([])
    ws_cc.append(["DETALLE POR PAR CIUDAD ORIGEN -> CIUDAD DESTINO (solo movimientos que NO salen del CD Ibarra)"])
    ws_cc.cell(row=ws_cc.max_row, column=1).font = FONT_TITLE
    headers_rutas = ["CIUDAD ORIGEN", "CIUDAD DESTINO", "N° GUÍAS", "COSTO TOTAL (USD)", "COSTO PROMEDIO / GUÍA (USD)"]
    ws_cc.append(headers_rutas)
    r_rutas_h = ws_cc.max_row
    for c_i in range(1, len(headers_rutas) + 1):
        cell = ws_cc.cell(row=r_rutas_h, column=c_i)
        cell.fill = NAVY_FILL
        cell.font = FONT_HEADER
        
    df_pares = datos_cruce["df_pares_rutas"]
    for r_idx, row in df_pares.iterrows():
        curr_r = ws_cc.max_row + 1
        ws_cc.append([
            row.get("CIUDAD ORIGEN", ""), row.get("CIUDAD DESTINO", ""),
            row.get("N_GUIAS", 0), row.get("COSTO_TOTAL", 0.0), row.get("COSTO PROMEDIO", 0.0)
        ])
        ws_cc.cell(row=curr_r, column=3).number_format = '#,##0'
        ws_cc.cell(row=curr_r, column=4).number_format = '$#,##0.00'
        ws_cc.cell(row=curr_r, column=5).number_format = '$#,##0.00'
        if curr_r % 2 == 0:
            for c_i in range(1, len(headers_rutas) + 1):
                ws_cc.cell(row=curr_r, column=c_i).fill = LIGHT_GRAY_FILL

    # --------------------------------------------------------------------------
    # PESTAÑA 3: Analisis Contenido
    # --------------------------------------------------------------------------
    ws_cont = wb.create_sheet(title="Analisis Contenido")
    ws_cont["A1"] = "ANÁLISIS DE COSTOS POR CATEGORÍA DE CONTENIDO"
    ws_cont["A1"].font = FONT_TITLE
    
    headers_cont = ["CATEGORÍA", "N° GUÍAS", "COSTO TOTAL (USD)", "% GUÍAS", "% COSTO"]
    ws_cont.append([])
    ws_cont.append(headers_cont)
    r_cont_h = 3
    for c_i in range(1, len(headers_cont) + 1):
        cell = ws_cont.cell(row=r_cont_h, column=c_i)
        cell.fill = RED_FILL
        cell.font = FONT_HEADER
        
    categorias_lista = [
        "DOCUMENTOS", "MERCADERÍA", "ACCESORIOS / PRENDAS", "VENTA WEB", 
        "DEVOLUCIÓN", "PUBLICIDAD", "MOBILIARIO / EQUIPOS", "INSUMOS / ADMINISTRATIVO", 
        "SIN ESPECIFICAR", "OTROS"
    ]
    
    start_r_cont = 4
    for idx, cat in enumerate(categorias_lista):
        r = start_r_cont + idx
        ws_cont.append([
            cat,
            f'=COUNTIFS(\'Detalle Carga\'!$J$2:$J${max_c},"{cat}")+COUNTIFS(\'Detalle Documentos\'!$J$2:$J${max_d},"{cat}")',
            f'=SUMIFS(\'Detalle Carga\'!$P$2:$P${max_c},\'Detalle Carga\'!$J$2:$J${max_c},"{cat}")+SUMIFS(\'Detalle Documentos\'!$P$2:$P${max_d},\'Detalle Documentos\'!$J$2:$J${max_d},"{cat}")',
            f'=B{r}/$B${start_r_cont+len(categorias_lista)}',
            f'=C{r}/$C${start_r_cont+len(categorias_lista)}'
        ])
        ws_cont.cell(row=r, column=2).number_format = '#,##0'
        ws_cont.cell(row=r, column=3).number_format = '$#,##0.00'
        ws_cont.cell(row=r, column=4).number_format = '0.0%'
        ws_cont.cell(row=r, column=5).number_format = '0.0%'
        if r % 2 == 0:
            for c_i in range(1, 6):
                ws_cont.cell(row=r, column=c_i).fill = LIGHT_GRAY_FILL
                
    tot_r_cont = start_r_cont + len(categorias_lista)
    ws_cont.append([
        "TOTAL",
        f'=SUM(B4:B{tot_r_cont-1})',
        f'=SUM(C4:C{tot_r_cont-1})',
        f'=SUM(D4:D{tot_r_cont-1})',
        f'=SUM(E4:E{tot_r_cont-1})'
    ])
    ws_cont.cell(row=tot_r_cont, column=2).number_format = '#,##0'
    ws_cont.cell(row=tot_r_cont, column=3).number_format = '$#,##0.00'
    ws_cont.cell(row=tot_r_cont, column=4).number_format = '0.0%'
    ws_cont.cell(row=tot_r_cont, column=5).number_format = '0.0%'
    for c_i in range(1, 6):
        cell = ws_cont.cell(row=tot_r_cont, column=c_i)
        cell.font = FONT_TOTAL
        cell.border = BORDER_TOTAL
        cell.fill = TOTAL_FILL

    # --------------------------------------------------------------------------
    # PESTAÑA 1: Resumen Ejecutivo
    # --------------------------------------------------------------------------
    ws_resumen = wb.create_sheet(title="Resumen Ejecutivo", index=0)
    
    ws_resumen.merge_cells("A1:D1")
    ws_resumen["A1"] = f"RESUMEN EJECUTIVO DE COSTOS DE TRANSPORTE - FASHION CLUB (AEROPOSTALE)"
    ws_resumen["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws_resumen["A1"].fill = NAVY_FILL
    ws_resumen["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws_resumen.merge_cells("A2:D2")
    ws_resumen["A2"] = f"Centro de Distribución Ibarra. Cruce Manifiesto de Recolección vs. Factura del Courier (Carga + Documentos + Seguro). Período: {mes_nombre} {anio}"
    ws_resumen["A2"].font = FONT_SUBTITLE
    ws_resumen["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # 1. INDICADORES DE GESTIÓN (KPI)
    ws_resumen["A4"] = "1. INDICADORES DE GESTION (KPI)"
    ws_resumen["A4"].font = FONT_TITLE
    
    kpis_tabla = [
        ("Guías en el Manifiesto (total del mes)", f'=COUNTA(\'Detalle Carga\'!$A$2:$A${max_c})+COUNTA(\'Detalle Documentos\'!$A$2:$A${max_d})+COUNTA(\'Guias Anuladas\'!$A$2:$A${max(2, len(df_anul)+1)})', '#,##0'),
        ("Guías facturadas - Carga", f'=COUNTA(\'Detalle Carga\'!$A$2:$A${max_c})', '#,##0'),
        ("Guías facturadas - Documentos", f'=COUNTA(\'Detalle Documentos\'!$A$2:$A${max_d})', '#,##0'),
        ("Guías con Seguro contratado", f'=COUNTA(\'Detalle Seguro\'!$A$2:$A${max(2, num_rows_seg+1)})', '#,##0'),
        ("Guías ANULADAS / no facturadas", f'=COUNTA(\'Guias Anuladas\'!$A$2:$A${max(2, len(df_anul)+1)})', '#,##0'),
        ("% de Anulación sobre el Manifiesto", '=B9/B5', '0.0%'),
    ]
    for idx, (label, formula, num_fmt) in enumerate(kpis_tabla):
        r = 5 + idx
        ws_resumen.cell(row=r, column=1, value=label).font = FONT_BOLD
        c_val = ws_resumen.cell(row=r, column=2, value=formula)
        c_val.font = Font(name="Arial", size=11, bold=True, color="991B1B")
        c_val.number_format = num_fmt
        c_val.alignment = Alignment(horizontal="right")
        ws_resumen.cell(row=r, column=1).border = BORDER_THIN
        ws_resumen.cell(row=r, column=2).border = BORDER_THIN

    # 2. COSTOS DEL PERIODO (USD)
    ws_resumen["A12"] = "2. COSTOS DEL PERIODO (USD)"
    ws_resumen["A12"].font = FONT_TITLE
    
    costos_tabla = [
        ("Costo Flete - Carga", f'=\'Detalle Carga\'!N{num_rows_carga+2}' if num_rows_carga>0 else 0.0),
        ("Costo Flete - Documentos", f'=\'Detalle Documentos\'!N{num_rows_doc+2}' if num_rows_doc>0 else 0.0),
        ("Costo Seguro (todas las guías)", f'=\'Detalle Carga\'!O{num_rows_carga+2}' if num_rows_carga>0 else (f'=\'Detalle Seguro\'!H{num_rows_seg+2}' if num_rows_seg>0 else 0.0)),
        ("COSTO TOTAL DEL PERIODO", '=SUM(B13:B15)'),
        ("Costo promedio por guía facturada", '=B16/(B6+B7)'),
    ]
    for idx, (label, formula) in enumerate(costos_tabla):
        r = 13 + idx
        ws_resumen.cell(row=r, column=1, value=label).font = FONT_BOLD
        c_val = ws_resumen.cell(row=r, column=2, value=formula)
        c_val.font = Font(name="Arial", size=11, bold=True, color="991B1B" if r!=16 else "002D62")
        c_val.number_format = '$#,##0.00'
        c_val.alignment = Alignment(horizontal="right")
        if r == 16:
            ws_resumen.cell(row=r, column=1).fill = TOTAL_FILL
            ws_resumen.cell(row=r, column=2).fill = TOTAL_FILL
            ws_resumen.cell(row=r, column=1).border = BORDER_TOTAL
            ws_resumen.cell(row=r, column=2).border = BORDER_TOTAL
        else:
            ws_resumen.cell(row=r, column=1).border = BORDER_THIN
            ws_resumen.cell(row=r, column=2).border = BORDER_THIN

    # 3. CIUDAD-CIUDAD vs. DESDE CD IBARRA
    ws_resumen["A19"] = "3. CIUDAD-CIUDAD vs. DESDE CD IBARRA"
    ws_resumen["A19"].font = FONT_TITLE
    
    headers_res_mov = ["TIPO DE MOVIMIENTO", "N° GUÍAS", "COSTO TOTAL (USD)", "% DEL COSTO TOTAL"]
    ws_resumen.append([])
    ws_resumen.append(headers_res_mov)
    r_h_m = 21
    for c_i in range(1, 5):
        cell = ws_resumen.cell(row=r_h_m, column=c_i)
        cell.fill = RED_FILL
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")
        
    ws_resumen.append(["DESDE CD IBARRA", "='Analisis Ciudad-Ciudad'!B4", "='Analisis Ciudad-Ciudad'!C4", "='Analisis Ciudad-Ciudad'!D4"])
    ws_resumen.append(["CIUDAD-CIUDAD", "='Analisis Ciudad-Ciudad'!B5", "='Analisis Ciudad-Ciudad'!C5", "='Analisis Ciudad-Ciudad'!D5"])
    for r in [22, 23]:
        ws_resumen.cell(row=r, column=1).border = BORDER_THIN
        ws_resumen.cell(row=r, column=2).border = BORDER_THIN
        ws_resumen.cell(row=r, column=3).border = BORDER_THIN
        ws_resumen.cell(row=r, column=4).border = BORDER_THIN
        ws_resumen.cell(row=r, column=2).number_format = '#,##0'
        ws_resumen.cell(row=r, column=3).number_format = '$#,##0.00'
        ws_resumen.cell(row=r, column=4).number_format = '0.0%'

    # 4. TOP 15 CIUDADES DESTINO POR COSTO
    ws_resumen["A25"] = "4. TOP 15 CIUDADES DESTINO POR COSTO"
    ws_resumen["A25"].font = FONT_TITLE
    
    headers_top_c = ["CIUDAD DESTINO", "N° GUÍAS", "COSTO TOTAL (USD)", "% DEL COSTO TOTAL"]
    ws_resumen.append([])
    ws_resumen.append(headers_top_c)
    r_h_top = 27
    for c_i in range(1, 5):
        cell = ws_resumen.cell(row=r_h_top, column=c_i)
        cell.fill = RED_FILL
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")
        
    df_top15 = datos_cruce["df_top_ciudades"]
    for idx, row in df_top15.iterrows():
        curr_r = 28 + idx
        n_g_val = row.get("N_GUIAS", row.get("N° GUIAS", row.get("N° GUÍAS", 0)))
        c_t_val = row.get("COSTO_TOTAL", row.get("COSTO TOTAL (USD)", 0.0))
        pct_t_val = row.get("PCT_COSTO_TOTAL", row.get("% DEL COSTO TOTAL", 0.0))
        ws_resumen.append([row.get("CIUDAD DESTINO", ""), n_g_val, c_t_val, pct_t_val / 100.0])
        ws_resumen.cell(row=curr_r, column=1).border = BORDER_THIN
        ws_resumen.cell(row=curr_r, column=2).border = BORDER_THIN
        ws_resumen.cell(row=curr_r, column=3).border = BORDER_THIN
        ws_resumen.cell(row=curr_r, column=4).border = BORDER_THIN
        ws_resumen.cell(row=curr_r, column=2).number_format = '#,##0'
        ws_resumen.cell(row=curr_r, column=3).number_format = '$#,##0.00'
        ws_resumen.cell(row=curr_r, column=4).number_format = '0.0%'
        if curr_r % 2 == 0:
            for c_i in range(1, 5):
                ws_resumen.cell(row=curr_r, column=c_i).fill = LIGHT_GRAY_FILL

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val:
                    val_str = str(val)
                    if val_str.startswith("="):
                        val_str = "1234567.89"
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def leer_distribucion_original(*args, **kwargs):
    """Alias de compatibilidad previa."""
    return None, 0, ""

def integrar_costos_en_distribucion(*args, **kwargs):
    """Alias de compatibilidad previa."""
    return pd.DataFrame(), {}

