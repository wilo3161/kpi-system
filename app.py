import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import time
import hashlib
import re
import unicodedata
import io
import os
import json
import qrcode
import requests
import imaplib
import email
import tempfile
import requests
from bs4 import BeautifulSoup
import re
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from io import BytesIO
from email.header import decode_header
from typing import Dict, List, Any, Optional
from reportlab.lib.pagesizes import letter, landscape, A5, A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib.colors import HexColor, black, white, grey, whitesmoke, beige
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- CONFIGURACION DE PAGINA ---
st.set_page_config(
    layout="wide",
    page_title="AEROPOSTALE ERP | Control Total",
    page_icon="👔",
    initial_sidebar_state="collapsed",
)

# ==============================================================================
# CONSTANTES GLOBALES (TIENDAS, USUARIOS, ETC.)
# ==============================================================================
TIENDAS_DATA = [
    {"Nombre de Tienda": "Aeropostale - (Cuenca) Mall del Rio", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "CUENCA", "Contacto": "Marco Eras", "Dirección": "Av. Felipe II y Autopista Sur - CC Mall del Rio", "Teléfono": "994570933"},
    {"Nombre de Tienda": "Aeropostale - 6 de Diciembre", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "QUITO", "Contacto": "Micaela Yépez", "Dirección": "Av. 6 de Diciembre y Thomas de Berlanga CC Riocentro UIO", "Teléfono": "987883889"},
    {"Nombre de Tienda": "Aeropostale - Paseo Ambato", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "AMBATO", "Contacto": "Franco Torres", "Dirección": "Manuelita Saenz y Pio Baroja, cerca al parque de las Flores CC Paseo Shopping", "Teléfono": "984951515"},
    {"Nombre de Tienda": "Aeropostale - Ambato", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "AMBATO", "Contacto": "Gabriela Urrutia", "Dirección": "Av. Atahualpa y Victor Hugo CC Mall de Los Andes", "Teléfono": "967239488"},
    {"Nombre de Tienda": "Aeropostale - Babahoyo", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "BABAHOYO", "Contacto": "Yomaira Sellan", "Dirección": "Av.Enrique Ponce Luque CC Paseo Shopping Babahoyo", "Teléfono": "981641355"},
    {"Nombre de Tienda": "Aeropostale - bomboli", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "SANTO DOMINGO", "Contacto": "Josselyn Navarrete", "Dirección": "Via Chone Diagonal a la Universidad Catolica CC Bpmbolí Shopping", "Teléfono": "933906346"},
    {"Nombre de Tienda": "Aeropostale - Bahía de Caráquez", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "BAHIA DE CARAQUEZ", "Contacto": "Nayely Orejuela", "Dirección": "Av. 3 de Noviembre - CC Paseo Shoping Bahía de caraquez", "Teléfono": "981131760"},
    {"Nombre de Tienda": "Aeropostale - Carapungo", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "QUITO", "Contacto": "María José Benalcazar", "Dirección": "Av. Simón Bolívar, Panamericana Norte y calle, Capitán Giovanni Calles - CC", "Teléfono": "997242323"},
    {"Nombre de Tienda": "Aeropostale - CCI", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "QUITO", "Contacto": "Carolina Procel", "Dirección": "Av. Amazonas y Naciones Unidas - CC Iñaquito", "Teléfono": "984048928"},
    {"Nombre de Tienda": "Aeropostale - Ceibos", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "GUAYAQUIL", "Contacto": "Angie Delgado", "Dirección": "Av. Del Bombero y San Eduardo - Riocentro Ceibos", "Teléfono": "999085369"},
    {"Nombre de Tienda": "Aeropostale - Centro Histórico", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "CUENCA", "Contacto": "Renata Sacari", "Dirección": "Av. Simón Bolívar y PadreA guirre Centro Histórico diagonal a a la chocolateri", "Teléfono": "980874018"},
    {"Nombre de Tienda": "Aeropostale - City Mall", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "GUAYAQUIL", "Contacto": "Jordan Guale", "Dirección": "Av. felipe Pezo y Av. Benjamín Carrión CC City Mall", "Teléfono": "962880194"},
    {"Nombre de Tienda": "Aeropostale - Condado Shopping", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "QUITO", "Contacto": "Mateo Recalde", "Dirección": "Av. Mariscal Sucre entre Av. La Prensa Y Jhon F. Kennedy - CC Condado Sh", "Teléfono": "993736447"},
    {"Nombre de Tienda": "Aeropostale - Daule", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "DAULE", "Contacto": "Alisson Ramirez", "Dirección": "Av. Vicente Piedrahita y Coronel Calletano Cestaris- Paseo Shoping Daule", "Teléfono": "978881886"},
    {"Nombre de Tienda": "Aeropostale - Dorado", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "GUAYAQUIL", "Contacto": "Oscar Alvarado", "Dirección": "Av. León Febres Cordero Ribadeneyra - Rio Centro Dorado", "Teléfono": "959098012"},
    {"Nombre de Tienda": "Aeropostale - Durán", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "DURÁN", "Contacto": "Yaritza Córdova", "Dirección": "Av. Boliche Panamericana - Paseo Shoping Durán Junto al Terminal", "Teléfono": "996359344"},
    {"Nombre de Tienda": "Aeropostale - el Coca", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "EL COCA", "Contacto": "Adriana Zurita", "Dirección": "Av. 9 de Octubre y Rio Curaray - Junto a Super Akí el Coca", "Teléfono": "989137928"},
    {"Nombre de Tienda": "Aeropostale - La Plaza", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "IBARRA", "Contacto": "Andrea Andrango", "Dirección": "Av. Mariano Acosta entre Inacio Canelos y Victor Gómez Jurado - CC La Plaz", "Teléfono": "978765143"},
    {"Nombre de Tienda": "Aeropostale - Lago Agrio", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "LAGO AGRIO", "Contacto": "Angie Maldonado", "Dirección": "Av. Quito y Pasaje Brazil - Junto a Super Akí Lago Agrio", "Teléfono": "989893309"},
    {"Nombre de Tienda": "Aeropostale - Machala", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "MACHALA", "Contacto": "Iris carpio", "Dirección": "Av. Paquisha y Vía Machala km. 2 1/2 - CC Paseo Shoping Machala", "Teléfono": "997260162"},
    {"Nombre de Tienda": "Aeropostale - Mall del Pacífico", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "MANTA", "Contacto": "Karina Figueroa", "Dirección": "Av. Malecón y Calle 23 - CC Mall del Pacifico", "Teléfono": "990614279"},
    {"Nombre de Tienda": "Aeropostale - Mall del Rio (Gye)", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "GUAYAQUIL", "Contacto": "Danna Peralta", "Dirección": "Av. Francisco de Orellana y Av. Guillermo Pareja - CC Mall del Rio", "Teléfono": "995609664"},
    {"Nombre de Tienda": "Aeropostale - Mall del Sol", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "GUAYAQUIL", "Contacto": "Kiara Dávalos", "Dirección": "Av. Juan tanca marengo, Carlos Aurelio Rubira Infante 14 NE y Pasaje 1A - C", "Teléfono": "992753549"},
    {"Nombre de Tienda": "Aeropostale - Mall del Sur", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "GUAYAQUIL", "Contacto": "Judith Asunción", "Dirección": "AV 25 de Julio junto al Hospital de IESS - CC Mall del Sur", "Teléfono": "999669429"},
    {"Nombre de Tienda": "Aeropostale - Manta", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "MANTA", "Contacto": "Yenny Alvia", "Dirección": "Av. 4 de noviembre Paseo Shoping Manta", "Teléfono": "995168732"},
    {"Nombre de Tienda": "Aeropostale - Milagro", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "MILAGRO", "Contacto": "Lady Silva", "Dirección": "Av. 12 de Octubre, entre presidente Jerónimo Carrión y presidente Javier Espi", "Teléfono": "985415948"},
    {"Nombre de Tienda": "Aeropostale - Peso Shopping Riobama", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "RIOBAMBA", "Contacto": "María Fernanda Ibarra", "Dirección": "Av. Antonio José de Sucre frente a la Universidad UNACH CC Paseo Shoppin", "Teléfono": "993438844"},
    {"Nombre de Tienda": "Aeropostale - Multiplaza Riobamba", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "RIOBAMBA", "Contacto": "Jennifer Jimenez", "Dirección": "Avenida Lizarzaburu y Agustin Torres - CC Multiplaza Riobamba", "Teléfono": "962636619"},
    {"Nombre de Tienda": "Aeropostale - Pasaje", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "PASAJE", "Contacto": "Jhonny Cun", "Dirección": "Av. Quito entrada a Pasaje y Redondel del León - Junto a Super Aki Pasaje", "Teléfono": "969586186"},
    {"Nombre de Tienda": "Aeropostale - Paseo Ambato", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "AMBATO", "Contacto": "Franco Torres", "Dirección": "Av.Pio Baroja Nesi y Av. Manuelita Saéns Paseo Shoping Ambato", "Teléfono": "984951515"},
    {"Nombre de Tienda": "Aeropostale - Pedernales", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "PEDERNALES", "Contacto": "Mónica Muñoz", "Dirección": "Av. García Moreno y Calle Pedernales - Junto a Aki Pedernales", "Teléfono": "989113061"},
    {"Nombre de Tienda": "Aeropostale - Península", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "PENINSULA", "Contacto": "Kenny Bohorquez", "Dirección": "Av. Carlos Espinosa y Av. Central CC Paseo Shopping La Peninsula", "Teléfono": "997432684"},
    {"Nombre de Tienda": "Aeropostale - Playas", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "PLAYAS", "Contacto": "Steven Ortiz", "Dirección": "Av. General Villamil - Paseo Shoping Playas", "Teléfono": "991871477"},
    {"Nombre de Tienda": "Aeropostale - Portoviejo", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "PORTOVIEJO", "Contacto": "Gissel Loor", "Dirección": "Jorge Washington entre Av. América y E30 CC Paseo Shopping Portoviejo", "Teléfono": "963683962"},
    {"Nombre de Tienda": "Aeropostale - Rio Centro Norte", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "GUAYAQUIL", "Contacto": "Doris Zambrano", "Dirección": "Av. Francisco de Orellana y Urb. Alcance CC Riocentro Norte", "Teléfono": "969705137"},
    {"Nombre de Tienda": "Aeropostale - San Luis", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "QUITO", "Contacto": "Karina Proaño", "Dirección": "Av. General Rumiñahui y Av. San Luis - CC San Luis Shoping", "Teléfono": "991879974"},
    {"Nombre de Tienda": "Aeropostale - Santo Domingo", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "SANTO DOMINGO", "Contacto": "Mateo Fruto", "Dirección": "Av. Abraham Calazacón y Av. Quito - Paseo Shoping Santo Domingo", "Teléfono": "967593039"},
    {"Nombre de Tienda": "Aeropostale - Cayambe", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "CAYAMBE", "Contacto": "Celeste Contreras", "Dirección": "Panamericana norte y camino del sol CC Altos de Cayambe", "Teléfono": "995414136"},
    {"Nombre de Tienda": "Aeropostale Quevedo", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "QUEVEDO", "Contacto": "Dayana León", "Dirección": "Av. Quito, frente a la policia Nacional CC Paseo Shopping Quevedo", "Teléfono": "981398074"},
    {"Nombre de Tienda": "Price Club - Portoviejo", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "PORTOVIEJO", "Contacto": "Dayana Merchan", "Dirección": "Jorge Washington entre Av. América y E30 CC Paseo Shopping Portoviejo", "Teléfono": "959877997"},
    {"Nombre de Tienda": "Price Club - Machala", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "MACHALA", "Contacto": "Yuleysi Delgado", "Dirección": "AV. 25 de Junio CC Oro Plaza", "Teléfono": "988087085"},
    {"Nombre de Tienda": "Price Club - Guayaquil", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "GUAYAQUIL", "Contacto": "Angie Delgado", "Dirección": "Pedro Carbo S/N y Luque junto a almacenes Estuardo Sánchez", "Teléfono": "999085369"},
    {"Nombre de Tienda": "Price Club - Ibarra", "Empresa": "Aeropostale", "Origen": "MATRIZ", "Destino": "IBARRA", "Contacto": "Silvia Urcuango", "Dirección": "Av. Victor Gómez Jurado y Rodrigo Miño junto a la cancha La Bombonera", "Teléfono": "982649058"},
]

USERS_DB = {
    "admin": {"password": hashlib.sha256("wilo3161".encode()).hexdigest(), "role": "Administrador", "name": "Administrador General", "email": "admin@aeropostale.com", "avatar": "👑"},
    "logistica": {"password": hashlib.sha256("log123".encode()).hexdigest(), "role": "Logística", "name": "Coordinador Logístico", "email": "logistica@aeropostale.com", "avatar": "🚚"},
    "ventas": {"password": hashlib.sha256("ven123".encode()).hexdigest(), "role": "Ventas", "name": "Ejecutivo de Ventas", "email": "ventas@aeropostale.com", "avatar": "💼"},
    "bodega": {"password": hashlib.sha256("bod123".encode()).hexdigest(), "role": "Bodega", "name": "Supervisor de Bodega", "email": "bodega@aeropostale.com", "avatar": "📦"},
}

# ... (Resto de constantes PRICE_CLUBS, TIENDAS_REGULARES, COLORS, etc. sin cambios) ...
PRICE_CLUBS = ["Price Club - Portoviejo", "Price Club - Machala", "Price Club - Guayaquil", "Price Club - Ibarra", "Price Club - Cuenca"]
TIENDAS_REGULARES = ['AERO CCI', 'AERO DAULE', 'AERO LAGO AGRIO', 'AERO MALL DEL RIO GYE', 'AERO PLAYAS', 'AEROPOSTALE 6 DE DICIEMBRE', 'AEROPOSTALE BOMBOLI', 'AEROPOSTALE CAYAMBE', 'AEROPOSTALE EL COCA', 'AEROPOSTALE PASAJE', 'AEROPOSTALE PEDERNALES', 'AMBATO', 'BABAHOYO', 'BAHIA DE CARAQUEZ', 'CARAPUNGO', 'CEIBOS', 'CONDADO SHOPPING', 'CUENCA', 'DURAN', 'LA PLAZA SHOPPING', 'MACHALA', 'MAL DEL SUR', 'MALL DEL PACIFICO', 'MALL DEL SOL', 'MANTA', 'MILAGRO', 'MULTIPLAZA RIOBAMBA', 'PASEO AMBATO', 'PENINSULA', 'PORTOVIEJO', 'QUEVEDO', 'RIOBAMBA', 'RIOCENTRO EL DORADO', 'RIOCENTRO NORTE', 'SAN LUIS', 'SANTO DOMINGO']
VENTAS_POR_MAYOR = ["VENTAS POR MAYOR", "MAYORISTA"]
TIENDA_WEB = ["TIENDA WEB", "WEB", "TIENDA MOVIL", "MOVIL"]
FALLAS = ["FALLAS"]

COLORS = {
    'PRICE CLUB': '#0033A0',
    'TIENDAS AEROPOSTALE': '#E4002B',
    'VENTAS POR MAYOR': '#10B981',
    'TIENDA WEB': '#8B5CF6',
    'FALLAS': '#F59E0B',
    'FUNDAS': '#EC4899'
}
GRADIENTS = {
    'PRICE CLUB': 'linear-gradient(135deg, #0033A015, #0033A030)',
    'TIENDAS AEROPOSTALE': 'linear-gradient(135deg, #E4002B15, #E4002B30)',
    'VENTAS POR MAYOR': 'linear-gradient(135deg, #10B98115, #10B98130)',
    'TIENDA WEB': 'linear-gradient(135deg, #8B5CF615, #8B5CF630)',
    'FALLAS': 'linear-gradient(135deg, #F59E0B15, #F59E0B30)',
    'FUNDAS': 'linear-gradient(135deg, #EC489915, #EC489930)'
}
# ==============================================================================
# FUNCIONES AUXILIARES GENERALES (sin cambios)
# ==============================================================================
def hash_password(password: str) -> str: return hashlib.sha256(str(password).encode()).hexdigest()
def extraer_entero(val) -> int:
    if pd.isna(val): return 0
    s = str(val).replace(',', '')
    match = re.search(r'\d+', s)
    return int(match.group()) if match else 0
def normalizar_texto(texto) -> str:
    if pd.isna(texto) or texto == "": return ""
    texto = str(texto)
    try: texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    except Exception: texto = texto.upper()
    texto = re.sub(r"[^A-Za-z0-9\s]", " ", texto.upper())
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto
def procesar_subtotal(valor) -> float:
    if pd.isna(valor): return 0.0
    try:
        if isinstance(valor, (int, float, np.number)): return float(valor)
        valor_str = str(valor).strip()
        valor_str = re.sub(r"[^\d.,-]", "", valor_str)
        if "," in valor_str and "." in valor_str:
            if valor_str.rfind(",") > valor_str.rfind("."): valor_str = valor_str.replace(".", "").replace(",", ".")
            else: valor_str = valor_str.replace(",", "")
        elif "," in valor_str: valor_str = valor_str.replace(",", ".")
        return float(valor_str) if valor_str else 0.0
    except Exception: return 0.0
def obtener_columna_piezas(df: pd.DataFrame) -> Optional[str]:
    posibles = ["PIEZAS", "CANTIDAD", "UNIDADES", "QTY", "CANT", "PZS", "BULTOS"]
    for col in df.columns:
        col_upper = str(col).upper()
        if any(p in col_upper for p in posibles): return col
    return None
def obtener_columna_fecha(df: pd.DataFrame) -> Optional[str]:
    posibles = ["FECHA", "FECHA ING", "FECHA INGRESO", "FECHA CREACION", "FECHA_ING", "FECHA_CREACION"]
    for col in df.columns:
        col_upper = str(col).upper()
        if any(p in col_upper for p in posibles): return col
    return None
def identificar_tipo_tienda(nombre) -> str:
    if pd.isna(nombre) or nombre == "": return "DESCONOCIDO"
    nombre_upper = normalizar_texto(nombre)
    if "JOFRE" in nombre_upper and "SANTANA" in nombre_upper: return "VENTAS AL POR MAYOR"
    nombres_personales = ["ROCIO", "ALEJANDRA", "ANGELICA", "DELGADO", "CRUZ", "LILIANA", "SALAZAR", "RICARDO", "SANCHEZ", "JAZMIN", "ALVARADO", "MELISSA", "CHAVEZ", "KARLA", "SORIANO", "ESTEFANIA", "GUALPA", "MARIA", "JESSICA", "PEREZ", "LOYO"]
    palabras = nombre_upper.split()
    for p in palabras:
        if len(p) > 2 and p in nombres_personales: return "VENTA WEB"
    patrones_fisicas = ["LOCAL", "AEROPOSTALE", "MALL", "PLAZA", "SHOPPING", "CENTRO COMERCIAL", "CC", "C.C", "TIENDA", "SUCURSAL", "PRICE", "CLUB", "DORADO", "CIUDAD", "RIOCENTRO", "PASEO", "PORTAL", "SOL", "CONDADO", "CITY", "CEIBOS", "IBARRA", "MATRIZ", "BODEGA", "FASHION", "GYE", "QUITO", "MACHALA", "PORTOVIEJO", "BABAHOYO", "MANTA", "AMBATO", "CUENCA", "ALMACEN", "PRATI"]
    for patron in patrones_fisicas:
        if patron in nombre_upper: return "TIENDA FÍSICA"
    if len(palabras) >= 3 or any(len(p) > 3 for p in palabras): return "TIENDA FÍSICA"
    return "VENTA WEB"
def to_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer: df.to_excel(writer, index=False)
    return output.getvalue()
def add_back_button(key: str = "back"):
    if st.button("⬅️ Volver", key=key):
        if 'current_page' in st.session_state: st.session_state.current_page = "Inicio"
        st.rerun()

# ==============================================================================
# BASE DE DATOS (MongoDB con fallback local) - Sin cambios
# ==============================================================================
class MockLocalDB:
    def update_password(self, username, new_password_hash):
        db = self._get_db()
        for user in db.get('users', []):
            if user.get('username') == username:
                user['password'] = new_password_hash
                return True
        return False

    def _get_db(self):
        if 'local_database' not in st.session_state:
            st.session_state.local_database = {
                'users': [
                    {'id':1, 'username':'admin', 'password': hash_password('wilo3161'), 'role':'Administrador', 'name':'Administrador General'},
                    {'id':2, 'username':'logistica', 'password': hash_password('log123'), 'role':'Logística', 'name':'Coordinador Logístico'},
                    {'id':3, 'username':'ventas', 'password': hash_password('ven123'), 'role':'Ventas', 'name':'Ejecutivo de Ventas'},
                    {'id':4, 'username':'Andres', 'password': hash_password('Andres145'), 'role':'Bodega', 'name':'Andrés Yépez'},
                    {'id':5, 'username':'Luis', 'password': hash_password('luis230499'), 'role':'Bodega', 'name':'Luis Perugachi'},
                    {'id':6, 'username':'Jessica', 'password': hash_password('bod123'), 'role':'Bodega', 'name':'Jessica Suarez'},
                    {'id':7, 'username':'Diana', 'password': hash_password('bod123'), 'role':'Bodega', 'name':'Diana García'},
                    {'id':8, 'username':'Jhonny', 'password': hash_password('bod123'), 'role':'Bodega', 'name':'Jhonny Villa'},
                    {'id':9, 'username':'Josue', 'password': hash_password('bod123'), 'role':'Bodega', 'name':'Josue Imbacuan'},
                    {'id':10, 'username':'Rocio', 'password': hash_password('temp123'), 'role':'Bodega', 'name':'Rocio Cadena'},
                    {'id':11, 'username':'JhonnyG', 'password': hash_password('temp123'), 'role':'Bodega', 'name':'Jhonny Guadalupe'},
                ],
                'kpis': self._generate_kpis_data()  
            }
        return st.session_state.local_database

    def _generate_kpis_data(self):
        kpis = []
        today = datetime.now()
        for i in range(30):
            date = today - timedelta(days=i)
            kpis.append({"id": i, "fecha": date.strftime("%Y-%m-%d"), "produccion": np.random.randint(800, 1500), "eficiencia": np.random.uniform(85, 98), "alertas": np.random.randint(0, 5), "costos": np.random.uniform(5000, 15000)})
        return kpis

    def query(self, table_name): 
        return self._get_db().get(table_name, [])

    def insert(self, table_name, data):
        db = self._get_db()
        if table_name not in db: 
            db[table_name] = []
        if 'id' not in data: 
            data['id'] = len(db[table_name]) + 1
        db[table_name].append(data)

    def delete(self, table_name, id):
        db = self._get_db()
        if table_name in db: 
            db[table_name] = [item for item in db[table_name] if item.get('id') != id]

    def update(self, table_name, id, update_data):
        db = self._get_db()
        if table_name in db:
            for item in db[table_name]:
                if item.get('id') == id:
                    item.update(update_data)
                    return True
        return False

    def authenticate(self, username, password):
        users = self.query('users')
        input_hash = hash_password(password)
        for user in users:
            if user.get('username') == username and user.get('password') == input_hash: 
                return user
        return None
    
 

try:
    from database import mongo_db as local_db, inicializar_datos_base, MockLocalDBFallback
    inicializar_datos_base()
    print("✅ Usando MongoDB Atlas")
except ImportError as e:
    print(f"⚠️ MongoDB no disponible, usando MockLocalDB local: {e}")
    local_db = MockLocalDB()

# ==============================================================================
# INICIALIZACIÓN DE SESSION STATE - Sin cambios
# ==============================================================================
def initialize_session_state():
    defaults = {
        "current_page": "Inicio",
        "module_data": {},
        "guias_registradas": [],
        "contador_guias": 1000,
        "qr_images": {},
        "logos": {},
        "gastos_datos": {
            "manifesto": None,
            "facturas": None,
            "resultado": None,
            "metricas": None,
            "resumen": None,
            "validacion": None,
            "guias_anuladas": None,
            "procesado": False,
        },
        "clasificacion_data": pd.DataFrame(),
        "clasificacion_loaded": False,
        "kdi_current_data": pd.DataFrame(),
        "kdi_loaded": False,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

    # Nuevos estados para Gestión de Equipo
    if "chat_gemini" not in st.session_state:
        st.session_state.chat_gemini = []
    if "equipo_contactos" not in st.session_state:
        st.session_state.equipo_contactos = {}
    if "prompt_rapido" not in st.session_state:
        st.session_state.prompt_rapido = ""

# ==============================================================================
# ESTILOS CSS - MEJORADO CON TENDENCIAS 2024/2025 (GLASSMORPHISM, NEUMORPHISM, ANIMACIONES FLUIDAS)
# ==============================================================================
def load_css():
    st.markdown("""
    <style>
    /* --- IMPORTACIÓN DE FUENTES PREMIUM --- */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=Space+Grotesk:wght@300;400;500;600&display=swap');
    
    /* --- RESET Y VARIABLES GLOBALES --- */
    * {
        margin: 0;
        padding: 0;
        box-sizing: border-box;
    }

    :root {
        --font-primary: 'Inter', sans-serif;
        --font-display: 'Plus Jakarta Sans', sans-serif;
        --font-mono: 'Space Grotesk', monospace;
        
        --bg-deep: #0b0f19;
        --bg-glass: rgba(20, 30, 45, 0.6);
        --border-glass: rgba(255, 255, 255, 0.08);
        --glow-blue: rgba(56, 189, 248, 0.4);
        --glow-purple: rgba(168, 85, 247, 0.4);
    }

    /* --- FONDO ANIMADO CON GRADIENTE DINÁMICO --- */
    .stApp {
        font-family: var(--font-primary);
        background: radial-gradient(circle at 20% 30%, #1a2335 0%, #0f172a 50%, #020617 100%);
        color: #f1f5f9;
    }

    .main-bg {
        position: fixed;
        top: 0;
        left: 0;
        width: 100vw;
        height: 100vh;
        background: 
            radial-gradient(circle at 15% 40%, rgba(56, 189, 248, 0.15) 0%, transparent 35%),
            radial-gradient(circle at 85% 60%, rgba(168, 85, 247, 0.15) 0%, transparent 40%),
            radial-gradient(circle at 50% 80%, rgba(236, 72, 153, 0.1) 0%, transparent 50%);
        z-index: -2;
        animation: bgPulse 15s ease-in-out infinite alternate;
    }

    @keyframes bgPulse {
        0% { opacity: 0.7; transform: scale(1); }
        100% { opacity: 1; transform: scale(1.05); }
    }

    /* --- PARTÍCULAS FLOTANTES (EFECTO MODERNO) --- */
    .particles {
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        z-index: -1;
        opacity: 0.4;
        background-image: 
            radial-gradient(2px 2px at 20px 30px, #eee, rgba(0,0,0,0)),
            radial-gradient(2px 2px at 40px 70px, #fff, rgba(0,0,0,0)),
            radial-gradient(1px 1px at 90px 40px, #fff, rgba(0,0,0,0)),
            radial-gradient(1px 1px at 130px 80px, #fff, rgba(0,0,0,0));
        background-repeat: repeat;
        background-size: 200px 200px;
        animation: floatParticles 120s linear infinite;
    }

    @keyframes floatParticles {
        0% { transform: translateY(0px); }
        100% { transform: translateY(-2000px); }
    }

    /* --- TÍTULO PRINCIPAL CON GLOW Y TIPOGRAFÍA DISPLAY --- */
    .brand-title {
        font-family: var(--font-display);
        font-size: 4.5rem;
        font-weight: 800;
        letter-spacing: 12px;
        margin-bottom: 10px;
        text-transform: uppercase;
        background: linear-gradient(135deg, #38bdf8 0%, #a78bfa 50%, #f472b6 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-shadow: 0 0 40px rgba(56, 189, 248, 0.3);
        animation: titleGlow 4s ease-in-out infinite alternate;
        transition: all 0.3s ease;
    }

    .brand-title:hover {
        letter-spacing: 16px;
        text-shadow: 0 0 60px var(--glow-purple);
    }

    @keyframes titleGlow {
        0% { filter: drop-shadow(0 0 15px rgba(56, 189, 248, 0.4)); }
        100% { filter: drop-shadow(0 0 30px rgba(168, 85, 247, 0.6)); }
    }

    .brand-subtitle {
        font-family: var(--font-mono);
        color: #94a3b8;
        font-size: 1rem;
        letter-spacing: 8px;
        margin-bottom: 60px;
        text-transform: uppercase;
        font-weight: 400;
        position: relative;
        display: inline-block;
    }

    .brand-subtitle::after {
        content: '';
        position: absolute;
        bottom: -15px;
        left: 50%;
        transform: translateX(-50%);
        width: 120px;
        height: 3px;
        background: linear-gradient(90deg, transparent, #38bdf8, #a78bfa, transparent);
        border-radius: 3px;
    }

    /* --- GRID DE MÓDULOS - GLASSMORPHISM CARDS --- */
    .modules-grid {
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 28px;
        padding: 20px 15px;
        margin-bottom: 50px;
    }

    .module-card {
        background: rgba(15, 25, 40, 0.65);
        backdrop-filter: blur(16px) saturate(180%);
        -webkit-backdrop-filter: blur(16px) saturate(180%);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 32px;
        height: 220px;
        transition: all 0.5s cubic-bezier(0.23, 1, 0.32, 1);
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
        padding: 25px 20px;
        position: relative;
        cursor: pointer;
        overflow: hidden;
        box-shadow: 0 20px 40px -12px rgba(0, 0, 0, 0.5), inset 0 1px 1px rgba(255, 255, 255, 0.05);
    }

    .module-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        background: radial-gradient(600px circle at var(--mouse-x) var(--mouse-y), rgba(255,255,255,0.1), transparent 40%);
        opacity: 0;
        transition: opacity 0.3s;
        z-index: 1;
        pointer-events: none;
    }

    .module-card:hover::before {
        opacity: 1;
    }

    .module-card:hover {
        transform: translateY(-12px) scale(1.02);
        border-color: rgba(56, 189, 248, 0.5);
        box-shadow: 0 30px 50px -12px rgba(0, 0, 0, 0.6), 0 0 0 1px rgba(56, 189, 248, 0.3), 0 0 30px rgba(56, 189, 248, 0.2);
        background: rgba(20, 35, 55, 0.75);
    }

    .card-icon {
        font-size: 3.8rem;
        margin-bottom: 20px;
        transition: all 0.5s cubic-bezier(0.34, 1.56, 0.64, 1);
        filter: drop-shadow(0 5px 10px rgba(0,0,0,0.3));
        z-index: 2;
    }

    .module-card:hover .card-icon {
        transform: scale(1.15) translateY(-8px);
        filter: drop-shadow(0 15px 20px rgba(56, 189, 248, 0.4));
    }

    .card-title {
        font-family: var(--font-display);
        color: white;
        font-size: 1.4rem;
        font-weight: 700;
        letter-spacing: 1.5px;
        text-transform: uppercase;
        text-align: center;
        margin-bottom: 8px;
        z-index: 2;
    }

    .card-description {
        font-family: var(--font-primary);
        color: #cbd5e1;
        font-size: 0.9rem;
        text-align: center;
        opacity: 0.9;
        line-height: 1.6;
        font-weight: 400;
        z-index: 2;
    }

    /* --- HEADERS DE MÓDULO CON ESTILO NEUMÓRFICO --- */
    .module-header {
        background: rgba(20, 30, 45, 0.5);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        padding: 2.5rem 2.5rem;
        border-radius: 48px;
        margin: 20px 0 40px 0;
        border: 1px solid rgba(255, 255, 255, 0.1);
        box-shadow: 0 25px 40px -15px rgba(0, 0, 0, 0.4), inset 0 1px 2px rgba(255, 255, 255, 0.08);
    }

    .header-title {
        font-family: var(--font-display);
        font-size: 2.8rem;
        font-weight: 800;
        color: white;
        margin-bottom: 0.5rem;
        display: flex;
        align-items: center;
        gap: 15px;
    }

    .header-icon {
        font-size: 3.2rem;
        background: linear-gradient(135deg, #38bdf8, #f472b6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        filter: drop-shadow(0 5px 10px rgba(0,0,0,0.3));
    }

    .header-text {
        background: linear-gradient(to right, #f8fafc, #e2e8f0);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }

    /* --- TARJETAS DE ESTADÍSTICAS (KPIs) CON BORDES LUMINOSOS --- */
    .stat-card {
        background: rgba(15, 25, 40, 0.6);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        border-radius: 24px;
        padding: 24px 20px;
        border: 1px solid rgba(255, 255, 255, 0.08);
        transition: all 0.4s ease;
        box-shadow: 0 15px 25px -10px rgba(0, 0, 0, 0.3);
        position: relative;
        overflow: hidden;
    }

    .stat-card::after {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 100%;
        height: 4px;
        background: linear-gradient(90deg, transparent, var(--accent-color), transparent);
        opacity: 0;
        transition: opacity 0.4s;
    }

    .stat-card:hover {
        transform: translateY(-6px);
        border-color: rgba(255, 255, 255, 0.2);
        box-shadow: 0 20px 30px -8px rgba(0, 0, 0, 0.5), 0 0 20px rgba(56, 189, 248, 0.15);
    }

    .stat-card:hover::after {
        opacity: 1;
    }

    .card-blue { --accent-color: #38bdf8; }
    .card-green { --accent-color: #10b981; }
    .card-red { --accent-color: #ef4444; }
    .card-purple { --accent-color: #a78bfa; }

    .stat-icon {
        font-size: 2.2rem;
        margin-bottom: 12px;
        opacity: 0.9;
    }

    .stat-title {
        font-family: var(--font-mono);
        color: #94a3b8;
        font-size: 0.85rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-bottom: 10px;
    }

    .stat-value {
        font-family: var(--font-display);
        color: white;
        font-size: 2.4rem;
        font-weight: 800;
        margin-bottom: 8px;
        line-height: 1.2;
    }

    .stat-change {
        font-size: 0.85rem;
        font-weight: 600;
        padding: 4px 10px;
        border-radius: 30px;
        display: inline-block;
        backdrop-filter: blur(5px);
    }

    .positive { background: rgba(16, 185, 129, 0.2); color: #34d399; border: 1px solid rgba(16, 185, 129, 0.3); }
    .negative { background: rgba(239, 68, 68, 0.2); color: #f87171; border: 1px solid rgba(239, 68, 68, 0.3); }

    /* --- BOTONES CON EFECTO HOLOGRÁFICO --- */
    .stButton > button {
        font-family: var(--font-display) !important;
        background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%) !important;
        color: #f8fafc !important;
        border: 1px solid rgba(56, 189, 248, 0.3) !important;
        border-radius: 40px !important;
        font-weight: 600 !important;
        letter-spacing: 0.5px !important;
        transition: all 0.4s cubic-bezier(0.23, 1, 0.32, 1) !important;
        box-shadow: 0 8px 20px -8px rgba(0, 0, 0, 0.4), inset 0 1px 1px rgba(255, 255, 255, 0.1) !important;
        position: relative;
        overflow: hidden;
    }

    .stButton > button::before {
        content: '';
        position: absolute;
        top: 0;
        left: -100%;
        width: 100%;
        height: 100%;
        background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.1), transparent);
        transition: left 0.6s;
    }

    .stButton > button:hover {
        transform: translateY(-4px) !important;
        border-color: #a78bfa !important;
        box-shadow: 0 15px 25px -8px rgba(56, 189, 248, 0.3), 0 0 0 1px rgba(168, 85, 247, 0.5) !important;
        background: linear-gradient(135deg, #2d3b52 0%, #1e293b 100%) !important;
    }

    .stButton > button:hover::before {
        left: 100%;
    }

    /* --- INPUTS Y CONTROLES DE FORMULARIO --- */
    div[data-baseweb="input"] > div,
    div[data-baseweb="select"] > div {
        background: rgba(15, 25, 40, 0.5) !important;
        backdrop-filter: blur(8px);
        border: 1px solid rgba(255, 255, 255, 0.1) !important;
        border-radius: 16px !important;
        transition: all 0.3s !important;
    }

    div[data-baseweb="input"]:hover > div,
    div[data-baseweb="select"]:hover > div {
        border-color: #38bdf8 !important;
        box-shadow: 0 0 15px rgba(56, 189, 248, 0.2);
    }

    /* --- PESTAÑAS (TABS) ESTILIZADAS --- */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: rgba(0, 0, 0, 0.2);
        padding: 6px;
        border-radius: 50px;
        border: 1px solid rgba(255, 255, 255, 0.05);
    }

    .stTabs [data-baseweb="tab"] {
        font-family: var(--font-display);
        border-radius: 40px !important;
        padding: 10px 24px !important;
        color: #94a3b8 !important;
        font-weight: 500 !important;
        transition: all 0.3s !important;
        background: transparent !important;
        border: none !important;
    }

    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #38bdf8, #a78bfa) !important;
        color: white !important;
        box-shadow: 0 5px 15px rgba(56, 189, 248, 0.3) !important;
    }

    /* --- TABLAS DE DATOS CON HOVER ANIMADO --- */
    .stDataFrame {
        border-radius: 24px !important;
        overflow: hidden !important;
        border: 1px solid rgba(255, 255, 255, 0.08) !important;
        background: rgba(10, 20, 30, 0.4) !important;
        backdrop-filter: blur(8px);
    }

    .stDataFrame [data-testid="stTable"] {
        border-radius: 24px;
    }

    /* --- FOOTER ELEGANTE --- */
    .app-footer {
        text-align: center;
        padding: 40px 20px;
        margin-top: 80px;
        color: #64748b;
        font-size: 0.9rem;
        border-top: 1px solid rgba(255, 255, 255, 0.05);
        background: rgba(0, 0, 0, 0.2);
        backdrop-filter: blur(12px);
        font-family: var(--font-mono);
    }

    /* --- SCROLLBAR CUSTOMIZADA --- */
    ::-webkit-scrollbar { width: 8px; height: 8px; }
    ::-webkit-scrollbar-track { background: rgba(0, 0, 0, 0.2); border-radius: 10px; }
    ::-webkit-scrollbar-thumb { background: #334155; border-radius: 10px; border: 2px solid transparent; background-clip: content-box; }
    ::-webkit-scrollbar-thumb:hover { background: #475569; }

    /* --- RESPONSIVE --- */
    @media (max-width: 1200px) { .modules-grid { grid-template-columns: repeat(2, 1fr); } }
    @media (max-width: 768px) { 
        .modules-grid { grid-template-columns: 1fr; } 
        .brand-title { font-size: 2.8rem; letter-spacing: 6px; }
        .header-title { font-size: 2rem; }
    }

    /* --- LOGIN SCREEN - VERSIÓN PREMIUM --- */
    .login-container {
        max-width: 420px;
        margin: 8vh auto;
        padding: 50px 40px;
        background: rgba(15, 25, 40, 0.5);
        backdrop-filter: blur(25px);
        -webkit-backdrop-filter: blur(25px);
        border: 1px solid rgba(255, 255, 255, 0.15);
        border-radius: 48px;
        box-shadow: 0 40px 60px -20px rgba(0, 0, 0, 0.6), inset 0 1px 1px rgba(255, 255, 255, 0.1);
        text-align: center;
    }

    .login-brand .main {
        font-family: var(--font-display);
        font-size: 2.8rem;
        font-weight: 800;
        letter-spacing: 8px;
        background: linear-gradient(to right, #38bdf8, #a78bfa, #f472b6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 5px;
    }

    .login-brand .sub {
        font-family: var(--font-mono);
        font-size: 0.8rem;
        color: #94a3b8;
        letter-spacing: 6px;
        text-transform: uppercase;
        margin-bottom: 30px;
    }

    .login-title {
        font-family: var(--font-display);
        font-size: 1.2rem;
        font-weight: 600;
        color: #e2e8f0;
        margin-bottom: 30px;
        letter-spacing: 2px;
    }

    </style>
    <div class="main-bg"></div>
    <div class="particles"></div>
    """, unsafe_allow_html=True)

# ==============================================================================
# AUTENTICACIÓN Y NAVEGACIÓN (LÓGICA SIN CAMBIOS MAYORES)
# ==============================================================================
def check_password():
    if "authenticated" in st.session_state and st.session_state.authenticated:
        return True
    
    # Se carga CSS específico para el login (ya incluido en load_css, pero aseguramos)
    load_css()
    
    st.markdown("""
    <style>
    @keyframes gradientBg { 0% { background-position: 0% 50%; } 50% { background-position: 100% 50%; } 100% { background-position: 0% 50%; } }
    .stApp { background: linear-gradient(-45deg, #020617, #0f172a, #1e1b4b, #0f172a); background-size: 400% 400%; animation: gradientBg 15s ease infinite; }
    </style>
    """, unsafe_allow_html=True)
    st.markdown("""
<style>
.login-container {
    text-align: center !important;
}
.login-brand .main, .login-brand .sub, .login-title {
    text-align: center !important;
}
/* Asegurar que los inputs no hereden centrado extraño */
div[data-baseweb="input"] {
    text-align: left !important;
}
</style>
""", unsafe_allow_html=True)
    
    col_left, col_center, col_right = st.columns([1, 2, 1])
    with col_center:
        st.markdown('<div class="login-container">', unsafe_allow_html=True)
        st.markdown('<div class="login-brand"><div class="main">AEROPOSTALE</div><div class="sub">ERP CONTROL TOTAL</div></div>', unsafe_allow_html=True)
        st.markdown('<div class="login-title">INICIAR SESIÓN</div>', unsafe_allow_html=True)
        
        username = st.text_input("Usuario", placeholder="Nombre de usuario...", key="login_user", label_visibility="collapsed")
        password = st.text_input("Contraseña", placeholder="Contraseña segura...", type="password", key="login_pass", label_visibility="collapsed")
        remember = st.checkbox("Recordarme", key="remember_me")
        
        login_btn = st.button("Ingresar al Sistema →", use_container_width=True, type="primary")
        st.markdown('<div style="margin-top: 30px; font-size: 0.8rem; color: #64748b;"> | Secure Connection</div>', unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)
        
        if login_btn:
            try:
                user_data = local_db.authenticate(username, password)
                if user_data:
                    st.session_state.authenticated = True
                    st.session_state.username = username
                    st.session_state.role = user_data.get("role", "Usuario")
                    st.session_state.user_name = user_data.get("name", username)
                    if remember:
                        st.session_state.remember_username = username
                    st.rerun()
            except Exception:
                pass
            
            #if username in USERS_DB:
                #if USERS_DB[username]["password"] == hash_password(password):
                    #st.session_state.authenticated = True
                    #st.session_state.username = username
                    #st.session_state.role = USERS_DB[username]["role"]
                    #st.session_state.user_name = USERS_DB[username]["name"]
                    #if remember:
                        #st.session_state.remember_username = username
                    #st.rerun()
                #else:
                    #st.error("❌ Contraseña incorrecta")
            #else:
               # st.error("❌ Usuario no existe")
    #return False

def show_header():
    col1, col2, col3 = st.columns([1, 3, 1])
    with col1:
        if st.button("🏠 Inicio", use_container_width=True):
            st.session_state.current_page = "Inicio"
            st.rerun()
    with col2:
        st.markdown(f"<div style='text-align: center; color: #CBD5E1; font-family: var(--font-mono);'><strong>{st.session_state.user_name}</strong> | {st.session_state.role} | {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>", unsafe_allow_html=True)
    with col3:
        if st.button("🚪 Salir", use_container_width=True):
            for key in ["authenticated", "username", "role", "user_name"]:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()
    st.markdown("---")

def navigate_to_module(module_key):
    st.session_state.current_page = module_key
    st.rerun()

def create_module_card(icon, title, description, module_key):
    card_html = f"""
    <div class="module-card" onclick="window.location.href='?page={module_key}'">
        <div class="card-icon">{icon}</div>
        <div class="card-title">{title}</div>
        <div class="card-description">{description}</div>
    </div>
    """
    st.markdown(card_html, unsafe_allow_html=True)
    if st.button(f"Acceder a {title}", key=f"btn_{module_key}", use_container_width=True):
        navigate_to_module(module_key)

def show_module_header(title_with_icon, subtitle):
    icon = title_with_icon[0] if title_with_icon else ""
    title_text = title_with_icon[1:].strip() if title_with_icon else ""
    st.markdown(f"""
    <div class="module-header">
        <h1 class="header-title"><span class="header-icon">{icon}</span> <span class="header-text">{title_text}</span></h1>
        <p style="color: #CBD5E1; font-size: 1.1rem; font-family: var(--font-primary);">{subtitle}</p>
    </div>
    """, unsafe_allow_html=True)

def show_main_page():
    load_css()
    st.markdown("""
<div style="text-align: center; width: 100%;">
    <div class="brand-title">AEROPOSTALE</div>
    <div class="brand-subtitle">Centro de Distribucion Ecuador | ERP</div>
</div>
""", unsafe_allow_html=True)
    
    all_modules = [
        {"icon": "📊", "title": "Dashboard KPIs", "description": "Dashboard en tiempo real con metricas operativas", "key": "dashboard_kpis"},
        {"icon": "💰", "title": "Reconciliacion", "description": "Conciliacion financiera y analisis de facturas", "key": "reconciliacion_v8"},
        {"icon": "📧", "title": "Auditoria de Correos", "description": "Analisis inteligente de novedades por email", "key": "auditoria_correos"},
        {"icon": "📦", "title": "Dashboard Logistico", "description": "Control de transferencias y distribucion", "key": "dashboard_logistico"},
        {"icon": "👥", "title": "Gestion de Equipo", "description": "Administracion del personal del centro", "key": "gestion_equipo"},
        {"icon": "🚚", "title": "Generar Guias", "description": "Sistema de envios con seguimiento QR", "key": "generar_guias"},
        {"icon": "📋", "title": "Control de Inventario", "description": "Gestion de stock en tiempo real", "key": "control_inventario"},
        {"icon": "📈", "title": "Reportes Avanzados", "description": "Analisis y estadisticas ejecutivas", "key": "reportes_avanzados"},
        {"icon": "⚙️", "title": "Configuracion", "description": "Personalizacion del sistema ERP", "key": "configuracion"},
    ]
    
    role = st.session_state.role
    if role == "Bodega":
        modules = [m for m in all_modules if m["key"] in ["generar_guias", "control_inventario"]]
    else:
        modules = all_modules
        
    cols = st.columns(3)
    for idx, module in enumerate(modules):
        with cols[idx % 3]:
            create_module_card(module["icon"], module["title"], module["description"], module["key"])
    
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown('<div class="app-footer"><p><strong>Sistema ERP </strong> • Desarrollado por Wilson Pérez • Logistica & Sistemas</p><p style="font-size: 0.85rem; color: #94A3B8;">© 2026 AEROPOSTALE Ecuador • Todos los derechos reservados</p></div>', unsafe_allow_html=True)

# ==============================================================================
# MÓDULO: DASHBOARD KPIs (Sin cambios en lógica, solo estética CSS aplicada)
# ==============================================================================
def show_dashboard_kpis():
    add_back_button(key="back_kpis")
    show_module_header("📊 Dashboard de KPIs", "Metricas en tiempo real del Centro de Distribucion")
    st.markdown('<div class="module-content">', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    with col1: fecha_inicio = st.date_input("📅 Fecha Inicio", datetime.now() - timedelta(days=30))
    with col2: fecha_fin = st.date_input("📅 Fecha Fin", datetime.now())
    with col3: st.selectbox("📈 Tipo de Metrica", ["Produccion", "Eficiencia", "Costos", "Alertas"])
    
    kpis_data = local_db.query("kpis")
    df_kpis = pd.DataFrame(kpis_data)
    
    if not df_kpis.empty:
        df_kpis["fecha"] = pd.to_datetime(df_kpis["fecha"])
        mask = (df_kpis["fecha"].dt.date >= fecha_inicio) & (df_kpis["fecha"].dt.date <= fecha_fin)
        df_filtered = df_kpis[mask]
        
        if not df_filtered.empty:
            col_k1, col_k2, col_k3, col_k4 = st.columns(4)
            with col_k1:
                prod_prom = df_filtered["produccion"].mean()
                st.markdown(f"<div class='stat-card card-blue'><div class='stat-icon'>🏭</div><div class='stat-title'>Produccion Promedio</div><div class='stat-value'>{prod_prom:,.0f}</div></div>", unsafe_allow_html=True)
            with col_k2:
                efic_prom = df_filtered["eficiencia"].mean()
                st.markdown(f"<div class='stat-card card-green'><div class='stat-icon'>⚡</div><div class='stat-title'>Eficiencia</div><div class='stat-value'>{efic_prom:.1f}%</div></div>", unsafe_allow_html=True)
            with col_k3:
                alert_total = df_filtered["alertas"].sum()
                st.markdown(f"<div class='stat-card card-red'><div class='stat-icon'>🚨</div><div class='stat-title'>Alertas Totales</div><div class='stat-value'>{alert_total}</div></div>", unsafe_allow_html=True)
            with col_k4:
                costo_prom = df_filtered["costos"].mean()
                st.markdown(f"<div class='stat-card card-purple'><div class='stat-icon'>💰</div><div class='stat-title'>Costo Promedio</div><div class='stat-value'>${costo_prom:,.0f}</div></div>", unsafe_allow_html=True)
            
            fig = px.line(df_filtered, x="fecha", y="produccion", title="Produccion Diaria", line_shape="spline")
            fig.update_traces(line=dict(color="#38bdf8", width=3))
            fig.update_layout(template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig, use_container_width=True)
            
            col_ch1, col_ch2 = st.columns(2)
            with col_ch1:
                fig2 = px.bar(df_filtered.tail(7), x=df_filtered.tail(7)["fecha"].dt.strftime("%a"), y="eficiencia", title="Eficiencia Semanal", color="eficiencia", color_continuous_scale="Viridis")
                fig2.update_layout(template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig2, use_container_width=True)
            with col_ch2:
                fig3 = px.scatter(df_filtered, x="produccion", y="costos", title="Relacion Produccion vs Costos", color="alertas", size="eficiencia", hover_data=["fecha"])
                fig3.update_layout(template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig3, use_container_width=True)
        else:
            st.warning("No hay datos para el rango de fechas seleccionado.")
    else:
        st.info("Cargando datos de KPIs...")
    st.markdown("</div>", unsafe_allow_html=True)

# ==============================================================================
# FUNCIONES AUXILIARES PARA RECONCILIACIÓN (CARGA Y PROCESAMIENTO) - SIN CAMBIOS
# ==============================================================================
def cargar_archivo_local(uploaded_file, nombre):
    try:
        if uploaded_file.name.endswith((".xlsx", ".xls")):
            excel_file = pd.ExcelFile(uploaded_file)
            hojas = excel_file.sheet_names
            hoja_seleccionada = None
            for hoja in hojas:
                temp_df = pd.read_excel(uploaded_file, sheet_name=hoja, nrows=5)
                if not temp_df.empty and len(temp_df.columns) > 1:
                    hoja_seleccionada = hoja
                    break
            if hoja_seleccionada is None: hoja_seleccionada = hojas[0]
            st.sidebar.info(f"Hoja seleccionada automáticamente: {hoja_seleccionada}")
            df = pd.read_excel(uploaded_file, sheet_name=hoja_seleccionada, header=0)
            df.columns = [str(col).strip().replace('\n', ' ').replace('\r', '') for col in df.columns]
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [' '.join(str(c).strip() for c in col if str(c) != 'nan').strip() for col in df.columns.values]
            st.sidebar.success(f"✓ {nombre}: {len(df):,} filas de Excel")
        elif uploaded_file.name.endswith(".csv"):
            encodings = ["utf-8", "latin-1", "cp1252", "ISO-8859-1"]
            df = None
            for encoding in encodings:
                try:
                    df = pd.read_csv(uploaded_file, encoding=encoding)
                    st.sidebar.success(f"✓ {nombre}: {len(df):,} filas de CSV ({encoding})")
                    break
                except UnicodeDecodeError:
                    if encoding == encodings[-1]: raise
                    continue
        else:
            st.sidebar.error(f"✗ Formato no soportado: {uploaded_file.name}")
            return None
        df = df.dropna(axis=1, how="all")
        df = df.dropna(how="all")
        if df.empty:
            st.sidebar.error(f"✗ {nombre}: Archivo vacío o sin datos válidos")
            return None
        return df
    except Exception as e:
        st.sidebar.error(f"✗ Error al cargar {nombre}: {str(e)}")
        return None

def procesar_gastos_reconciliacion(manifesto, facturas, config):
    # ... (Mantener TODO el código original de esta función, es muy extenso y funciona correctamente) ...
    # Por brevedad se omite la reescritura completa aquí, pero se debe pegar el código original sin modificar.
    # Asegúrate de que el código original de 'procesar_gastos_reconciliacion' esté aquí.
    try:
        # 1. PREPARAR MANIFIESTO
        st.info("📦 Procesando manifiesto...")
        columnas_manifesto = manifesto.columns.tolist()
        col_guia_m = config["guia_m"]
        if col_guia_m not in columnas_manifesto:
            for col in columnas_manifesto:
                if "GUIA" in str(col).upper() or "GUÍA" in str(col).upper():
                    col_guia_m = col
                    break
            if col_guia_m not in columnas_manifesto:
                raise ValueError(f"No se encontró columna de guía en el manifiesto. Columnas: {columnas_manifesto}")
        col_subtotal_m = config.get("subtotal_m", "")
        if not col_subtotal_m or col_subtotal_m not in columnas_manifesto:
            for col in columnas_manifesto:
                if "SUBT" in str(col).upper() or "TOTAL" in str(col).upper() or "VALOR" in str(col).upper():
                    col_subtotal_m = col
                    break
            if not col_subtotal_m or col_subtotal_m not in columnas_manifesto:
                col_subtotal_m = columnas_manifesto[-1]
        col_ciudad_m = config.get("ciudad_destino", "")
        if not col_ciudad_m or col_ciudad_m not in columnas_manifesto:
            for col in columnas_manifesto:
                if "CIUDAD" in str(col).upper() or "DES" in str(col).upper() or "DESTINO" in str(col).upper():
                    col_ciudad_m = col
                    break
            if not col_ciudad_m or col_ciudad_m not in columnas_manifesto:
                col_ciudad_m = "CIUDAD"
                manifesto[col_ciudad_m] = "DESCONOCIDA"
        col_piezas_m = obtener_columna_piezas(manifesto)
        if col_piezas_m:
            st.info(f"✓ Columna de piezas detectada: {col_piezas_m}")
        else:
            st.warning("⚠ No se encontró columna de número de piezas. Se usará valor por defecto de 1 por guía.")
            manifesto["PIEZAS"] = 1
            col_piezas_m = "PIEZAS"
        col_fecha_m = obtener_columna_fecha(manifesto)
        if col_fecha_m:
            st.info(f"✓ Columna de fecha detectada: {col_fecha_m}")
        col_destinatario_m = None
        posibles_destinatario = ["DESTINATARIO", "CONSIGNATARIO", "CLIENTE", "NOMBRE", "RAZON SOCIAL"]
        for col in posibles_destinatario:
            if col in columnas_manifesto:
                col_destinatario_m = col
                break
        if not col_destinatario_m:
            for col in columnas_manifesto:
                col_upper = str(col).upper()
                if any(p in col_upper for p in ["DEST", "CONSIG", "CLIEN", "NOMB", "RAZON"]):
                    col_destinatario_m = col
                    break
        if not col_destinatario_m:
            manifesto["DESTINATARIO_MANIFIESTO"] = "TIENDA " + manifesto[col_ciudad_m].astype(str)
            col_destinatario_m = "DESTINATARIO_MANIFIESTO"
        otras_columnas_importantes = []
        for col in manifesto.columns:
            col_upper = str(col).upper()
            if any(p in col_upper for p in ["FECHA", "ORIGEN", "SERVICIO", "TRANSPORTE", "PESO", "FLETE"]):
                otras_columnas_importantes.append(col)
        columnas_manifiesto = [col_guia_m, col_subtotal_m, col_ciudad_m, col_destinatario_m, col_piezas_m]
        if col_fecha_m:
            columnas_manifiesto.append(col_fecha_m)
        columnas_manifiesto += otras_columnas_importantes
        columnas_manifiesto = list(dict.fromkeys(columnas_manifiesto))
        df_m = manifesto[columnas_manifiesto].copy()
        df_m["GUIA"] = df_m[col_guia_m].astype(str).str.strip()
        df_m["SUBTOTAL_MANIFIESTO"] = df_m[col_subtotal_m].apply(procesar_subtotal)
        df_m["CIUDAD_MANIFIESTO"] = df_m[col_ciudad_m].fillna("DESCONOCIDA").astype(str)
        df_m["DESTINATARIO_MANIFIESTO"] = df_m[col_destinatario_m].fillna("DESTINATARIO DESCONOCIDO").astype(str)
        df_m["PIEZAS_MANIFIESTO"] = pd.to_numeric(df_m[col_piezas_m], errors="coerce").fillna(1)
        if col_fecha_m:
            try:
                df_m["FECHA_MANIFIESTO"] = pd.to_datetime(df_m[col_fecha_m], errors="coerce")
            except:
                df_m["FECHA_MANIFIESTO"] = df_m[col_fecha_m].astype(str)
        df_m["GUIA_LIMPIA"] = df_m["GUIA"].str.upper()
        st.success(f"✓ Manifiesto procesado: {len(df_m):,} guías, {df_m['PIEZAS_MANIFIESTO'].sum():,.0f} piezas totales")
    except Exception as e:
        st.error(f"Error al procesar manifiesto: {str(e)}")
        st.error(f"Columnas disponibles en manifiesto: {manifesto.columns.tolist()}")
        raise

    try:
        st.info("🧾 Procesando facturas...")
        col_guia_f = config.get("guia_f", "")
        col_subtotal_f = config.get("subtotal", "")
        if isinstance(col_guia_f, list):
            col_guia_f = col_guia_f[0] if col_guia_f else ""
        if isinstance(col_subtotal_f, list):
            col_subtotal_f = col_subtotal_f[0] if col_subtotal_f else ""
        facturas.columns = [str(c).strip().replace('\n', ' ').replace('\r', '') for c in facturas.columns]
        if isinstance(facturas.columns, pd.MultiIndex):
            facturas.columns = [' '.join(str(x).strip() for x in col if str(x) != 'nan').strip() for col in facturas.columns.values]
        columnas_facturas = facturas.columns.tolist()
        if not col_guia_f or col_guia_f not in columnas_facturas:
            for col in columnas_facturas:
                if "GUIA" in str(col).upper() or "GUÍA" in str(col).upper():
                    col_guia_f = col
                    break
            if not col_guia_f or col_guia_f not in columnas_facturas:
                raise ValueError(f"No se encontró columna de guía en las facturas. Columnas: {columnas_facturas}")
        if not col_subtotal_f or col_subtotal_f not in columnas_facturas:
            for col in columnas_facturas:
                if any(x in str(col).upper() for x in ["SUBTOTAL", "TOTAL", "IMPORTE", "VALOR"]):
                    col_subtotal_f = col
                    break
            if not col_subtotal_f or col_subtotal_f not in columnas_facturas:
                col_subtotal_f = columnas_facturas[-1]
        df_f = facturas[[col_guia_f, col_subtotal_f]].copy()
        if isinstance(df_f.columns, pd.MultiIndex):
            df_f.columns = ['_'.join(str(c).strip() for c in col if str(c) != 'nan').strip() for col in df_f.columns.values]
            col_guia_f = df_f.columns[0]
            col_subtotal_f = df_f.columns[1]
        df_f["GUIA_FACTURA"] = df_f.iloc[:, 0].astype(str).str.strip()
        df_f["SUBTOTAL_FACTURA"] = df_f.iloc[:, 1].apply(procesar_subtotal)
        df_f["GUIA_LIMPIA"] = df_f["GUIA_FACTURA"].str.upper()
        st.success(f"✓ Facturas procesadas: {len(df_f):,} registros")
    except Exception as e:
        st.error(f"Error al procesar facturas: {str(e)}")
        st.error(f"Columnas disponibles en facturas: {facturas.columns.tolist()}")
        raise

    try:
        st.info("🔗 Uniendo datos por guía...")
        df_completo = pd.merge(
            df_m,
            df_f[["GUIA_LIMPIA", "SUBTOTAL_FACTURA"]],
            on="GUIA_LIMPIA",
            how="left",
        )
        df_completo["DESTINATARIO"] = df_completo["DESTINATARIO_MANIFIESTO"]
        df_completo["CIUDAD"] = df_completo["CIUDAD_MANIFIESTO"]
        df_completo["PIEZAS"] = df_completo["PIEZAS_MANIFIESTO"]
        df_completo["ESTADO"] = df_completo["SUBTOTAL_FACTURA"].apply(
            lambda x: "FACTURADA" if pd.notna(x) and float(x) > 0 else "ANULADA"
        )
        df_completo["SUBTOTAL"] = df_completo["SUBTOTAL_FACTURA"].fillna(0)
        df_completo["DIFERENCIA"] = df_completo["SUBTOTAL_MANIFIESTO"] - df_completo["SUBTOTAL"]
        df_completo["TIPO"] = df_completo["DESTINATARIO"].apply(identificar_tipo_tienda)
        df_completo["NOMBRE_NORMALIZADO"] = df_completo["DESTINATARIO"].apply(normalizar_texto)
        def crear_grupo(fila):
            tipo = fila["TIPO"]
            nombre = fila["NOMBRE_NORMALIZADO"]
            ciudad = normalizar_texto(fila["CIUDAD"])
            if tipo == "VENTA WEB":
                palabras = nombre.split()
                if len(palabras) >= 2:
                    return f"VENTA WEB - {palabras[0]} {palabras[1]}"
                else:
                    return f"VENTA WEB - {nombre}"
            elif tipo == "VENTAS AL POR MAYOR":
                return "VENTAS AL POR MAYOR - JOFRE SANTANA"
            elif tipo == "TIENDA FÍSICA":
                grupo_ciudad = f"{ciudad} - " if ciudad != "DESCONOCIDA" else ""
                palabras = nombre.split()
                if len(palabras) > 0:
                    nombre_grupo = " ".join(palabras[: min(3, len(palabras))])
                    return f"{grupo_ciudad}{nombre_grupo}"
                else:
                    return f"{grupo_ciudad}TIENDA"
            else:
                return f"DESCONOCIDO - {nombre[:20]}"
        df_completo["GRUPO"] = df_completo.apply(crear_grupo, axis=1)
        guias_facturadas = df_completo[df_completo["ESTADO"] == "FACTURADA"].shape[0]
        guias_anuladas = df_completo[df_completo["ESTADO"] == "ANULADA"].shape[0]
        total_piezas = df_completo["PIEZAS"].sum()
        guias_anuladas_df = df_completo[df_completo["ESTADO"] == "ANULADA"].copy()
        st.success(f"✓ Datos unidos: {len(df_completo):,} registros")
        st.info(f"  • Guías facturadas: {guias_facturadas:,}")
        st.info(f"  • Guías anuladas: {guias_anuladas:,}")
        st.info(f"  • Piezas totales: {total_piezas:,}")
    except Exception as e:
        st.error(f"Error al unir datos: {str(e)}")
        raise

    try:
        st.info("📊 Calculando métricas por grupo...")
        df_facturadas = df_completo[df_completo["ESTADO"] == "FACTURADA"]
        metricas = (
            df_facturadas.groupby("GRUPO")
            .agg(
                GUIAS=("GUIA_LIMPIA", "count"),
                PIEZAS=("PIEZAS", "sum"),
                SUBTOTAL=("SUBTOTAL", "sum"),
                SUBTOTAL_MANIFIESTO=("SUBTOTAL_MANIFIESTO", "sum"),
                DIFERENCIA=("DIFERENCIA", "sum"),
                DESTINATARIOS=("DESTINATARIO", lambda x: ", ".join(sorted(set(str(d) for d in x if pd.notna(d) and str(d) != ""))[:5])),
                CIUDADES=("CIUDAD", lambda x: ", ".join(sorted(set(str(c) for c in x if pd.notna(c) and str(c) != ""))[:3])),
                TIPO=("TIPO", lambda x: x.mode()[0] if not x.mode().empty else "DESCONOCIDO"),
            )
            .reset_index()
        )
        total_general = metricas["SUBTOTAL"].sum()
        if total_general > 0:
            metricas["PORCENTAJE"] = (metricas["SUBTOTAL"] / total_general * 100).round(2)
            metricas["PROMEDIO_POR_PIEZA"] = (metricas["SUBTOTAL"] / metricas["PIEZAS"]).round(2)
        else:
            metricas["PORCENTAJE"] = 0.0
            metricas["PROMEDIO_POR_PIEZA"] = 0.0
        metricas["PIEZAS_POR_GUIA"] = (metricas["PIEZAS"] / metricas["GUIAS"]).round(2)
        metricas = metricas.sort_values("SUBTOTAL", ascending=False)
        st.success(f"✓ Métricas calculadas: {len(metricas):,} grupos identificados")
    except Exception as e:
        st.error(f"Error al calcular métricas: {str(e)}")
        raise

    try:
        st.info("📋 Generando resumen por tipo...")
        resumen = (
            df_facturadas.groupby("TIPO")
            .agg(
                TIENDAS=("GRUPO", "nunique"),
                GUIAS=("GUIA_LIMPIA", "count"),
                PIEZAS=("PIEZAS", "sum"),
                SUBTOTAL=("SUBTOTAL", "sum"),
            )
            .reset_index()
        )
        if total_general > 0:
            resumen["PORCENTAJE"] = (resumen["SUBTOTAL"] / total_general * 100).round(2)
        else:
            resumen["PORCENTAJE"] = 0.0
        resumen = resumen.sort_values("SUBTOTAL", ascending=False)
        st.success("✓ Resumen por tipo generado")
    except Exception as e:
        st.error(f"Error al calcular resumen: {str(e)}")
        raise

    try:
        st.info("✅ Realizando validación...")
        total_manifiesto = df_completo["SUBTOTAL_MANIFIESTO"].sum()
        total_facturas = df_completo["SUBTOTAL"].sum()
        total_piezas_manifesto = df_completo["PIEZAS"].sum()
        validacion = {
            "total_manifiesto": total_manifiesto,
            "total_facturas": total_facturas,
            "diferencia": abs(total_manifiesto - total_facturas),
            "porcentaje": (abs(total_manifiesto - total_facturas) / total_manifiesto * 100) if total_manifiesto > 0 else 0,
            "coincide": abs(total_manifiesto - total_facturas) < 0.01,
            "guias_procesadas": len(df_completo),
            "guias_facturadas": guias_facturadas,
            "guias_anuladas": guias_anuladas,
            "piezas_totales": total_piezas_manifesto,
            "grupos_identificados": len(metricas),
            "porcentaje_facturadas": (guias_facturadas / len(df_completo) * 100) if len(df_completo) > 0 else 0,
            "porcentaje_anuladas": (guias_anuladas / len(df_completo) * 100) if len(df_completo) > 0 else 0,
        }
        st.success("✓ Validación completada")
    except Exception as e:
        st.error(f"Error al realizar validación: {str(e)}")
        raise

    return df_completo, metricas, resumen, validacion, guias_anuladas_df

# ==============================================================================
# MÓDULO: RECONCILIACIÓN (GESTIÓN DE GASTOS POR TIENDA) - COMPLETO SIN CAMBIOS
# ==============================================================================
def show_reconciliacion_v8():
    """Modulo de reconciliacion financiera y gestion de gastos por tienda"""
    add_back_button(key="back_reconciliacion")
    show_module_header(
        "💰 Gestión de Gastos por Tienda",
        "Conciliación financiera y análisis de facturas"
    )
    st.markdown('<div class="module-content">', unsafe_allow_html=True)

    if 'gastos_datos' not in st.session_state:
        st.session_state.gastos_datos = {
            'manifesto': None, 'facturas': None, 'resultado': None,
            'metricas': None, 'resumen': None, 'validacion': None,
            'guias_anuladas': None, 'procesado': False
        }

    # --- Funciones de exportación (mantener originales) ---
    def generar_excel_con_formato_exacto(metricas_filt, resultado, guias_anuladas, manifesto_original, filtros_aplicados=None):
        # ... (código original de generación de excel) ...
        try:
            output = BytesIO()
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Reporte"
            hoja1_data = metricas_filt[['GRUPO', 'SUBTOTAL']].copy().sort_values('GRUPO')
            ws1.append(["", ""])
            ws1.append(["", ""])
            ws1.append(["Etiquetas de fila", "Suma de SUBTOTAL"])
            for _, row in hoja1_data.iterrows():
                ws1.append([row['GRUPO'], row['SUBTOTAL']])
            ws1.append(["Total general", hoja1_data['SUBTOTAL'].sum()])
            for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for cell in ws1[3]:
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for row in range(4, ws1.max_row + 1):
                ws1.cell(row=row, column=2).number_format = '#,##0.00'
            ws1.column_dimensions['A'].width = 50
            ws1.column_dimensions['B'].width = 20

            ws2 = wb.create_sheet(title="Tiendas")
            columnas = ["GRUPO", "GUIAS", "PIEZAS", "SUBTOTAL", "DESTINATARIOS", "CIUDADES", "TIPO", "PORCENTAJE", "PROMEDIO_POR_PIEZA", "PIEZAS_POR_GUIA"]
            ws2.append(columnas)
            for _, row in metricas_filt.iterrows():
                ws2.append([row['GRUPO'], int(row['GUIAS']), int(row['PIEZAS']), row['SUBTOTAL'],
                            row['DESTINATARIOS'], row['CIUDADES'], row['TIPO'], row['PORCENTAJE'],
                            row['PROMEDIO_POR_PIEZA'], row['PIEZAS_POR_GUIA']])
            ws2.append(["" for _ in range(len(columnas))])
            ultima_fila_datos = ws2.max_row - 1
            total_row = ["" for _ in range(len(columnas))]
            total_row[0] = "Total general"
            total_row[3] = f"=SUBTOTAL(109,D2:D{ultima_fila_datos})"
            ws2.append(total_row)
            for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=len(columnas)):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    if cell.column in [2,3,8,9,10]:
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
            for cell in ws2[1]:
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for row in range(2, ws2.max_row + 1):
                ws2.cell(row=row, column=4).number_format = '#,##0.00'
                ws2.cell(row=row, column=8).number_format = '0.00'
                ws2.cell(row=row, column=9).number_format = '0.00'
                ws2.cell(row=row, column=10).number_format = '0.00'
            for cell in ws2[ws2.max_row]:
                cell.font = Font(bold=True)
                if cell.column == 4:
                    cell.number_format = '#,##0.00'
            anchos = [40,10,10,15,50,20,20,15,20,20]
            for i, ancho in enumerate(anchos,1):
                ws2.column_dimensions[get_column_letter(i)].width = ancho

            if not guias_anuladas.empty:
                ws3 = wb.create_sheet(title="Guias Anuladas")
                posibles_columnas = ['FECHA_MANIFIESTO', 'GUIA', 'ORIGEN', 'GUIA2', 'DESTINATARIO_MANIFIESTO',
                                     'SERVICIO', 'TRANSPORTE', 'PIEZAS_MANIFIESTO', 'PESO', 'FLETE',
                                     'SUBTOTAL_MANIFIESTO', 'CIUDAD_MANIFIESTO', 'ESTADO']
                columnas_anuladas = [col for col in posibles_columnas if col in guias_anuladas.columns]
                mapeo_nombres = {'FECHA_MANIFIESTO':'FECHA', 'DESTINATARIO_MANIFIESTO':'DESTINATARIO',
                                 'PIEZAS_MANIFIESTO':'NUMERO DE PIEZAS', 'SUBTOTAL_MANIFIESTO':'SUBTOTAL',
                                 'CIUDAD_MANIFIESTO':'CIUDAD'}
                columnas_anuladas = [mapeo_nombres.get(col, col) for col in columnas_anuladas]
                ws3.append(columnas_anuladas)
                for _, row in guias_anuladas.iterrows():
                    fila_data = []
                    for col_original in posibles_columnas:
                        if col_original in guias_anuladas.columns:
                            valor = row[col_original]
                            if col_original == 'FECHA_MANIFIESTO' and pd.notna(valor) and hasattr(valor, 'strftime'):
                                fila_data.append(valor.strftime('%d/%m/%Y %H:%M'))
                            elif col_original == 'PIEZAS_MANIFIESTO':
                                fila_data.append(int(valor) if pd.notna(valor) else 0)
                            elif col_original == 'SUBTOTAL_MANIFIESTO':
                                fila_data.append(float(valor) if pd.notna(valor) else 0.0)
                            else:
                                fila_data.append(valor if pd.notna(valor) else '')
                    ws3.append(fila_data)
                ws3.append(["" for _ in range(len(columnas_anuladas))])
                total_row = ["" for _ in range(len(columnas_anuladas))]
                total_row[0] = "Total guías anuladas"
                if 'NUMERO DE PIEZAS' in columnas_anuladas:
                    idx_piezas = columnas_anuladas.index('NUMERO DE PIEZAS') + 1
                    total_row[idx_piezas-1] = f"=SUBTOTAL(109,{get_column_letter(idx_piezas)}2:{get_column_letter(idx_piezas)}{ws3.max_row-1})"
                if 'SUBTOTAL' in columnas_anuladas:
                    idx_subtotal = columnas_anuladas.index('SUBTOTAL') + 1
                    total_row[idx_subtotal-1] = f"=SUBTOTAL(109,{get_column_letter(idx_subtotal)}2:{get_column_letter(idx_subtotal)}{ws3.max_row-1})"
                ws3.append(total_row)
                for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=len(columnas_anuladas)):
                    for cell in row:
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for cell in ws3[1]:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
                for row in range(2, ws3.max_row + 1):
                    if 'NUMERO DE PIEZAS' in columnas_anuladas:
                        col_idx = columnas_anuladas.index('NUMERO DE PIEZAS') + 1
                        ws3.cell(row=row, column=col_idx).number_format = '0'
                    if 'SUBTOTAL' in columnas_anuladas:
                        col_idx = columnas_anuladas.index('SUBTOTAL') + 1
                        ws3.cell(row=row, column=col_idx).number_format = '#,##0.00'
                anchos_anuladas = [18,15,15,20,15,40,15,15,15,10,10,15,20,15]
                for i, ancho in enumerate(anchos_anuladas[:len(columnas_anuladas)]):
                    ws3.column_dimensions[get_column_letter(i+1)].width = ancho

            ws4 = wb.create_sheet(title="Detalle")
            columnas_detalle = []
            if 'FECHA_MANIFIESTO' in resultado.columns:
                columnas_detalle.append('FECHA_MANIFIESTO')
            columnas_detalle += ['GUIA', 'ESTADO', 'GRUPO', 'DESTINATARIO', 'CIUDAD', 'PIEZAS',
                                 'SUBTOTAL_MANIFIESTO', 'SUBTOTAL', 'DIFERENCIA', 'TIPO']
            columnas_detalle = [col for col in columnas_detalle if col in resultado.columns]
            detalle_df = resultado[columnas_detalle].copy()
            mapeo_detalle = {'FECHA_MANIFIESTO':'FECHA', 'GUIA':'GUIA', 'ESTADO':'ESTADO', 'GRUPO':'GRUPO',
                             'DESTINATARIO':'DESTINATARIO', 'CIUDAD':'CIUDAD', 'PIEZAS':'PIEZAS',
                             'SUBTOTAL_MANIFIESTO':'SUBTOTAL MANIFIESTO', 'SUBTOTAL':'SUBTOTAL FACTURA',
                             'DIFERENCIA':'DIFERENCIA', 'TIPO':'TIPO'}
            detalle_df = detalle_df.rename(columns={k:v for k,v in mapeo_detalle.items() if k in detalle_df.columns})
            ws4.append(list(detalle_df.columns))
            for _, row in detalle_df.iterrows():
                fila_data = []
                for col in detalle_df.columns:
                    valor = row[col]
                    if col == 'FECHA' and pd.notna(valor) and hasattr(valor, 'strftime'):
                        fila_data.append(valor.strftime('%d/%m/%Y %H:%M'))
                    elif col in ['PIEZAS']:
                        fila_data.append(int(valor) if pd.notna(valor) else 0)
                    elif col in ['SUBTOTAL MANIFIESTO', 'SUBTOTAL FACTURA', 'DIFERENCIA']:
                        fila_data.append(float(valor) if pd.notna(valor) else 0.0)
                    else:
                        fila_data.append(valor if pd.notna(valor) else '')
                ws4.append(fila_data)
            for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, min_col=1, max_col=len(detalle_df.columns)):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for cell in ws4[1]:
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for row in range(2, ws4.max_row + 1):
                for col_idx, col_name in enumerate(detalle_df.columns, 1):
                    if col_name in ['PIEZAS']:
                        ws4.cell(row=row, column=col_idx).number_format = '0'
                    elif col_name in ['SUBTOTAL MANIFIESTO', 'SUBTOTAL FACTURA', 'DIFERENCIA']:
                        ws4.cell(row=row, column=col_idx).number_format = '#,##0.00'
            anchos_detalle = [20,15,12,40,30,20,10,15,15,15,20]
            for i, ancho in enumerate(anchos_detalle[:len(detalle_df.columns)]):
                ws4.column_dimensions[get_column_letter(i+1)].width = ancho

            wb.save(output)
            output.seek(0)
            return output
        except Exception as e:
            st.error(f"Error al generar Excel con formato: {str(e)}")
            return None

    def generar_pdf_reporte(metricas, resumen, validacion, filtros_aplicados=None):
        # ... (código original de generación de PDF) ...
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                pdf_path = tmp_file.name
            doc = SimpleDocTemplate(pdf_path, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=12, alignment=1)
            subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Heading2'], fontSize=12, spaceAfter=6)
            normal_style = styles['Normal']
            elements.append(Paragraph("REPORTE EJECUTIVO - GESTIÓN DE GASTOS POR TIENDA", title_style))
            elements.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("MÉTRICAS PRINCIPALES", subtitle_style))
            metricas_data = [
                ["Total Facturado", f"${validacion['total_facturas']:,.2f}"],
                ["Total Manifiesto", f"${validacion['total_manifiesto']:,.2f}"],
                ["Diferencia", f"${validacion['diferencia']:,.2f} ({validacion['porcentaje']:.2f}%)"],
                ["Guías Procesadas", f"{validacion['guias_procesadas']:,}"],
                ["Guías Facturadas", f"{validacion['guias_facturadas']:,} ({validacion['porcentaje_facturadas']:.1f}%)"],
                ["Guías Anuladas", f"{validacion['guias_anuladas']:,} ({validacion['porcentaje_anuladas']:.1f}%)"],
                ["Piezas Totales", f"{validacion['piezas_totales']:,}"],
                ["Grupos Identificados", f"{validacion['grupos_identificados']:,}"]
            ]
            metricas_table = Table(metricas_data, colWidths=[200, 150])
            metricas_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), grey), ('TEXTCOLOR', (0,0), (-1,0), whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 10), ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), beige), ('GRID', (0,0), (-1,-1), 1, black)
            ]))
            elements.append(metricas_table)
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("RESUMEN POR TIPO DE TIENDA", subtitle_style))
            if not resumen.empty:
                resumen_data = [["TIPO","TIENDAS","GUÍAS","PIEZAS","SUBTOTAL","%"]]
                for _, row in resumen.iterrows():
                    resumen_data.append([row['TIPO'], str(int(row['TIENDAS'])), str(int(row['GUIAS'])),
                                        str(int(row['PIEZAS'])), f"${row['SUBTOTAL']:,.2f}", f"{row['PORCENTAJE']:.2f}%"])
                resumen_table = Table(resumen_data, colWidths=[120,80,80,80,100,80])
                resumen_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), grey), ('TEXTCOLOR', (0,0), (-1,0), whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,-1), 9), ('BOTTOMPADDING', (0,0), (-1,0), 12),
                    ('BACKGROUND', (0,1), (-1,-1), beige), ('GRID', (0,0), (-1,-1), 1, black)
                ]))
                elements.append(resumen_table)
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("TOP 15 GRUPOS POR GASTO", subtitle_style))
            if not metricas.empty:
                top_15 = metricas.head(15)
                grupos_data = [["GRUPO","GUÍAS","PIEZAS","SUBTOTAL","%","PROM/PIEZA"]]
                for _, row in top_15.iterrows():
                    grupos_data.append([row['GRUPO'][:30] + "..." if len(row['GRUPO'])>30 else row['GRUPO'],
                                        str(int(row['GUIAS'])), str(int(row['PIEZAS'])),
                                        f"${row['SUBTOTAL']:,.2f}", f"{row['PORCENTAJE']:.2f}%",
                                        f"${row['PROMEDIO_POR_PIEZA']:,.2f}"])
                grupos_table = Table(grupos_data, colWidths=[150,60,60,90,60,80])
                grupos_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), grey), ('TEXTCOLOR', (0,0), (-1,0), whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,-1), 8), ('BOTTOMPADDING', (0,0), (-1,0), 12),
                    ('BACKGROUND', (0,1), (-1,-1), beige), ('GRID', (0,0), (-1,-1), 1, black)
                ]))
                elements.append(grupos_table)
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("ANÁLISIS EJECUTIVO", subtitle_style))
            analisis_text = f"""
            <b>Validación:</b> {"✅ COINCIDENCIA EXACTA" if validacion['coincide'] else "⚠ CON DIFERENCIAS"}<br/>
            <b>Facturación:</b> {validacion['porcentaje_facturadas']:.1f}% de guías facturadas ({validacion['guias_facturadas']:,} guías)<br/>
            <b>Anulaciones:</b> {validacion['porcentaje_anuladas']:.1f}% de guías anuladas ({validacion['guias_anuladas']:,} guías)<br/>
            <b>Distribución:</b> {resumen.iloc[0]['TIPO'] if not resumen.empty else 'N/A'} representa el {resumen.iloc[0]['PORCENTAJE'] if not resumen.empty else 0:.1f}% del total facturado<br/>
            <b>Eficiencia:</b> Promedio de {validacion['piezas_totales']/validacion['guias_procesadas']:.1f} piezas por guía<br/>
            <b>Recomendación:</b> {"Revisar guías anuladas para optimizar facturación" if validacion['guias_anuladas'] > 0 else "Proceso de facturación eficiente"}
            """
            elements.append(Paragraph(analisis_text, normal_style))
            doc.build(elements)
            return pdf_path
        except Exception as e:
            st.error(f"Error al generar PDF: {str(e)}")
            return None

    # --- Sidebar para carga de archivos ---
    with st.sidebar:
        st.header("📁 Carga de Archivos")
        st.markdown("**Formatos soportados:** Excel (.xlsx, .xls) y CSV")
        uploaded_manifesto = st.file_uploader("Manifiesto (con DESTINATARIO y PIEZAS)", type=["csv", "xlsx", "xls"], key="manifesto_upload")
        uploaded_facturas = st.file_uploader("Facturas (solo GUÍA y VALOR)", type=["csv", "xlsx", "xls"], key="facturas_upload")
        if uploaded_manifesto and uploaded_facturas:
            if st.button("📥 Cargar Archivos", type="primary", use_container_width=True):
                with st.spinner("Cargando archivos..."):
                    manifesto = cargar_archivo_local(uploaded_manifesto, "Manifiesto")
                    facturas = cargar_archivo_local(uploaded_facturas, "Facturas")
                    if manifesto is not None and facturas is not None:
                        st.session_state.gastos_datos["manifesto"] = manifesto
                        st.session_state.gastos_datos["facturas"] = facturas
                        st.rerun()

    # --- Contenido principal (MANTENER LÓGICA ORIGINAL COMPLETA) ---
    if st.session_state.gastos_datos["manifesto"] is not None:
        manifesto = st.session_state.gastos_datos["manifesto"]
        facturas = st.session_state.gastos_datos["facturas"]
        st.header("⚙️ Configuración de Procesamiento")
        st.subheader("🔍 Detección Automática de Columnas")
        col1, col2 = st.columns(2)
        with col1:
            st.write("**Manifiesto**")
            guia_opciones_m = [col for col in manifesto.columns if "GUIA" in str(col).upper() or "GUÍA" in str(col).upper()]
            if not guia_opciones_m:
                guia_opciones_m = manifesto.columns.tolist()
            guia_m = st.selectbox("Columna Guía", guia_opciones_m, index=0 if guia_opciones_m else 0)
            subtotal_opciones_m = [col for col in manifesto.columns if "SUBT" in str(col).upper() or "TOTAL" in str(col).upper() or "VALOR" in str(col).upper()]
            if not subtotal_opciones_m:
                subtotal_opciones_m = manifesto.columns.tolist()
            subtotal_m = st.selectbox("Columna Subtotal/Valor", subtotal_opciones_m, index=0 if subtotal_opciones_m else 0)
            ciudad_opciones_m = [col for col in manifesto.columns if "CIUDAD" in str(col).upper() or "DES" in str(col).upper() or "DESTINO" in str(col).upper()]
            if not ciudad_opciones_m:
                ciudad_opciones_m = manifesto.columns.tolist()
            ciudad_destino = st.selectbox("Columna Ciudad Destino", ciudad_opciones_m, index=0 if ciudad_opciones_m else 0)
            st.caption("⚠ **IMPORTANTE:** Para agrupar gastos se usará SOLO el DESTINATARIO del manifiesto")
            destinatario_opciones_m = [col for col in manifesto.columns if any(p in str(col).upper() for p in ["DEST", "CONSIG", "CLIEN", "NOMB", "RAZON"])]
            if destinatario_opciones_m:
                st.info(f"Columnas de destinatario detectadas: {', '.join(destinatario_opciones_m)}")
        with col2:
            st.write("**Facturas**")
            guia_opciones_f = [col for col in facturas.columns if "GUIA" in str(col).upper() or "GUÍA" in str(col).upper()]
            if not guia_opciones_f:
                guia_opciones_f = facturas.columns.tolist()
            guia_f = st.selectbox("Columna Guía (Facturas)", guia_opciones_f, index=0 if guia_opciones_f else 0)
            subtotal_opciones_f = [col for col in facturas.columns if any(x in str(col).upper() for x in ["SUBTOTAL", "TOTAL", "IMPORTE", "VALOR"])]
            if not subtotal_opciones_f:
                subtotal_opciones_f = facturas.columns.tolist()
            subtotal_f = st.selectbox("Columna Subtotal (Facturas)", subtotal_opciones_f, index=0 if subtotal_opciones_f else 0)
            if subtotal_f in facturas.columns:
                st.caption("Previsualización de valores de facturas:")
                valores = facturas[subtotal_f].dropna().head(3).tolist()
                for i, val in enumerate(valores, 1):
                    st.code(f"{i}. {val}")
        config = {"guia_m": guia_m, "subtotal_m": subtotal_m, "ciudad_destino": ciudad_destino, "guia_f": guia_f, "subtotal": subtotal_f}
        if st.button("🚀 Procesar Gastos por Tienda", type="primary", use_container_width=True):
            with st.spinner("Procesando y validando datos..."):
                try:
                    resultado, metricas, resumen, validacion, guias_anuladas = procesar_gastos_reconciliacion(manifesto, facturas, config)
                    st.session_state.gastos_datos["resultado"] = resultado
                    st.session_state.gastos_datos["metricas"] = metricas
                    st.session_state.gastos_datos["resumen"] = resumen
                    st.session_state.gastos_datos["validacion"] = validacion
                    st.session_state.gastos_datos["guias_anuladas"] = guias_anuladas
                    st.session_state.gastos_datos["procesado"] = True
                    st.success("✅ Procesamiento completado exitosamente")
                    if validacion["coincide"]:
                        st.balloons()
                        st.success("✅ Validación exitosa: Los totales coinciden dentro del margen aceptable")
                    else:
                        st.warning(f"⚠ Hay una diferencia de ${validacion['diferencia']:,.2f} ({validacion['porcentaje']:.2f}%)")
                    st.info(f"""
                    🎯 **Resumen del procesamiento:**
                    - Guías totales procesadas: {validacion["guias_procesadas"]:,}
                    - Guías facturadas: {validacion["guias_facturadas"]:,} ({validacion["porcentaje_facturadas"]:.1f}%)
                    - **Guías anuladas:** {validacion["guias_anuladas"]:,} ({validacion["porcentaje_anuladas"]:.1f}%)
                    - Piezas totales: {validacion["piezas_totales"]:,}
                    - Grupos de tiendas identificados: {validacion["grupos_identificados"]:,}
                    - **Total facturado:** ${validacion["total_facturas"]:,.2f}
                    - **Total manifiesto:** ${validacion["total_manifiesto"]:,.2f}
                    """)
                except Exception as e:
                    st.error(f"Error en el procesamiento: {str(e)}")
                    st.exception(e)

    # --- Mostrar resultados si ya se procesó (MANTENER TODOS LOS TABS ORIGINALES) ---
    if st.session_state.gastos_datos["procesado"]:
        resultado = st.session_state.gastos_datos["resultado"]
        metricas = st.session_state.gastos_datos["metricas"]
        resumen = st.session_state.gastos_datos["resumen"]
        validacion = st.session_state.gastos_datos["validacion"]
        guias_anuladas = st.session_state.gastos_datos["guias_anuladas"]
        tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
            "📊 Resumen", "✅ Validación", "🏪 Todas las Tiendas", "🚫 Guías Anuladas",
            "🌎 Geografía", "📋 Datos", "💾 Exportar", "📄 Reporte PDF"
        ])

        with tab1:
            # ... (Mantener toda la lógica del tab1) ...
            st.header("📊 Resumen Ejecutivo")
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1: st.metric("Grupos de Tiendas", f"{len(metricas):,}")
            with col2: st.metric("Total Guías", f"{len(resultado):,}")
            with col3: st.metric("Guías Anuladas", f"{validacion['guias_anuladas']:,}")
            with col4: st.metric("Piezas Totales", f"{validacion['piezas_totales']:,}")
            with col5: st.metric("Total Facturado", f"${metricas['SUBTOTAL'].sum():,.2f}")
            st.subheader("Distribución por Tipo de Tienda")
            if not resumen.empty:
                fig = px.pie(resumen, values="SUBTOTAL", names="TIPO", title="Distribución de Gastos por Tipo de Tienda", hole=0.4, color_discrete_sequence=px.colors.qualitative.Set3)
                fig.update_traces(textposition="inside", textinfo="percent+label")
                fig.update_layout(template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig, use_container_width=True)
            st.subheader("Resumen por Tipo de Tienda")
            if not resumen.empty:
                resumen_display = resumen.copy()
                resumen_display["PIEZAS_POR_GUIA"] = (resumen_display["PIEZAS"] / resumen_display["GUIAS"]).round(2)
                st.dataframe(resumen_display.style.format({"SUBTOTAL": "${:,.2f}", "PORCENTAJE": "{:.2f}%", "PIEZAS_POR_GUIA": "{:.2f}"}), use_container_width=True, hide_index=True)

        with tab2:
            # ... (Mantener toda la lógica del tab2) ...
            st.header("✅ Validación de Totales")
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1: st.metric("Total Manifiesto", f"${validacion['total_manifiesto']:,.2f}")
            with col2: st.metric("Total Facturas", f"${validacion['total_facturas']:,.2f}")
            with col3: st.metric("Diferencia", f"${validacion['diferencia']:,.2f}")
            with col4: st.metric("% Diferencia", f"{validacion['porcentaje']:.2f}%")
            with col5: st.metric("Guías Anuladas", f"{validacion['guias_anuladas']:,}")
            if validacion['coincide']: st.success("✅ **Validación Exitosa**\n\nLos totales coinciden dentro del margen aceptable.")
            else: st.warning(f"⚠ **Validación con Diferencia**\n\nHay una diferencia de **${validacion['diferencia']:,.2f}** ({validacion['porcentaje']:.2f}%). Revisar guías anuladas.")
            st.info(f"""
            **📈 Métricas de Coincidencia:**
            - Guías totales procesadas: {validacion['guias_procesadas']:,}
            - Guías facturadas: {validacion['guias_facturadas']:,} ({validacion['porcentaje_facturadas']:.1f}%)
            - **Guías anuladas:** {validacion['guias_anuladas']:,} ({validacion['porcentaje_anuladas']:.1f}%)
            - Piezas totales: {validacion['piezas_totales']:,}
            - Promedio piezas/guía: {(validacion['piezas_totales']/validacion['guias_procesadas']):.1f}
            """)

        with tab3:
            # ... (Mantener toda la lógica del tab3) ...
            st.header("🏪 Gastos por Tienda/Grupo - TODAS LAS TIENDAS")
            st.info("**🏷️ Clasificación Automática:**\n- **TIENDA FÍSICA:** Nombres comerciales (MALL, PLAZA, CENTRO COMERCIAL, etc.)\n- **VENTA WEB:** Nombres personales (ROCIO, ALEJANDRA, MARIA, etc.)\n- **VENTAS AL POR MAYOR:** JOFRE SANTANA (clasificado automáticamente)")
            col_f1, col_f2, col_f3, col_f4 = st.columns(4)
            with col_f1: min_gasto = st.number_input("Gasto mínimo", value=0.0, step=10.0, format="%.2f", key="min_gasto_todas")
            with col_f2: tipo_filtro = st.multiselect("Tipo", metricas['TIPO'].unique() if 'TIPO' in metricas.columns else [], default=[], key="tipo_filtro_todas")
            with col_f3: ciudad_filtro = st.text_input("Ciudad (parcial)", "", key="ciudad_filtro_todas")
            with col_f4: min_piezas = st.number_input("Mín. piezas", value=0, step=1, key="min_piezas_todas")
            buscar_grupo = st.text_input("🔍 Buscar grupo, destinatario o ciudad:", key="buscar_grupo_todas")
            col_o1, col_o2 = st.columns(2)
            with col_o1: orden_campo = st.selectbox("Ordenar por campo:", ['SUBTOTAL', 'GUIAS', 'PIEZAS', 'PORCENTAJE', 'PROMEDIO_POR_PIEZA'], key="orden_campo_todas")
            with col_o2: orden_direccion = st.selectbox("Dirección:", ['Descendente', 'Ascendente'], key="orden_direccion_todas")
            metricas_filt = metricas.copy()
            if min_gasto > 0: metricas_filt = metricas_filt[metricas_filt['SUBTOTAL'] >= min_gasto]
            if tipo_filtro: metricas_filt = metricas_filt[metricas_filt['TIPO'].isin(tipo_filtro)]
            if ciudad_filtro: metricas_filt = metricas_filt[metricas_filt['CIUDADES'].str.contains(ciudad_filtro, case=False, na=False)]
            if min_piezas > 0: metricas_filt = metricas_filt[metricas_filt['PIEZAS'] >= min_piezas]
            if buscar_grupo: metricas_filt = metricas_filt[metricas_filt['GRUPO'].str.contains(buscar_grupo, case=False, na=False) | metricas_filt['DESTINATARIOS'].str.contains(buscar_grupo, case=False, na=False) | metricas_filt['CIUDADES'].str.contains(buscar_grupo, case=False, na=False)]
            if orden_direccion == 'Descendente': metricas_filt = metricas_filt.sort_values(orden_campo, ascending=False)
            else: metricas_filt = metricas_filt.sort_values(orden_campo, ascending=True)
            st.subheader(f"📋 Todas las Tiendas ({len(metricas_filt):,} grupos)")
            if len(metricas_filt) > 0:
                fig = px.bar(metricas_filt.head(30), x='SUBTOTAL', y='GRUPO', orientation='h', title='Distribución de Gastos por Tienda', color='TIPO', text='SUBTOTAL', hover_data=['GUIAS', 'PIEZAS', 'PORCENTAJE', 'DESTINATARIOS', 'CIUDADES', 'PROMEDIO_POR_PIEZA'], color_discrete_sequence=px.colors.qualitative.Set3)
                fig.update_traces(texttemplate='$%{text:,.2f}', textposition='outside')
                fig.update_layout(yaxis={'categoryorder': 'total ascending'}, height=max(400, len(metricas_filt.head(30)) * 25), showlegend=True, template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            if not metricas_filt.empty:
                metricas_display = metricas_filt[['GRUPO','TIPO','GUIAS','PIEZAS','SUBTOTAL','PORCENTAJE','PROMEDIO_POR_PIEZA','DESTINATARIOS','CIUDADES']].copy()
                styled_df = metricas_display.style.format({'SUBTOTAL':'${:,.2f}','PORCENTAJE':'{:.2f}%','PROMEDIO_POR_PIEZA':'${:,.2f}'}).background_gradient(subset=['SUBTOTAL','PORCENTAJE'], cmap='Blues')
                st.dataframe(styled_df, use_container_width=True, height=600)
                st.subheader("📈 Estadísticas de la Tabla Filtrada")
                col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                with col_s1: st.metric("Grupos Mostrados", len(metricas_filt))
                with col_s2: st.metric("Subtotal Total", f"${metricas_filt['SUBTOTAL'].sum():,.2f}")
                with col_s3: st.metric("Piezas Totales", f"{metricas_filt['PIEZAS'].sum():,.0f}")
                with col_s4: st.metric("Promedio por Pieza", f"${(metricas_filt['SUBTOTAL'].sum()/metricas_filt['PIEZAS'].sum()):.2f}")

        with tab4:
            # ... (Mantener toda la lógica del tab4) ...
            st.header("🚫 Guías Anuladas")
            if not guias_anuladas.empty:
                st.info(f"📊 **Total de guías anuladas:** {len(guias_anuladas):,} ({validacion['porcentaje_anuladas']:.1f}% del total)")
                col_g1, col_g2, col_g3, col_g4 = st.columns(4)
                with col_g1: st.metric("Total Guías", len(guias_anuladas))
                with col_g2:
                    total_manifiesto_anuladas = guias_anuladas['SUBTOTAL_MANIFIESTO'].sum()
                    st.metric("Valor Manifiesto", f"${total_manifiesto_anuladas:,.2f}")
                with col_g3: st.metric("Piezas", f"{guias_anuladas['PIEZAS'].sum():,}")
                with col_g4: st.metric("Destinatarios Únicos", guias_anuladas['DESTINATARIO'].nunique())
                st.subheader("🔍 Filtros para Guías Anuladas")
                col_fa1, col_fa2, col_fa3 = st.columns(3)
                with col_fa1: min_valor = st.number_input("Valor mínimo", value=0.0, step=1.0, key="min_valor_anuladas")
                with col_fa2: buscar_destinatario = st.text_input("Buscar destinatario", "", key="buscar_dest_anuladas")
                with col_fa3: buscar_ciudad = st.text_input("Buscar ciudad", "", key="buscar_ciudad_anuladas")
                guias_anuladas_filt = guias_anuladas.copy()
                if min_valor > 0: guias_anuladas_filt = guias_anuladas_filt[guias_anuladas_filt['SUBTOTAL_MANIFIESTO'] >= min_valor]
                if buscar_destinatario: guias_anuladas_filt = guias_anuladas_filt[guias_anuladas_filt['DESTINATARIO'].str.contains(buscar_destinatario, case=False, na=False)]
                if buscar_ciudad: guias_anuladas_filt = guias_anuladas_filt[guias_anuladas_filt['CIUDAD'].str.contains(buscar_ciudad, case=False, na=False)]
                st.subheader(f"📋 Guías Anuladas Filtradas ({len(guias_anuladas_filt):,})")
                columnas_mostrar = ['FECHA_MANIFIESTO', 'GUIA', 'DESTINATARIO', 'CIUDAD', 'PIEZAS', 'SUBTOTAL_MANIFIESTO', 'TIPO']
                columnas_disponibles = [col for col in columnas_mostrar if col in guias_anuladas_filt.columns]
                guias_display = guias_anuladas_filt[columnas_disponibles].copy()
                if 'FECHA_MANIFIESTO' in guias_display.columns:
                    guias_display['FECHA_MANIFIESTO'] = pd.to_datetime(guias_display['FECHA_MANIFIESTO'], errors='coerce').dt.strftime('%d/%m/%Y')
                if not guias_display.empty:
                    st.dataframe(guias_display.style.format({'SUBTOTAL_MANIFIESTO':'${:,.2f}'}), use_container_width=True, height=400)
                    anuladas_por_tipo = guias_anuladas_filt.groupby('TIPO').agg(CANTIDAD=('GUIA','count'), VALOR=('SUBTOTAL_MANIFIESTO','sum')).reset_index()
                    if not anuladas_por_tipo.empty:
                        fig = px.bar(anuladas_por_tipo, x='VALOR', y='TIPO', orientation='h', title='Valor de Guías Anuladas por Tipo', color='CANTIDAD', text='VALOR', color_continuous_scale='Reds')
                        fig.update_traces(texttemplate='$%{text:,.2f}', textposition='outside')
                        fig.update_layout(yaxis={'categoryorder':'total ascending'}, height=400, template="plotly_dark")
                        st.plotly_chart(fig, use_container_width=True)
                    st.download_button(label="📥 Descargar Guías Anuladas (CSV)", data=guias_display.to_csv(index=False), file_name=f"guias_anuladas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", mime="text/csv", use_container_width=True)
            else:
                st.success("✅ No hay guías anuladas. Todas las guías del manifiesto tienen factura asociada.")

        with tab5:
            # ... (Mantener toda la lógica del tab5) ...
            st.header("🌎 Distribución Geográfica")
            if 'CIUDAD' in resultado.columns:
                df_facturadas = resultado[resultado['ESTADO'] == 'FACTURADA']
                ciudad_agrupada = df_facturadas.groupby('CIUDAD').agg({'GUIA_LIMPIA':'count','PIEZAS':'sum','SUBTOTAL':'sum'}).reset_index()
                ciudad_agrupada = ciudad_agrupada.rename(columns={'GUIA_LIMPIA':'TOTAL_GUIAS','PIEZAS':'TOTAL_PIEZAS','SUBTOTAL':'TOTAL_SUBTOTAL'}).sort_values('TOTAL_SUBTOTAL', ascending=False)
                st.subheader("Top 15 Ciudades por Gasto (Facturadas)")
                if not ciudad_agrupada.empty:
                    fig = px.bar(ciudad_agrupada.head(15), x='TOTAL_SUBTOTAL', y='CIUDAD', orientation='h', title='Distribución por Ciudad', color='TOTAL_PIEZAS', color_continuous_scale='Viridis', text='TOTAL_SUBTOTAL', hover_data=['TOTAL_GUIAS','TOTAL_PIEZAS'])
                    fig.update_traces(texttemplate='$%{text:,.2f}', textposition='outside')
                    fig.update_layout(yaxis={'categoryorder':'total ascending'}, height=600, template="plotly_dark")
                    st.plotly_chart(fig, use_container_width=True)
                    st.subheader("Tabla Detallada por Ciudad")
                    ciudad_agrupada['PORCENTAJE'] = (ciudad_agrupada['TOTAL_SUBTOTAL'] / ciudad_agrupada['TOTAL_SUBTOTAL'].sum() * 100).round(2)
                    ciudad_agrupada['PROMEDIO_POR_PIEZA'] = (ciudad_agrupada['TOTAL_SUBTOTAL'] / ciudad_agrupada['TOTAL_PIEZAS']).round(2)
                    st.dataframe(ciudad_agrupada.style.format({'TOTAL_SUBTOTAL':'${:,.2f}','PORCENTAJE':'{:.2f}%','PROMEDIO_POR_PIEZA':'${:,.2f}'}), use_container_width=True, height=400)

        with tab6:
            # ... (Mantener toda la lógica del tab6) ...
            st.header("📋 Datos Detallados")
            col_fd1, col_fd2, col_fd3 = st.columns(3)
            with col_fd1: grupo_filtro = st.multiselect("Filtrar por Grupo:", sorted(resultado['GRUPO'].unique()), default=[], key="filtro_grupo_detalle")
            with col_fd2: estado_filtro = st.selectbox("Estado:", ['TODOS','FACTURADA','ANULADA'], key="estado_filtro_detalle")
            with col_fd3: min_subtotal = st.number_input("Mínimo subtotal:", min_value=0.0, value=0.0, step=1.0, key="min_subtotal_detalle")
            buscar = st.text_input("🔍 Buscar por guía, destinatario o ciudad:", key="buscar_detalle")
            datos_filt = resultado.copy()
            if grupo_filtro: datos_filt = datos_filt[datos_filt['GRUPO'].isin(grupo_filtro)]
            if estado_filtro == 'FACTURADA': datos_filt = datos_filt[datos_filt['ESTADO'] == 'FACTURADA']
            elif estado_filtro == 'ANULADA': datos_filt = datos_filt[datos_filt['ESTADO'] == 'ANULADA']
            datos_filt = datos_filt[datos_filt['SUBTOTAL_MANIFIESTO'] >= min_subtotal]
            if buscar: datos_filt = datos_filt[datos_filt['GUIA_LIMPIA'].astype(str).str.contains(buscar, case=False, na=False) | datos_filt['DESTINATARIO'].astype(str).str.contains(buscar, case=False, na=False) | datos_filt['CIUDAD'].astype(str).str.contains(buscar, case=False, na=False) | datos_filt['GRUPO'].str.contains(buscar, case=False, na=False)]
            columnas_mostrar = ['FECHA_MANIFIESTO','GUIA_LIMPIA','ESTADO','GRUPO','DESTINATARIO','CIUDAD','PIEZAS','SUBTOTAL_MANIFIESTO','SUBTOTAL','DIFERENCIA','TIPO']
            columnas_disponibles = [col for col in columnas_mostrar if col in datos_filt.columns]
            st.subheader(f"Registros filtrados: {len(datos_filt):,}")
            if not datos_filt.empty:
                datos_display = datos_filt[columnas_disponibles].copy()
                if 'FECHA_MANIFIESTO' in datos_display.columns:
                    datos_display['FECHA_MANIFIESTO'] = pd.to_datetime(datos_display['FECHA_MANIFIESTO'], errors='coerce').dt.strftime('%d/%m/%Y')
                styled_datos = datos_display.style.format({'SUBTOTAL_MANIFIESTO':'${:,.2f}','SUBTOTAL':'${:,.2f}','DIFERENCIA':'${:,.2f}'})
                def color_estado(val):
                    if val == 'ANULADA': return 'background-color: #ffcccc'
                    elif val == 'FACTURADA': return 'background-color: #ccffcc'
                    return ''
                styled_datos = styled_datos.applymap(color_estado, subset=['ESTADO'])
                st.dataframe(styled_datos, use_container_width=True, height=500)
                st.download_button(label="📥 Descargar datos filtrados (CSV)", data=datos_display.to_csv(index=False), file_name=f"datos_filtrados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", mime="text/csv", use_container_width=True)
            st.subheader("📈 Estadísticas de los Datos Filtrados")
            if not datos_filt.empty:
                col_ed1, col_ed2, col_ed3, col_ed4 = st.columns(4)
                with col_ed1:
                    st.metric("Registros", len(datos_filt))
                    st.caption(f"Facturadas: {len(datos_filt[datos_filt['ESTADO']=='FACTURADA']):,}")
                    st.caption(f"Anuladas: {len(datos_filt[datos_filt['ESTADO']=='ANULADA']):,}")
                with col_ed2: st.metric("Subtotal Total", f"${datos_filt['SUBTOTAL'].sum():,.2f}")
                with col_ed3: st.metric("Piezas Totales", f"{datos_filt['PIEZAS'].sum():,.0f}")
                with col_ed4: st.metric("Promedio por Pieza", f"${(datos_filt['SUBTOTAL'].sum()/datos_filt['PIEZAS'].sum()):.2f}")

        with tab7:
            # ... (Mantener toda la lógica del tab7) ...
            st.header("💾 Exportar Resultados")
            formato = st.radio("Seleccione el formato de exportación:", ['Excel Formato Exacto (.xlsx)', 'Excel Normal (.xlsx)', 'CSV (.csv)', 'JSON (.json)'], horizontal=True, key="formato_export")
            nombre_base = st.text_input("Nombre base para los archivos:", value=f"gastos_tiendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}", key="nombre_base_export")
            if st.button("📥 Generar y Descargar", use_container_width=True, type="primary"):
                with st.spinner("Generando archivo..."):
                    if formato == 'Excel Formato Exacto (.xlsx)':
                        excel_output = generar_excel_con_formato_exacto(metricas, resultado, guias_anuladas, manifesto)
                        if excel_output:
                            st.download_button(label="⬇️ Descargar Excel Formato Exacto", data=excel_output.getvalue(), file_name=f"{nombre_base}_formato_exacto.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    elif formato == 'Excel Normal (.xlsx)':
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            metricas.to_excel(writer, sheet_name='Gastos por Grupo', index=False)
                            resultado.to_excel(writer, sheet_name='Datos Completos', index=False)
                            resumen.to_excel(writer, sheet_name='Resumen por Tipo', index=False)
                            pd.DataFrame([validacion]).to_excel(writer, sheet_name='Validación', index=False)
                            if not guias_anuladas.empty:
                                guias_anuladas.to_excel(writer, sheet_name='Guias Anuladas', index=False)
                        st.download_button(label="⬇️ Descargar Excel Normal", data=output.getvalue(), file_name=f"{nombre_base}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    elif formato == 'CSV (.csv)':
                        csv_data = resultado.to_csv(index=False)
                        st.download_button(label="⬇️ Descargar CSV", data=csv_data, file_name=f"{nombre_base}.csv", mime="text/csv", use_container_width=True)
                    else:
                        json_data = resultado.to_json(orient='records', indent=2)
                        st.download_button(label="⬇️ Descargar JSON", data=json_data, file_name=f"{nombre_base}.json", mime="application/json", use_container_width=True)

        with tab8:
            # ... (Mantener toda la lógica del tab8) ...
            st.header("📄 Generar Reporte PDF Ejecutivo")
            st.info("**📋 Contenido del Reporte PDF:**\n1. Métricas principales del análisis\n2. Resumen por tipo de tienda\n3. Top 15 grupos por gasto\n4. Análisis ejecutivo con guías anuladas\n5. Recomendaciones")
            if st.button("🖨️ Generar Reporte PDF", type="primary", use_container_width=True):
                with st.spinner("Generando reporte PDF..."):
                    pdf_path = generar_pdf_reporte(metricas, resumen, validacion)
                    if pdf_path:
                        with open(pdf_path, "rb") as pdf_file:
                            pdf_bytes = pdf_file.read()
                        st.download_button(label="📥 Descargar Reporte PDF", data=pdf_bytes, file_name=f"reporte_ejecutivo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf", mime="application/pdf", use_container_width=True)
    else:
        st.info("👆 **Por favor, carga los archivos de manifiesto y facturas desde el panel lateral para comenzar el análisis.**")
        st.markdown("### 📋 Estructura esperada del Manifiesto:")
        st.code("""FECHA,GUIA,ORIGEN,GUIA2,DESTINATARIO,SERVICIO,TRANSPORTE,PIEZAS,PESO,FLETE,SUBTOTAL
46004,20176386,DURAN,LC52450203,Lider celeley zamb CAR,SEC,,1,16,3.41,3.41
46006,LC52390589,GUAYAQUIL,LC52516378,COMERCIALIZADO CAR,PRI,,1,4,2.15,2.15""")
        st.markdown("### 📋 Estructura esperada de Facturas:")
        st.code("""GUIA,SUBTOTAL
LC52450203,3.41
LC52516378,2.15""")
    st.markdown('</div>', unsafe_allow_html=True)

# ==============================================================================
# MÓDULO: AUDITORIA DE CORREOS (SIMPLIFICADO)
# ==============================================================================
def show_auditoria_correos():
    add_back_button(key="back_auditoria")
    show_module_header("📧 Auditoria de Correos", "Analisis inteligente de novedades por email")
    st.markdown('<div class="module-content">', unsafe_allow_html=True)
    st.info("Módulo de auditoría de correos. Configurar credenciales en el panel lateral.")
    st.markdown('</div>', unsafe_allow_html=True)

# ==============================================================================
# CLASIFICACIÓN INTELIGENTE DE PRODUCTOS (TextileClassifier, DataProcessor, etc.) - SIN CAMBIOS
# ==============================================================================
GENDER_MAP = {
    'GIRLS': 'Mujer', 'WOMEN': 'Mujer', 'WOMENS': 'Mujer', 'WOMAN': 'Mujer',
    'LADIES': 'Mujer', 'FEMALE': 'Mujer', 'MUJER': 'Mujer', 'LADY': 'Mujer',
    'GUYS': 'Hombre', 'MEN': 'Hombre', 'MENS': 'Hombre', 'MAN': 'Hombre',
    'HOMBRE': 'Hombre', 'MALE': 'Hombre', 'BOY': 'Hombre', 'BOYS': 'Hombre',
    'UNISEX': 'Unisex', 'KIDS': 'Niño/a', 'CHILD': 'Niño/a', 'CHILDREN': 'Niños',
    'BABY': 'Bebé', 'INFANT': 'Bebé', 'TODDLER': 'Bebé'
}
CATEGORY_MAP = {
    'TEES': 'Camiseta', 'TEE': 'Camiseta', 'T-SHIRT': 'Camiseta', 'TSHIRT': 'Camiseta',
    'TANK': 'Camiseta sin mangas', 'TANK TOP': 'Camiseta sin mangas',
    'TOP': 'Top', 'TOPS': 'Top', 'BABY TEE': 'Top', 'BABYTEE': 'Top',
    'BARE': 'Blusa', 'CORE': 'Blusa', 'BLOUSE': 'Blusa',
    'GRAPHIC': 'Camiseta estampada', 'GRAPHICS': 'Camiseta estampada',
    'POLO': 'Polo', 'POLOS': 'Polo', 'SHIRT': 'Camisa',
    'BUTTON-DOWN': 'Camisa', 'BUTTONDOWN': 'Camisa',
    'DRESS': 'Vestido', 'DRESSES': 'Vestido', 'SUNDRESS': 'Vestido',
    'PANTS': 'Pantalón', 'PANT': 'Pantalón', 'TROUSERS': 'Pantalón',
    'JEANS': 'Jeans', 'JEAN': 'Jeans', 'DENIM': 'Jeans',
    'BOOTCUT': 'Pantalón bootcut', 'SKINNY': 'Pantalón ajustado',
    'STRAIGHT': 'Pantalón recto', 'FLARE': 'Pantalón campana',
    'SHORTS': 'Short', 'SHORT': 'Short', 'BERMUDA': 'Bermuda',
    'JACKET': 'Chaqueta', 'JACKETS': 'Chaqueta', 'HOODIE': 'Sudadera',
    'SWEATSHIRT': 'Sudadera', 'SWEATER': 'Suéter', 'BLAZER': 'Blazer',
    'BVD': 'Ropa interior', 'BOXER': 'Boxer', 'BOXERS': 'Boxer',
    'UNDERWEAR': 'Ropa interior', 'BRIEF': 'Calzoncillo',
    'BELT': 'Cinturón', 'BELTS': 'Cinturón', 'HAT': 'Gorro',
    'CAP': 'Gorra', 'SOCKS': 'Medias', 'SOCK': 'Medias',
    'WOVEN': 'Tejido', 'KNIT': 'Tejido', 'KNITTED': 'Tejido',
    'SOLID': 'Sólido', 'BASIC': 'Básico', 'BASICO': 'Básico',
    'LEATHER': 'Cuero', 'SUMMER': 'Verano', 'WINTER': 'Invierno',
    'SPRING': 'Primavera', 'FALL': 'Otoño', 'AUTUMN': 'Otoño',
    'FASHION': 'Moda', 'SS': 'Primavera-Verano', 'FW': 'Otoño-Invierno'
}
COLOR_MAP = {
    'BLACK': 'Negro', 'BLACK/DARK': 'Negro', 'DARK BLACK': 'Negro Oscuro',
    'WHITE': 'Blanco', 'WHITE/LIGHT': 'Blanco', 'CREAM': 'Crema',
    'RED': 'Rojo', 'RASPBERRY': 'Frambuesa', 'EARTH RED': 'Rojo Tierra',
    'BLUE': 'Azul', 'NAVY': 'Azul Marino', 'LIGHT BLUE': 'Azul Claro', 'ROYAL': 'Azul Rey',
    'GREEN': 'Verde', 'GREEN GABLES': 'Verde Gabardina', 'OLIVE': 'Verde Oliva',
    'YELLOW': 'Amarillo', 'GOLD': 'Dorado', 'MUSTARD': 'Mostaza',
    'PINK': 'Rosa', 'HOT PINK': 'Rosa Fuerte', 'BLUSH': 'Rosa Palo',
    'PURPLE': 'Morado', 'VIOLET': 'Violeta', 'LILAC': 'Lila', 'LAVENDER': 'Lavanda',
    'ORANGE': 'Naranja', 'PEACH': 'Durazno', 'CORAL': 'Coral',
    'BROWN': 'Marrón', 'TAN': 'Café Claro', 'COFFEE': 'Café',
    'GRAY': 'Gris', 'GREY': 'Gris', 'SILVER': 'Plateado', 'CHARCOAL': 'Gris Carbón',
    'BEIGE': 'Beige', 'KHAKI': 'Caqui', 'CAMEL': 'Camello',
    'BLEACH': 'Blanqueado', 'LIGHT WASH': 'Lavado Claro', 'DARK WASH': 'Lavado Oscuro',
    'DARK': 'Oscuro', 'LIGHT': 'Claro', 'TENDER': 'Tender (Amarillo Suave)',
    'MULTI': 'Multicolor', 'MULTICOLOR': 'Multicolor', 'MIXED': 'Multicolor'
}
SIZE_HIERARCHY = [
    '3XS', 'XXS', '2XS', 'XS', 'XSMALL', 'EXTRA SMALL',
    'S', 'SMALL', 'M', 'MEDIUM', 'L', 'LARGE', 
    'XL', 'XLARGE', 'EXTRA LARGE', '1XL',
    '2XL', 'XXL', 'XXLARGE', '3XL', 'XXXL', 'XXXLARGE',
    '4XL', 'XXXXL', '5XL', 'OS', 'ONE SIZE', 'UNICA', 'U'
]
SIZE_NORMALIZATION = {
    'XSMALL': 'XS', 'EXTRA SMALL': 'XS', 'XXS': '2XS', '2XS': '2XS',
    'SMALL': 'S', 'MEDIUM': 'M', 'LARGE': 'L',
    'XLARGE': 'XL', 'EXTRA LARGE': 'XL', '1XL': 'XL',
    'XXLARGE': '2XL', 'XXL': '2XL', '2XL': '2XL',
    'XXXLARGE': '3XL', 'XXXL': '3XL', '3XL': '3XL',
    'XXXXL': '4XL', '4XL': '4XL', '5XL': '5XL',
    'OS': 'Única', 'ONE SIZE': 'Única', 'UNICA': 'Única', 'U': 'Única'
}
IGNORE_WORDS = {'AERO', 'OF', 'THE', 'AND', 'IN', 'WITH', 'FOR', 'BY', 'ON', 'AT', 'TO', 'FROM'}

class TextileClassifier:
    def __init__(self):
        self.gender_map = GENDER_MAP
        self.category_map = CATEGORY_MAP
        self.color_map = COLOR_MAP
        self.size_hierarchy = SIZE_HIERARCHY
        self.size_norm = SIZE_NORMALIZATION
        self.ignore_words = IGNORE_WORDS
        
    def classify_product(self, product_name: str) -> dict:
        # ... (Código original de classify_product) ...
        if not product_name or not isinstance(product_name, str):
            return self._get_empty_classification()
        product_name = str(product_name).upper().strip()
        words = product_name.split()
        if not words:
            return self._get_empty_classification()
        color_idx, color_norm = self._detect_color_index(words)
        size_idx, size_norm = self._detect_size_index(words)
        gender_phrase, gender_norm = self._extract_gender_phrase(words)
        desc_words = self._extract_description(words, gender_phrase, color_idx, size_idx)
        descripcion = ' '.join(desc_words) if desc_words else ''
        category_info = self._detect_category(words)
        style_info = self._detect_style(words)
        classification = {
            'Genero_Raw': gender_phrase,
            'Genero': gender_norm,
            'Descripcion': descripcion,
            'Categoria': category_info['categoria'],
            'Subcategoria': category_info['subcategoria'],
            'Color': color_norm,
            'Talla': size_norm,
            'Material': style_info['material'],
            'Estilo': style_info['estilo']
        }
        return self._clean_classification(classification)
    
    def _extract_gender_phrase(self, words):
        if not words:
            return '', 'Unisex'
        if len(words) >= 2 and words[0] == 'AERO' and words[1] in self.gender_map:
            phrase = f"{words[0]} {words[1]}"
            norm = self.gender_map[words[1]]
            return phrase, norm
        if words[0] in self.gender_map:
            return words[0], self.gender_map[words[0]]
        for i, w in enumerate(words):
            if w in self.gender_map:
                phrase = ' '.join(words[:i+1])
                return phrase, self.gender_map[w]
        return words[0], 'Unisex'
    
    def _detect_color_index(self, words):
        for i, w in enumerate(words):
            if w in self.color_map:
                return i, self.color_map[w]
            for color_key, color_val in self.color_map.items():
                if color_key in w and len(w) < len(color_key) + 3:
                    return i, color_val
        return None, 'No Especificado'
    
    def _detect_size_index(self, words):
        for i, w in enumerate(words):
            if w in self.size_hierarchy:
                return i, self._normalize_size(w)
            if w in self.size_norm:
                return i, self.size_norm[w]
        return None, 'Única'
    
    def _extract_description(self, words, gender_phrase, color_idx, size_idx):
        if not gender_phrase:
            return []
        gender_words = gender_phrase.split()
        start_idx = len(gender_words)
        end_candidates = []
        if color_idx is not None:
            end_candidates.append(color_idx)
        if size_idx is not None:
            end_candidates.append(size_idx)
        if end_candidates:
            end_idx = min(end_candidates)
        else:
            end_idx = len(words)
        if end_idx <= start_idx:
            return []
        desc_words = words[start_idx:end_idx]
        desc_words = [w for w in desc_words if w not in self.ignore_words and len(w) > 1]
        return desc_words
    
    def _normalize_size(self, size):
        return self.size_norm.get(size, size)
    
    def _detect_category(self, words):
        categoria = 'General'
        subcategoria = ''
        found_categories = []
        for word in words:
            if word in self.category_map:
                found_categories.append(self.category_map[word])
        if found_categories:
            priority = ['Camiseta', 'Pantalón', 'Jeans', 'Vestido', 'Chaqueta', 'Sudadera', 'Top', 'Short', 'Polo']
            for p in priority:
                if p in found_categories:
                    categoria = p
                    others = [c for c in found_categories if c != p]
                    subcategoria = ' / '.join(others[:2])
                    break
            else:
                categoria = found_categories[0]
                if len(found_categories) > 1:
                    subcategoria = ' / '.join(found_categories[1:3])
        return {'categoria': categoria, 'subcategoria': subcategoria}
    
    def _detect_style(self, words):
        material = ''
        estilo = ''
        materials = {'COTTON':'Algodón', 'POLYESTER':'Poliéster', 'DENIM':'Denim', 'LEATHER':'Cuero', 'WOOL':'Lana', 'SILK':'Seda', 'LINEN':'Lino', 'RAYON':'Rayón', 'SPANDEX':'Spandex', 'NYLON':'Nylon'}
        styles = {'GRAPHIC':'Estampado', 'PRINTED':'Estampado', 'LOGO':'Con Logo', 'SOLID':'Liso', 'STRIPED':'Rayado', 'PLAID':'Cuadros', 'BASIC':'Básico', 'PREMIUM':'Premium', 'FASHION':'Moda', 'VINTAGE':'Vintage', 'CLASSIC':'Clásico', 'SLIM':'Slim Fit', 'RELAXED':'Relaxed Fit', 'OVERSIZED':'Oversized'}
        for word in words:
            if word in materials:
                material = materials[word]
            if word in styles:
                estilo = styles[word]
        return {'material': material, 'estilo': estilo}
    
    def _clean_classification(self, classification):
        if not classification['Genero']:
            classification['Genero'] = 'Unisex'
        if not classification['Categoria']:
            classification['Categoria'] = 'General'
        if not classification['Descripcion']:
            classification['Descripcion'] = '-'
        for key in classification:
            if isinstance(classification[key], str):
                classification[key] = classification[key].strip()
        return classification
    
    def _get_empty_classification(self):
        return {
            'Genero_Raw': '', 'Genero': 'Unisex', 'Descripcion': '-',
            'Categoria': 'General', 'Subcategoria': '',
            'Color': 'No Especificado', 'Talla': 'Única',
            'Material': '', 'Estilo': ''
        }

class DataProcessor:
    def __init__(self):
        self.classifier = TextileClassifier()
    
def process_excel_file(self, file) -> pd.DataFrame:
    try:
        if hasattr(file, 'name'):
            filename = file.name.lower()
            if filename.endswith('.csv'):
                df = pd.read_csv(file, encoding='utf-8')
            else:
                df = pd.read_excel(file, engine='openpyxl')
        else:
            df = pd.read_excel(file)

        df.columns = df.columns.str.strip().str.upper()

        # --- Detección de columna de producto (PRIORIZA coincidencia exacta) ---
        product_col = None
        # Primero buscar columna exactamente "PRODUCTO"
        if 'PRODUCTO' in df.columns:
            product_col = 'PRODUCTO'
        else:
            # Si no existe exacta, buscar la primera que CONTENGA "PRODUCTO" pero NO "CODIGO"
            product_aliases = ['PRODUCTO', 'ITEM', 'DESCRIPCION', 'DESCRIPTION', 'ARTICULO', 'NOMBRE', 'NAME', 'ITEM DESCRIPTION']
            for alias in product_aliases:
                matching_cols = [c for c in df.columns if alias in c and 'CODIGO' not in c]
                if matching_cols:
                    product_col = matching_cols[0]
                    break
        # Si aún nada, intentar detectar por contenido típico de producto
        if product_col is None:
            for col in df.columns:
                if len(df) > 0:
                    sample = str(df[col].iloc[0])
                    if len(sample) > 10 and any(x in sample.upper() for x in ['AERO', 'GIRLS', 'WOMEN', 'MEN', 'BANANA', 'PRICE CLUB']):
                        product_col = col
                        break
        if product_col is None:
            product_col = df.columns[0]  # último recurso

        # Si la columna seleccionada no se llama exactamente 'PRODUCTO', renombrar eliminando duplicados
        if product_col != 'PRODUCTO':
            if 'PRODUCTO' in df.columns:
                df = df.drop(columns=['PRODUCTO'])  # eliminar la columna 'PRODUCTO' preexistente
            df = df.rename(columns={product_col: 'PRODUCTO'})

        df = df.dropna(subset=['PRODUCTO'])
        df['PRODUCTO'] = df['PRODUCTO'].astype(str).str.strip()

        # --- Columna de bodega destino ---
        bodega_col = None
        bodega_aliases = ['BODEGA RECIBE', 'BODEGA DESTINO', 'SUCURSAL DESTINO', 'DESTINO', 'BODEGA', 'SUCURSAL', 'TIENDA', 'STORE']
        for alias in bodega_aliases:
            matching_cols = [c for c in df.columns if alias in c]
            if matching_cols:
                bodega_col = matching_cols[0]
                break
        if bodega_col:
            if bodega_col != 'BODEGA_RECIBE':
                if 'BODEGA_RECIBE' in df.columns:
                    df = df.drop(columns=['BODEGA_RECIBE'])
                df = df.rename(columns={bodega_col: 'BODEGA_RECIBE'})

        # --- Columna de cantidad ---
        cant_col = None
        cant_aliases = ['CANTIDAD', 'QUANTITY', 'UNIDADES', 'QTY', 'CANT']
        for alias in cant_aliases:
            matching_cols = [c for c in df.columns if alias in c]
            if matching_cols:
                cant_col = matching_cols[0]
                break
        if cant_col:
            if cant_col != 'CANTIDAD':
                if 'CANTIDAD' in df.columns:
                    df = df.drop(columns=['CANTIDAD'])
                df = df.rename(columns={cant_col: 'CANTIDAD'})
            df['CANTIDAD'] = pd.to_numeric(df['CANTIDAD'].astype(str).str.replace(',', '').str.replace(' ', ''), errors='coerce').fillna(0)

        # --- Columna de costo ---
        costo_col = None
        costo_aliases = ['COSTO', 'COST', 'PRECIO', 'COSTO UNITARIO']
        for alias in costo_aliases:
            matching_cols = [c for c in df.columns if alias in c]
            if matching_cols:
                costo_col = matching_cols[0]
                break
        if costo_col:
            if costo_col != 'COSTO':
                if 'COSTO' in df.columns:
                    df = df.drop(columns=['COSTO'])
                df = df.rename(columns={costo_col: 'COSTO'})
            df['COSTO'] = pd.to_numeric(df['COSTO'].astype(str).str.replace(',', '').str.replace(' ', ''), errors='coerce').fillna(0)

        # --- Columna de total (opcional, calculamos si no existe) ---
        if 'TOTAL' not in df.columns:
            if 'CANTIDAD' in df.columns and 'COSTO' in df.columns:
                df['TOTAL'] = df['CANTIDAD'] * df['COSTO']

        # --- Columna de fecha ---
        fecha_col = None
        fecha_aliases = ['FECHA', 'DATE', 'DIA', 'DAY']
        for alias in fecha_aliases:
            matching_cols = [c for c in df.columns if alias in c]
            if matching_cols:
                fecha_col = matching_cols[0]
                break
        if fecha_col:
            if fecha_col != 'FECHA':
                if 'FECHA' in df.columns:
                    df = df.drop(columns=['FECHA'])
                df = df.rename(columns={fecha_col: 'FECHA'})
            df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce', dayfirst=True)

        if df.empty or 'PRODUCTO' not in df.columns:
            st.error("No se encontraron datos válidos para clasificar.")
            return pd.DataFrame()

        df = self._classify_products(df)
        return df

    except Exception as e:
        st.error(f"Error procesando archivo: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return pd.DataFrame()

def mostrar_clasificacion_inteligente():
    st.markdown("### 📂 Cargar archivo de productos para clasificación inteligente")
    st.markdown("Soporta archivos: **Excel (.xlsx, .xls)** y **CSV (.csv)**")
    if 'clasificacion_data' not in st.session_state:
        st.session_state.clasificacion_data = pd.DataFrame()
    if 'clasificacion_loaded' not in st.session_state:
        st.session_state.clasificacion_loaded = False
    processor = DataProcessor()
    col_up1, col_up2 = st.columns([3, 1])
    with col_up1:
        uploaded_file = st.file_uploader("Selecciona tu archivo", type=['xlsx','xls','csv'], key="clasificador_file_uploader")
    with col_up2:
        if st.button("🔄 Limpiar datos", key="btn_limpiar_clasificacion", use_container_width=True):
            st.session_state.clasificacion_data = pd.DataFrame()
            st.session_state.clasificacion_loaded = False
            st.rerun()
    if uploaded_file is not None:
        with st.spinner("🔍 Procesando y clasificando productos..."):
            try:
                df_procesado = processor.process_excel_file(uploaded_file)
                if not df_procesado.empty:
                    st.session_state.clasificacion_data = df_procesado
                    st.session_state.clasificacion_loaded = True
                    st.success(f"✅ **{len(df_procesado)}** productos clasificados exitosamente")
                else:
                    st.error("❌ El archivo no contiene datos válidos para clasificar.")
                    st.session_state.clasificacion_loaded = False
            except Exception as e:
                st.error(f"❌ Error al procesar: {str(e)}")
                st.session_state.clasificacion_loaded = False
    if not st.session_state.clasificacion_loaded:
        st.info("👆 **Sube un archivo** para comenzar el análisis.")
        with st.expander("📋 Ver formato esperado"):
            ejemplo_df = pd.DataFrame({'PRODUCTO': ['AERO GIRLS SS FASHION GRAPHICS TENDER YELLOW XLARGE BABY TEE','AERO WOMEN DENIM JACKET BLUE M','MENS COTTON T-SHIRT BLACK L','KIDS SUMMER SHORTS RED S']})
            st.dataframe(ejemplo_df, use_container_width=True)
            st.markdown("**Extracción automática:**\n- **Género**: \"AERO GIRLS\" → Mujer\n- **Descripción**: \"SS FASHION GRAPHICS TENDER\"\n- **Color**: \"YELLOW\"\n- **Talla**: \"XLARGE\" → XL")
        st.markdown("---")
        st.subheader("🧠 Demo: Probar Clasificador")
        producto_demo = st.text_input("Ingresa nombre de producto", "AERO GIRLS SS FASHION GRAPHICS TENDER YELLOW XLARGE BABY TEE", key="input_demo_clasificador")
        if st.button("🔍 Clasificar", key="btn_demo_clasificar"):
            classifier = TextileClassifier()
            resultado = classifier.classify_product(producto_demo)
            st.markdown("#### 🏷️ Resultado:")
            cols = st.columns(4)
            campos = [("Género (Raw)", resultado['Genero_Raw'], "🏷️"), ("Género", resultado['Genero'], "👤"), ("Descripción", resultado['Descripcion'], "📝"), ("Categoría", resultado['Categoria'], "👕"), ("Color", resultado['Color'], "🎨"), ("Talla", resultado['Talla'], "📏"), ("Material", resultado['Material'], "🧵"), ("Estilo", resultado['Estilo'], "✨")]
            for i, (label, value, icon) in enumerate(campos):
                with cols[i % 4]:
                    st.metric(f"{icon} {label}", value if value else "N/A")
        return
    data = st.session_state.clasificacion_data
    if data.empty:
        st.error("No hay datos para mostrar.")
        return
    total_unidades = int(data['CANTIDAD'].sum()) if 'CANTIDAD' in data.columns else len(data)
    col_k1, col_k2, col_k3, col_k4, col_k5 = st.columns(5)
    with col_k1: st.metric("📦 Total Productos", f"{len(data):,}")
    with col_k2: st.metric("📊 Unidades Totales", f"{total_unidades:,}")
    with col_k3:
        if 'Genero' in data.columns:
            gen_dist = data['Genero'].value_counts()
            st.metric("👤 Género Top", gen_dist.index[0] if len(gen_dist) > 0 else "N/A")
    with col_k4:
        if 'Color' in data.columns:
            col_dist = data['Color'].value_counts()
            st.metric("🎨 Color Top", col_dist.index[0] if len(col_dist) > 0 else "N/A")
    with col_k5:
        if 'Talla' in data.columns:
            talla_dist = data['Talla'].value_counts()
            st.metric("📏 Talla Top", talla_dist.index[0] if len(talla_dist) > 0 else "N/A")
    st.markdown("---")
    st.subheader("📈 Análisis por Dimensiones")
    tab_dims = st.tabs(["🏪 Por Tienda/Bodega", "🎨 Por Color", "📏 Por Talla", "👤 Por Género", "📋 Tabla Dinámica"])
    cantidad_col = 'CANTIDAD' if 'CANTIDAD' in data.columns else None
    bodega_col = None
    for col in ['BODEGA_RECIBE', 'Bodega Recibe', 'Bodega Destino', 'Sucursal Destino']:
        if col in data.columns:
            bodega_col = col
            break
    with tab_dims[0]:
        st.markdown("### 🏪 Análisis por Tienda/Bodega")
        if bodega_col:
            if cantidad_col:
                tienda_stats = data.groupby(bodega_col).agg({cantidad_col:'sum','PRODUCTO':'count'}).reset_index()
                tienda_stats.columns = ['Tienda','Unidades','Productos']
                tienda_stats = tienda_stats.sort_values('Unidades', ascending=False)
            else:
                tienda_stats = data.groupby(bodega_col).size().reset_index(name='Productos')
                tienda_stats.columns = ['Tienda','Productos']
                tienda_stats = tienda_stats.sort_values('Productos', ascending=False)
            col_t1, col_t2 = st.columns([2,1])
            with col_t1:
                fig = px.bar(tienda_stats.head(15), x='Tienda', y='Unidades' if cantidad_col else 'Productos', title="Distribución por Tienda", color='Unidades' if cantidad_col else 'Productos', color_continuous_scale='Viridis')
                fig.update_layout(xaxis_tickangle=-45, template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            with col_t2:
                st.dataframe(tienda_stats.head(20), use_container_width=True, height=500)
            if 'Genero' in data.columns and cantidad_col:
                pivot_tienda_genero = pd.pivot_table(data, values=cantidad_col, index=bodega_col, columns='Genero', aggfunc='sum', fill_value=0, margins=True, margins_name='TOTAL')
                st.dataframe(pivot_tienda_genero.style.format('{:,.0f}'), use_container_width=True)
        else:
            st.warning("No se encontró columna de tienda/bodega en los datos.")
    with tab_dims[1]:
        st.markdown("### 🎨 Análisis por Color")
        if 'Color' in data.columns:
            if cantidad_col:
                color_stats = data.groupby('Color').agg({cantidad_col:['sum','count'],'PRODUCTO':'nunique'}).reset_index()
                color_stats.columns = ['Color','Unidades','Productos','SKUs Únicos']
                color_stats = color_stats.sort_values('Unidades', ascending=False)
            else:
                color_stats = data.groupby('Color').agg({'PRODUCTO':['count','nunique']}).reset_index()
                color_stats.columns = ['Color','Productos','SKUs Únicos']
                color_stats = color_stats.sort_values('Productos', ascending=False)
            col_c1, col_c2 = st.columns([2,1])
            with col_c1:
                fig = px.pie(color_stats.head(15), values='Unidades' if cantidad_col else 'Productos', names='Color', title="Top 15 Colores", hole=0.4)
                fig.update_layout(template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            with col_c2:
                st.dataframe(color_stats.style.format({'Unidades':'{:,}','Productos':'{:,}','SKUs Únicos':'{:,}'}), use_container_width=True, height=400)
            if 'Talla' in data.columns and cantidad_col:
                pivot_color_talla = pd.pivot_table(data, values=cantidad_col, index='Color', columns='Talla', aggfunc='sum', fill_value=0, margins=True, margins_name='TOTAL')
                st.dataframe(pivot_color_talla.style.format('{:,.0f}'), use_container_width=True)
        else:
            st.warning("No hay datos de color disponibles.")
    with tab_dims[2]:
        st.markdown("### 📏 Análisis por Talla")
        if 'Talla' in data.columns:
            talla_order = ['2XS','XS','S','M','L','XL','2XL','3XL','4XL','Única']
            if cantidad_col:
                talla_stats = data.groupby('Talla').agg({cantidad_col:['sum','count'],'PRODUCTO':'nunique'}).reset_index()
                talla_stats.columns = ['Talla','Unidades','Productos','SKUs Únicos']
            else:
                talla_stats = data.groupby('Talla').agg({'PRODUCTO':['count','nunique']}).reset_index()
                talla_stats.columns = ['Talla','Productos','SKUs Únicos']
            talla_stats['Talla_Order'] = talla_stats['Talla'].apply(lambda x: talla_order.index(x) if x in talla_order else 999)
            talla_stats = talla_stats.sort_values('Talla_Order')
            col_ta1, col_ta2 = st.columns([2,1])
            with col_ta1:
                fig = px.bar(talla_stats, x='Talla', y='Unidades' if cantidad_col else 'Productos', title="Distribución por Talla", color='Unidades' if cantidad_col else 'Productos', color_continuous_scale='Plasma')
                fig.update_layout(template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            with col_ta2:
                st.dataframe(talla_stats[['Talla','Unidades' if cantidad_col else 'Productos','SKUs Únicos']].style.format({'Unidades':'{:,}','Productos':'{:,}','SKUs Únicos':'{:,}'}), use_container_width=True, height=400)
            if 'Genero' in data.columns and cantidad_col:
                pivot_talla_genero = pd.pivot_table(data, values=cantidad_col, index='Talla', columns='Genero', aggfunc='sum', fill_value=0, margins=True, margins_name='TOTAL')
                pivot_talla_genero = pivot_talla_genero.reindex([t for t in talla_order if t in pivot_talla_genero.index] + [t for t in pivot_talla_genero.index if t not in talla_order])
                st.dataframe(pivot_talla_genero.style.format('{:,.0f}'), use_container_width=True)
        else:
            st.warning("No hay datos de talla disponibles.")
    with tab_dims[3]:
        st.markdown("### 👤 Análisis por Género")
        if 'Genero' in data.columns:
            if cantidad_col:
                genero_stats = data.groupby('Genero').agg({cantidad_col:['sum','count'],'PRODUCTO':'nunique'}).reset_index()
                genero_stats.columns = ['Género','Unidades','Productos','SKUs Únicos']
                genero_stats = genero_stats.sort_values('Unidades', ascending=False)
            else:
                genero_stats = data.groupby('Genero').agg({'PRODUCTO':['count','nunique']}).reset_index()
                genero_stats.columns = ['Género','Productos','SKUs Únicos']
                genero_stats = genero_stats.sort_values('Productos', ascending=False)
            col_g1, col_g2 = st.columns([2,1])
            with col_g1:
                fig = px.pie(genero_stats, values='Unidades' if cantidad_col else 'Productos', names='Género', title="Distribución por Género", color_discrete_sequence=['#FF6B6B','#4ECDC4','#45B7D1','#96CEB4','#FFEAA7'])
                fig.update_traces(textposition='inside', textinfo='percent+label')
                fig.update_layout(template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            with col_g2:
                st.dataframe(genero_stats.style.format({'Unidades':'{:,}','Productos':'{:,}','SKUs Únicos':'{:,}'}), use_container_width=True, height=300)
            if 'Categoria' in data.columns and cantidad_col:
                pivot_gen_cat = pd.pivot_table(data, values=cantidad_col, index='Genero', columns='Categoria', aggfunc='sum', fill_value=0, margins=True, margins_name='TOTAL')
                st.dataframe(pivot_gen_cat.style.format('{:,.0f}'), use_container_width=True)
                gen_cat_data = data.groupby(['Genero','Categoria']).agg({cantidad_col:'sum'}).reset_index()
                gen_cat_data.columns = ['Género','Categoría','Cantidad']
                fig_stack = px.bar(gen_cat_data, x='Género', y='Cantidad', color='Categoría', title="Distribución por Género y Categoría", barmode='stack')
                fig_stack.update_layout(template="plotly_dark")
                st.plotly_chart(fig_stack, use_container_width=True)
        else:
            st.warning("No hay datos de género disponibles.")
    with tab_dims[4]:
        st.markdown("### 📋 Tabla Dinámica Personalizada")
        available_dims = []
        if 'Genero' in data.columns: available_dims.append('Genero')
        if 'Color' in data.columns: available_dims.append('Color')
        if 'Talla' in data.columns: available_dims.append('Talla')
        if 'Categoria' in data.columns: available_dims.append('Categoria')
        if bodega_col: available_dims.append('Tienda/Bodega')
        if available_dims:
            col_cfg1, col_cfg2, col_cfg3 = st.columns(3)
            filas = col_cfg1.selectbox("Filas (índice)", available_dims, index=0)
            columnas = col_cfg2.selectbox("Columnas", available_dims, index=min(1, len(available_dims)-1) if len(available_dims)>1 else 0)
            valores = col_cfg3.selectbox("Valores a agregar", ['Cantidad (suma)' if cantidad_col else 'Conteo de productos', 'Conteo de productos', 'SKUs únicos'], index=0)
            if filas and columnas and filas != columnas:
                fila_col = 'BODEGA_RECIBE' if filas == 'Tienda/Bodega' and 'BODEGA_RECIBE' in data.columns else filas
                col_col = 'BODEGA_RECIBE' if columnas == 'Tienda/Bodega' and 'BODEGA_RECIBE' in data.columns else columnas
                if 'Cantidad' in valores and cantidad_col:
                    agg_val = cantidad_col
                    agg_func = 'sum'
                elif 'SKUs' in valores:
                    agg_val = 'PRODUCTO'
                    agg_func = 'nunique'
                else:
                    agg_val = cantidad_col if cantidad_col else 'PRODUCTO'
                    agg_func = 'sum' if cantidad_col else 'count'
                try:
                    pivot_custom = pd.pivot_table(data, values=agg_val, index=fila_col, columns=col_col, aggfunc=agg_func, fill_value=0, margins=True, margins_name='TOTAL')
                    st.dataframe(pivot_custom.style.format('{:,.0f}'), use_container_width=True)
                    csv = pivot_custom.to_csv()
                    st.download_button("📥 Descargar CSV", data=csv, file_name=f"tabla_dinamica_{filas}_vs_{columnas}.csv", mime="text/csv")
                except Exception as e:
                    st.error(f"Error generando tabla: {str(e)}")
            else:
                st.info("Selecciona dimensiones diferentes para filas y columnas.")
        else:
            st.warning("No hay dimensiones disponibles para crear tablas dinámicas.")
    st.markdown("---")
    st.subheader("🔍 Filtros y Detalle")
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    filtro_gen = 'Todos'
    filtro_col = 'Todos'
    filtro_talla = 'Todos'
    filtro_bod = 'Todas'
    with col_f1:
        if 'Genero' in data.columns:
            gen_opts = ['Todos'] + sorted(data['Genero'].unique())
            filtro_gen = st.selectbox("👤 Género", gen_opts, key="filtro_gen")
    with col_f2:
        if 'Color' in data.columns:
            col_opts = ['Todos'] + sorted(data['Color'].unique())
            filtro_col = st.selectbox("🎨 Color", col_opts, key="filtro_col")
    with col_f3:
        if 'Talla' in data.columns:
            talla_opts = ['Todos'] + sorted(data['Talla'].unique())
            filtro_talla = st.selectbox("📏 Talla", talla_opts, key="filtro_talla")
    with col_f4:
        if bodega_col:
            bod_opts = ['Todas'] + sorted(data[bodega_col].unique())
            filtro_bod = st.selectbox("🏪 Tienda", bod_opts, key="filtro_bod")
    filtered = data.copy()
    if filtro_gen != 'Todos' and 'Genero' in filtered.columns:
        filtered = filtered[filtered['Genero'] == filtro_gen]
    if filtro_col != 'Todos' and 'Color' in filtered.columns:
        filtered = filtered[filtered['Color'] == filtro_col]
    if filtro_talla != 'Todos' and 'Talla' in filtered.columns:
        filtered = filtered[filtered['Talla'] == filtro_talla]
    if filtro_bod != 'Todas' and bodega_col:
        filtered = filtered[filtered[bodega_col] == filtro_bod]
    st.markdown(f"**📋 Mostrando {len(filtered)} registros filtrados de {len(data)} totales**")
    display_cols = [c for c in ['PRODUCTO','Genero_Raw','Genero','Descripcion','Categoria','Color','Talla','CANTIDAD','BODEGA_RECIBE','FECHA'] if c in filtered.columns]
    if not display_cols:
        display_cols = filtered.columns.tolist()[:10]
    st.dataframe(filtered[display_cols].sort_values(by='CANTIDAD' if 'CANTIDAD' in filtered.columns else display_cols[0], ascending=False), use_container_width=True, height=400)
    if len(filtered) > 0:
        st.markdown("---")
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            filtered.to_excel(writer, sheet_name='Datos Filtrados', index=False)
            resumen_export = []
            if 'Genero' in filtered.columns: resumen_export.append(['Géneros', filtered['Genero'].nunique()])
            if 'Color' in filtered.columns: resumen_export.append(['Colores', filtered['Color'].nunique()])
            if 'Talla' in filtered.columns: resumen_export.append(['Tallas', filtered['Talla'].nunique()])
            if 'CANTIDAD' in filtered.columns: resumen_export.append(['Total Unidades', int(filtered['CANTIDAD'].sum())])
            resumen_export.append(['Total Productos', len(filtered)])
            pd.DataFrame(resumen_export, columns=['Métrica','Valor']).to_excel(writer, sheet_name='Resumen', index=False)
        st.download_button(label="📥 Descargar Excel", data=buffer.getvalue(), file_name=f"clasificacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# ==============================================================================
# FUNCIONES PARA DASHBOARD LOGÍSTICO (Transferencias Diarias) - SIN CAMBIOS
# ==============================================================================
def clasificar_transferencia_diaria(row: pd.Series) -> str:
    sucursal = str(row.get('Sucursal Destino', row.get('Bodega Destino', ''))).upper()
    cantidad = row.get('Cantidad_Entera', 0)
    if any(kw in sucursal for kw in ['PRICE', 'OIL', 'CITY MALL']): return 'Price Club'
    if cantidad >= 500 and cantidad % 100 == 0: return 'Fundas'
    if any(kw in sucursal for kw in TIENDA_WEB): return 'Tienda Web'
    if any(kw in sucursal for kw in FALLAS): return 'Fallas'
    if any(kw in sucursal for kw in VENTAS_POR_MAYOR): return 'Ventas por Mayor'
    if any(tienda.upper() in sucursal for tienda in TIENDAS_REGULARES): return 'Tiendas'
    tiendas_kw = ['AERO', 'MALL', 'CENTRO', 'SHOPPING', 'PLAZA', 'RIOCENTRO']
    if any(kw in sucursal for kw in tiendas_kw): return 'Tiendas'
    return 'Ventas por Mayor'

def procesar_transferencias_diarias(df: pd.DataFrame) -> Dict:
    df = df.dropna(subset=['Secuencial'])
    df['Secuencial'] = df['Secuencial'].astype(str).str.strip()
    df = df[df['Secuencial'] != '']
    cant_col = 'Cantidad Prendas' if 'Cantidad Prendas' in df.columns else 'Cantidad'
    df['Cantidad_Entera'] = df[cant_col].apply(extraer_entero)
    df['Categoria'] = df.apply(clasificar_transferencia_diaria, axis=1)
    suc_por_cat = df.groupby('Categoria')['Sucursal Destino'].nunique().to_dict() if 'Sucursal Destino' in df.columns else {}
    resumen = {
        'total_unidades': int(df['Cantidad_Entera'].sum()),
        'transferencias': int(df['Secuencial'].nunique()),
        'por_categoria': {},
        'detalle_categoria': {},
        'conteo_sucursales': suc_por_cat,
        'df_procesado': df
    }
    for cat in ['Price Club', 'Tiendas', 'Ventas por Mayor', 'Tienda Web', 'Fallas', 'Fundas']:
        df_cat = df[df['Categoria'] == cat]
        resumen['por_categoria'][cat] = int(df_cat['Cantidad_Entera'].sum())
        resumen['detalle_categoria'][cat] = {
            'cantidad': int(df_cat['Cantidad_Entera'].sum()),
            'transf': int(df_cat['Secuencial'].nunique()),
            'unicas': int(df_cat['Sucursal Destino'].nunique()) if 'Sucursal Destino' in df_cat.columns else 0
        }
    return resumen

def generar_reporte_excel_simple(df: pd.DataFrame, fecha_inicio, fecha_fin) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Datos', index=False)
        resumen = pd.DataFrame({
            'Métrica': ['Fecha inicio', 'Fecha fin', 'Total unidades', 'Número de registros'],
            'Valor': [fecha_inicio, fecha_fin, df['Cantidad'].sum() if 'Cantidad' in df.columns else 0, len(df)]
        })
        resumen.to_excel(writer, sheet_name='Resumen', index=False)
    return output.getvalue()

def mostrar_kpi_diario():
    if 'kdi_current_data' not in st.session_state:
        st.session_state.kdi_current_data = pd.DataFrame()
        st.session_state.kdi_loaded = False

    processor = DataProcessor()
    st.markdown("### 📂 Cargar archivo de transferencias diarias")
    col_up1, col_up2 = st.columns([3, 1])
    with col_up1:
        uploaded = st.file_uploader("Seleccionar archivo Excel", type=['xlsx','xls','csv'], key="kdi_upload", label_visibility="collapsed")
    with col_up2:
        if st.button("🔄 Limpiar datos", key="kdi_clear"):
            st.session_state.kdi_current_data = pd.DataFrame()
            st.session_state.kdi_loaded = False
            st.rerun()

    if uploaded:
        with st.spinner("Procesando archivo..."):
            new_data = processor.process_excel_file(uploaded)
            if not new_data.empty:
                st.session_state.kdi_current_data = new_data
                st.session_state.kdi_loaded = True
                st.success("Datos cargados exitosamente")
            else:
                st.warning("El archivo no pudo ser procesado. Revise el formato.")

    if not st.session_state.kdi_loaded or st.session_state.kdi_current_data.empty:
        st.info("👆 Sube un archivo para comenzar el análisis.")
        with st.expander("📋 Estructura esperada del archivo"):
            st.markdown("**Columnas requeridas (se detectan automáticamente):**\n"
                        "- `Producto`: nombre del producto\n"
                        "- `Fecha`: fecha de la transferencia\n"
                        "- `Cantidad`: unidades\n"
                        "- `Bodega Recibe` (o similar): tienda destino\n"
                        "- `Costo` (opcional): costo unitario\n"
                        "Otras columnas como `Línea`, `Grupo`, `Categoría` enriquecen el análisis.")
        return

    st.markdown("### 🔍 Filtros")
    data = st.session_state.kdi_current_data
    filtered = data.copy()

    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    with col_f1:
        if 'FECHA' in filtered.columns and not filtered.empty:
            min_d = filtered['FECHA'].min().date()
            max_d = filtered['FECHA'].max().date()
            dr = st.date_input("Rango de fechas", [min_d, max_d], key="kdi_fecha")
            if len(dr) == 2:
                mask = (filtered['FECHA'].dt.date >= dr[0]) & (filtered['FECHA'].dt.date <= dr[1])
                filtered = filtered[mask].copy()
    with col_f2:
        if 'BODEGA_RECIBE' in filtered.columns:
            opts = ['Todas'] + sorted(filtered['BODEGA_RECIBE'].dropna().unique())
            sel = st.selectbox("Bodega", opts, key="kdi_bod")
            if sel != 'Todas':
                filtered = filtered[filtered['BODEGA_RECIBE'] == sel]
    with col_f3:
        if 'Genero' in filtered.columns:
            opts = ['Todos'] + sorted(filtered['Genero'].dropna().unique())
            sel = st.selectbox("Género", opts, key="kdi_gen")
            if sel != 'Todos':
                filtered = filtered[filtered['Genero'] == sel]
    with col_f4:
        if 'Categoria' in filtered.columns:
            opts = ['Todas'] + sorted(filtered['Categoria'].dropna().unique())
            sel = st.selectbox("Categoría", opts, key="kdi_cat")
            if sel != 'Todas':
                filtered = filtered[filtered['Categoria'] == sel]

    st.markdown("---")
    if filtered.empty:
        st.warning("No hay datos con los filtros actuales.")
        return

    # --- KPIs principales ---
    total_units = int(filtered['CANTIDAD'].sum()) if 'CANTIDAD' in filtered.columns else 0
    total_cost = filtered['TOTAL'].sum() if 'TOTAL' in filtered.columns else 0
    avg_cost_unit = total_cost / total_units if total_units > 0 else 0
    n_bodegas = filtered['BODEGA_RECIBE'].nunique() if 'BODEGA_RECIBE' in filtered.columns else 0
    n_transfers = filtered['Secuencial - Factura'].nunique() if 'Secuencial - Factura' in filtered.columns else len(filtered)
    n_products = filtered['PRODUCTO'].nunique() if 'PRODUCTO' in filtered.columns else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1: st.metric("📦 Unidades Totales", f"{total_units:,}")
    with k2: st.metric("💰 Costo Total", f"${total_cost:,.2f}")
    with k3: st.metric("🏷️ Costo Prom. Unitario", f"${avg_cost_unit:,.2f}")
    with k4: st.metric("🏪 Bodegas Destino", n_bodegas)
    with k5: st.metric("📋 Transferencias", n_transfers)

    st.markdown("---")
    st.markdown("### 📊 Análisis por Dimensiones")
    dim_tab1, dim_tab2, dim_tab3, dim_tab4, dim_tab5, dim_tab6 = st.tabs([
        "🎨 Color", "📏 Talla", "⚧ Género", "🏷️ Categoría/Departamento",
        "📦 Productos", "🔍 Análisis Avanzado"
    ])

    with dim_tab1:
        if 'Color' in filtered.columns:
            col_stats = filtered.groupby('Color').agg({'CANTIDAD':['sum','count']}).reset_index()
            col_stats.columns = ['Color','Unidades','Frecuencia']
            col_stats['Porcentaje'] = (col_stats['Unidades']/total_units*100).round(2)
            col_stats = col_stats.sort_values('Unidades', ascending=False)
            col1, col2 = st.columns([2,1])
            with col1:
                fig = px.pie(col_stats, values='Unidades', names='Color', title="Distribución por Color", color_discrete_sequence=px.colors.qualitative.Set3)
                fig.update_traces(textposition='inside', textinfo='percent+label')
                fig.update_layout(template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            with col2:
                st.dataframe(col_stats[['Color','Unidades','Porcentaje']].style.format({'Unidades':'{:,}','Porcentaje':'{:.2f}%'}), use_container_width=True, height=400)
        else:
            st.info("No hay datos de color disponibles")

    with dim_tab2:
        if 'Talla' in filtered.columns:
            talla_stats = filtered.groupby('Talla').agg({'CANTIDAD':['sum','count']}).reset_index()
            talla_stats.columns = ['Talla','Unidades','Frecuencia']
            talla_stats['Porcentaje'] = (talla_stats['Unidades']/total_units*100).round(2)
            size_order = ['XS','S','M','L','XL','2XL','3XL','4XL','Única']
            talla_stats['Talla_Order'] = talla_stats['Talla'].apply(lambda x: size_order.index(x) if x in size_order else 999)
            talla_stats = talla_stats.sort_values('Talla_Order')
            col1, col2 = st.columns([2,1])
            with col1:
                fig = px.bar(talla_stats, x='Talla', y='Unidades', title="Distribución por Talla", color='Unidades', color_continuous_scale='Viridis')
                fig.update_layout(template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            with col2:
                st.dataframe(talla_stats[['Talla','Unidades','Porcentaje']].style.format({'Unidades':'{:,}','Porcentaje':'{:.2f}%'}), use_container_width=True, height=400)
        else:
            st.info("No hay datos de talla disponibles")

    with dim_tab3:
        if 'Genero' in filtered.columns:
            gen_stats = filtered.groupby('Genero').agg({'CANTIDAD':['sum','count']}).reset_index()
            gen_stats.columns = ['Genero','Unidades','Frecuencia']
            gen_stats['Porcentaje'] = (gen_stats['Unidades']/total_units*100).round(2)
            gen_stats = gen_stats.sort_values('Unidades', ascending=False)
            col1, col2 = st.columns([2,1])
            with col1:
                fig = px.pie(gen_stats, values='Unidades', names='Genero', title="Distribución por Género", color_discrete_sequence=['#FF6B6B','#4ECDC4','#45B7D1','#96CEB4'])
                fig.update_traces(textposition='inside', textinfo='percent+label')
                fig.update_layout(template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            with col2:
                st.dataframe(gen_stats[['Genero','Unidades','Porcentaje']].style.format({'Unidades':'{:,}','Porcentaje':'{:.2f}%'}), use_container_width=True, height=400)
        else:
            st.info("No hay datos de género disponibles")

    with dim_tab4:
        if 'Categoria' in filtered.columns:
            cat_stats = filtered.groupby('Categoria').agg({'CANTIDAD':['sum','count'], 'TOTAL':'sum'}).reset_index()
            cat_stats.columns = ['Categoria','Unidades','Frecuencia','Costo Total']
            cat_stats['Porcentaje'] = (cat_stats['Unidades']/total_units*100).round(2)
            cat_stats = cat_stats.sort_values('Unidades', ascending=False)
            col1, col2 = st.columns([2,1])
            with col1:
                fig = px.bar(cat_stats, x='Categoria', y='Unidades', title="Distribución por Categoría", color='Costo Total', color_continuous_scale='Plasma')
                fig.update_layout(template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            with col2:
                st.dataframe(cat_stats[['Categoria','Unidades','Costo Total','Porcentaje']].style.format({'Unidades':'{:,}','Costo Total':'${:,.2f}','Porcentaje':'{:.2f}%'}), use_container_width=True, height=400)
        else:
            st.info("No hay datos de categoría disponible")

    with dim_tab5:
        st.markdown("### 📊 Análisis por Producto (Top 10 por Costo Total)")
        if 'PRODUCTO' in filtered.columns and 'TOTAL' in filtered.columns:
            top_productos = filtered.groupby('PRODUCTO').agg({'CANTIDAD':'sum','TOTAL':'sum'}).sort_values('TOTAL', ascending=False).head(10).reset_index()
            top_productos.columns = ['Producto','Unidades','Costo Total']
            top_productos['% Costo'] = (top_productos['Costo Total'] / total_cost * 100).round(2)
            col_p1, col_p2 = st.columns([2,1])
            with col_p1:
                fig = px.bar(top_productos, x='Costo Total', y='Producto', orientation='h', title='Top 10 Productos por Costo Total', color='Costo Total', color_continuous_scale='Blues')
                fig.update_layout(yaxis={'categoryorder':'total ascending'}, template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
            with col_p2:
                st.dataframe(top_productos[['Producto','Unidades','Costo Total','% Costo']].style.format({'Unidades':'{:,}','Costo Total':'${:,.2f}','% Costo':'{:.2f}%'}), use_container_width=True, height=350)
        else:
            st.info("No hay datos de costo para este análisis.")

    with dim_tab6:
        st.subheader("🌳 Jerarquía de Productos (Línea → Grupo → Categoría)")
        if all(col in filtered.columns for col in ['LINEA','GRUPO','CATEGORIA']):
            # Preparar datos para Sunburst
            sunburst_cols = ['LINEA','GRUPO','CATEGORIA']
            df_sun = filtered.dropna(subset=sunburst_cols)
            if not df_sun.empty:
                fig_sun = px.sunburst(df_sun, path=['LINEA','GRUPO','CATEGORIA'], values='CANTIDAD', 
                                      title='Distribución por Línea → Grupo → Categoría',
                                      color='CANTIDAD', color_continuous_scale='Viridis')
                fig_sun.update_layout(template="plotly_dark")
                st.plotly_chart(fig_sun, use_container_width=True)

        st.subheader("🗺️ Mapa de Calor – Envíos por Tienda y Día")
        if 'BODEGA_RECIBE' in filtered.columns and 'FECHA' in filtered.columns:
            df_hm = filtered.copy()
            df_hm['DIA_SEMANA'] = df_hm['FECHA'].dt.day_name().str.capitalize()
            dias_orden = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
            df_hm['DIA_SEMANA'] = pd.Categorical(df_hm['DIA_SEMANA'], categories=dias_orden, ordered=True)
            heat_data = df_hm.pivot_table(
                index='BODEGA_RECIBE', columns='DIA_SEMANA', 
                values='CANTIDAD', aggfunc='sum', fill_value=0
            )
            # Limitar a 15 tiendas con más movimiento
            top_tiendas = heat_data.sum(axis=1).nlargest(15).index
            heat_data = heat_data.loc[top_tiendas]
            fig_hm = px.imshow(heat_data, text_auto=True, aspect='auto', 
                               title='Intensidad de envíos (unidades)',
                               labels=dict(x="Día de la semana", y="Tienda", color="Unidades"),
                               color_continuous_scale='YlOrRd')
            fig_hm.update_layout(template="plotly_dark")
            st.plotly_chart(fig_hm, use_container_width=True)

        st.subheader("📊 Costo por Línea de Negocio")
        if 'LINEA' in filtered.columns and 'TOTAL' in filtered.columns:
            linea_cost = filtered.groupby('LINEA').agg({'CANTIDAD':'sum','TOTAL':'sum'}).sort_values('TOTAL', ascending=False).reset_index()
            linea_cost['% Costo'] = (linea_cost['TOTAL'] / total_cost * 100).round(2)
            col_a, col_b = st.columns([2,1])
            with col_a:
                fig_lc = px.bar(linea_cost, x='LINEA', y='TOTAL', title='Costo Total por Línea', color='TOTAL', color_continuous_scale='Teal')
                fig_lc.update_layout(template="plotly_dark")
                st.plotly_chart(fig_lc, use_container_width=True)
            with col_b:
                st.dataframe(linea_cost.style.format({'CANTIDAD':'{:,}','TOTAL':'${:,.2f}','% Costo':'{:.2f}%'}), use_container_width=True)

        st.subheader("🏷️ Distribución por Marca")
        if 'MARCA' in filtered.columns:
            marca_stats = filtered.groupby('MARCA').agg({'CANTIDAD':'sum','TOTAL':'sum'}).sort_values('TOTAL', ascending=False).reset_index()
            marca_stats['% Costo'] = (marca_stats['TOTAL'] / total_cost * 100).round(2)
            col_ma1, col_ma2 = st.columns(2)
            with col_ma1:
                fig_marca = px.pie(marca_stats, values='CANTIDAD', names='MARCA', title='Unidades por Marca')
                fig_marca.update_traces(textposition='inside', textinfo='percent+label')
                fig_marca.update_layout(template="plotly_dark")
                st.plotly_chart(fig_marca, use_container_width=True)
            with col_ma2:
                st.dataframe(marca_stats.style.format({'CANTIDAD':'{:,}','TOTAL':'${:,.2f}','% Costo':'{:.2f}%'}), use_container_width=True)

    # --- Secciones finales (top bodegas, tendencia diaria, detalle, descarga) ---
    st.markdown("---")
    col_g1, col_g2 = st.columns(2)
    with col_g1:
        st.subheader("📊 Top 10 Bodegas por Unidades")
        if 'BODEGA_RECIBE' in filtered.columns:
            top_bod = filtered.groupby('BODEGA_RECIBE')['CANTIDAD'].sum().nlargest(10)
            if not top_bod.empty:
                fig = px.bar(x=top_bod.values, y=top_bod.index, orientation='h', color=top_bod.values, color_continuous_scale='Viridis', labels={'x':'Unidades','y':''})
                fig.update_layout(height=350, template="plotly_dark")
                st.plotly_chart(fig, use_container_width=True)
    with col_g2:
        st.subheader("📈 Tendencia Diaria de Costos")
        if 'FECHA' in filtered.columns and 'TOTAL' in filtered.columns:
            daily = filtered.groupby(filtered['FECHA'].dt.date)['TOTAL'].sum().reset_index()
            daily.columns = ['Fecha','Costo Total']
            fig = px.line(daily, x='Fecha', y='Costo Total', markers=True)
            fig.update_layout(template="plotly_dark")
            st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.subheader("📋 Detalle de Transferencias")
    cols_display = ['FECHA','BODEGA_RECIBE','PRODUCTO','CANTIDAD','COSTO','TOTAL','LINEA','GRUPO','CATEGORIA','MARCA']
    cols_display = [c for c in cols_display if c in filtered.columns]
    st.dataframe(filtered[cols_display].sort_values('FECHA', ascending=False), use_container_width=True, height=300)

    st.markdown("---")
    st.subheader("📄 Generar Reporte")
    col_r1, col_r2, col_r3 = st.columns([1,1,2])
    with col_r1:
        r_start = st.date_input("Fecha inicio", filtered['FECHA'].min().date() if 'FECHA' in filtered.columns else datetime.now().date(), key="r_start")
    with col_r2:
        r_end = st.date_input("Fecha fin", filtered['FECHA'].max().date() if 'FECHA' in filtered.columns else datetime.now().date(), key="r_end")
    with col_r3:
        if st.button("📥 Descargar reporte Excel", use_container_width=True):
            with st.spinner("Generando reporte..."):
                excel_bytes = generar_reporte_excel_simple(filtered, r_start, r_end)
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                st.download_button(label="⬇️ Descargar", data=excel_bytes, file_name=f"KPI_Diario_{ts}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
def mostrar_dashboard_transferencias():
    st.markdown("""
    <div class='main-header'>
        <h1 class='header-title'>📊 Dashboard de Transferencias Diarias</h1>
        <div class='header-subtitle'>Análisis de distribución por categorías y sucursales</div>
    </div>
    """, unsafe_allow_html=True)
    tab1, tab2, tab3 = st.tabs(["📊 Transferencias Diarias", "📦 Mercadería en Tránsito (KPI Diario)", "📈 Análisis de Stock"])
    with tab1:
        st.markdown("""
        <div class='filter-panel'>
            <h4>📂 Carga de Archivo de Transferencias</h4>
            <p class='section-description'>Sube el archivo Excel de transferencias diarias (ej: 322026.xlsx)</p>
        </div>
        """, unsafe_allow_html=True)
        col_u1, col_u2 = st.columns([3,1])
        with col_u1:
            file_diario = st.file_uploader("Selecciona el archivo Excel", type=['xlsx'], key="diario_transferencias", label_visibility="collapsed")
        with col_u2:
            if st.button("🔄 Limpiar", use_container_width=True):
                st.rerun()
        if not file_diario:
            st.info("👆 **Por favor, sube un archivo Excel desde el panel superior para comenzar el análisis.**")
            ejemplo_data = pd.DataFrame({'Secuencial':['TR001','TR002','TR003','TR004'], 'Sucursal Destino':['PRICE CLUB QUITO','AERO MALL DEL SOL','VENTAS POR MAYOR','TIENDA WEB'], 'Cantidad Prendas':[1500,245,5000,120], 'Bodega Destino':['BODEGA CENTRAL','BODEGA NORTE','BODEGA CENTRAL','BODEGA WEB']})
            st.dataframe(ejemplo_data, use_container_width=True)
            st.markdown("### 📝 Columnas requeridas:\n1. **Secuencial**: Número único de transferencia\n2. **Sucursal Destino** o **Bodega Destino**: Nombre de la tienda destino\n3. **Cantidad Prendas**: Cantidad de unidades a transferir")
        else:
            try:
                df_diario = pd.read_excel(file_diario)
                st.success(f"✅ Archivo cargado exitosamente: {file_diario.name}")
                with st.expander("🔍 Vista previa del archivo cargado", expanded=True):
                    st.dataframe(df_diario.head(10), use_container_width=True)
                columnas_requeridas = ['Secuencial','Cantidad Prendas']
                columnas_destino = ['Sucursal Destino','Bodega Destino']
                faltan_requeridas = [col for col in columnas_requeridas if col not in df_diario.columns]
                if faltan_requeridas:
                    st.error(f"❌ **Columnas faltantes:** {faltan_requeridas}")
                else:
                    tiene_destino = any(col in df_diario.columns for col in columnas_destino)
                    if not tiene_destino:
                        st.error("❌ **No se encontró columna de destino.**")
                    else:
                        res = procesar_transferencias_diarias(df_diario)
                        st.header("📈 KPIs por Categoría")
                        categorias_display = {'Price Club':'PRICE CLUB','Tiendas':'TIENDAS AEROPOSTALE','Ventas por Mayor':'VENTAS POR MAYOR','Tienda Web':'TIENDA WEB','Fallas':'FALLAS','Fundas':'FUNDAS'}
                        sucursales_esperadas = {'Price Club':PRICE_CLUBS,'Tiendas':TIENDAS_REGULARES,'Ventas por Mayor':VENTAS_POR_MAYOR,'Tienda Web':TIENDA_WEB,'Fallas':FALLAS,'Fundas':None}
                        color_keys = {'Price Club':'PRICE CLUB','Tiendas':'TIENDAS AEROPOSTALE','Ventas por Mayor':'VENTAS POR MAYOR','Tienda Web':'TIENDA WEB','Fallas':'FALLAS','Fundas':'FUNDAS'}
                        cols = st.columns(3)
                        for i, (cat, cat_display) in enumerate(categorias_display.items()):
                            cantidad = res['por_categoria'].get(cat,0)
                            sucursales_activas = res['conteo_sucursales'].get(cat,0)
                            esperadas = sucursales_esperadas.get(cat)
                            color_key = color_keys.get(cat)
                            bg_gradient = GRADIENTS.get(color_key, 'linear-gradient(135deg, #f0f0f015, #e0e0e030)')
                            border_color = COLORS.get(color_key, '#cccccc')
                            with cols[i%3]:
                                if cat == 'Fundas':
                                    st.markdown(f"<div style='background:{bg_gradient};padding:20px;border-radius:10px;border-left:5px solid {border_color};margin-bottom:15px;'><div style='font-size:12px;color:#666;text-transform:uppercase;margin-bottom:5px;'>{cat_display}</div><div style='font-size:32px;font-weight:bold;color:{border_color};margin-bottom:5px;'>{cantidad:,}</div><div style='font-size:11px;color:#888;'>Múltiplos de 100 ≥ 500 unidades</div></div>", unsafe_allow_html=True)
                                else:
                                    st.markdown(f"<div style='background:{bg_gradient};padding:20px;border-radius:10px;border-left:5px solid {border_color};margin-bottom:15px;'><div style='font-size:12px;color:#666;text-transform:uppercase;margin-bottom:5px;'>{cat_display}</div><div style='font-size:32px;font-weight:bold;color:{border_color};margin-bottom:5px;'>{cantidad:,}</div><div style='font-size:11px;color:#888;'>{sucursales_activas} sucursales | {len(esperadas) if esperadas else 'N/A'} esperadas</div></div>", unsafe_allow_html=True)
                            if i==2:
                                cols = st.columns(3)
                        st.divider()
                        col1, col2 = st.columns([2,1])
                        with col1:
                            df_pie = pd.DataFrame({'Categoria':list(res['por_categoria'].keys()),'Unidades':list(res['por_categoria'].values())})
                            df_pie = df_pie[df_pie['Unidades']>0]
                            if not df_pie.empty:
                                color_map_pie = {'Price Club':COLORS['PRICE CLUB'],'Tiendas':COLORS['TIENDAS AEROPOSTALE'],'Ventas por Mayor':COLORS['VENTAS POR MAYOR'],'Tienda Web':COLORS['TIENDA WEB'],'Fallas':COLORS['FALLAS'],'Fundas':COLORS['FUNDAS']}
                                fig_pie = px.pie(df_pie, values='Unidades', names='Categoria', title="Distribución por Categoría", color='Categoria', color_discrete_map=color_map_pie, hole=0.3)
                                fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                                fig_pie.update_layout(template="plotly_dark")
                                st.plotly_chart(fig_pie, use_container_width=True)
                        with col2:
                            st.subheader("TOTAL GENERAL")
                            st.markdown(f"<div style='background:linear-gradient(135deg,#667eea20,#764ba240);padding:20px;border-radius:10px;text-align:center;'><div style='font-size:36px;font-weight:bold;'>{res['total_unidades']:,}</div></div>", unsafe_allow_html=True)
                            promedio = res['total_unidades']/res['transferencias'] if res['transferencias']>0 else 0
                            st.metric("PROMEDIO X TRANSFERENCIA", f"{promedio:,.0f}")
                            categorias_activas = sum(1 for v in res['por_categoria'].values() if v>0)
                            st.metric("CATEGORÍAS ACTIVAS", f"{categorias_activas}/6")
                            porcentaje_fundas = (res['por_categoria'].get('Fundas',0)/res['total_unidades'])*100 if res['total_unidades']>0 else 0
                            st.metric("% FUNDAS", f"{porcentaje_fundas:.1f}%")
                        st.divider()
                        st.header("📄 Detalle por Secuencial")
                        df_detalle = res['df_procesado'][['Sucursal Destino','Secuencial','Cantidad_Entera','Categoria']].copy()
                        with st.expander("📋 Resumen Estadístico", expanded=True):
                            st.dataframe(pd.DataFrame.from_dict(res['detalle_categoria'], orient='index').reset_index().rename(columns={'index':'Categoria','cantidad':'Unidades','transf':'Transferencias','unicas':'Sucursales Únicas'}), use_container_width=True)
                        col_d1, col_d2 = st.columns([1,4])
                        with col_d1:
                            excel_data = to_excel(df_detalle)
                            st.download_button(label="📥 Descargar Excel", data=excel_data, file_name=f"detalle_transferencias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", use_container_width=True)
                        st.dataframe(df_detalle.rename(columns={'Sucursal Destino':'Sucursal','Cantidad_Entera':'Cantidad','Categoria':'Categoría'}), use_container_width=True, height=400)
            except Exception as e:
                st.error(f"❌ **Error al procesar el archivo:** {str(e)}")
    with tab2:
        mostrar_kpi_diario()
    with tab3:
        st.header("📈 Análisis de Stock y Ventas")
        st.info("🚧 Módulo en desarrollo. Próximamente: análisis ABC, rotación de inventario y predicciones.")

def show_logistica():
    add_back_button(key="back_logistica")
    show_module_header("📦 Dashboard Logístico", "Control de transferencias y distribución de mercadería")
    st.markdown('<div class="module-content">', unsafe_allow_html=True)
    mostrar_dashboard_transferencias()
    st.markdown('</div>', unsafe_allow_html=True)

# ==============================================================================
# MÓDULO: GESTIÓN DE EQUIPO (SIMPLIFICADO, PERO FUNCIONAL) - SIN CAMBIOS
# ==============================================================================
def show_gestion_equipo():
    add_back_button(key="back_equipo")
    show_module_header("👥 Gestión de Equipo", "Directorio del equipo y asistente inteligente para comunicaciones")
    st.markdown('<div class="module-content">', unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────────
    # DATOS DEL EQUIPO (hardcoded)
    # ─────────────────────────────────────────────────────────────────
    EQUIPO_LOGISTICO = {
        "Liderazgo y Control": [
            {"nombre": "Wilson Pérez", "cargo": "Jefe de Logística / Centro de Distribución",
             "area": "Liderazgo", "whatsapp": "", "email": ""}
        ],
        "Transferencias": [
            {"nombre": "César Yépez",    "cargo": "Transferencias Fashion",
             "area": "Transferencias", "whatsapp": "", "email": ""},
            {"nombre": "Luis Perugachi", "cargo": "Pivote de Transferencias y Distribución",
             "area": "Transferencias", "whatsapp": "", "email": ""},
            {"nombre": "Josué Imbacuán","cargo": "Transferencias Tempo",
             "area": "Transferencias", "whatsapp": "", "email": ""}
        ],
        "Distribución": [
            {"nombre": "Jessica Suárez", "cargo": "Distribución",
             "area": "Distribución",   "whatsapp": "", "email": ""},
            {"nombre": "Norma Paredes",  "cargo": "Distribución (recientemente integrada)",
             "area": "Distribución",  "whatsapp": "", "email": ""}
        ],
        "Empaque, Guías y Envíos": [
            {"nombre": "Jhonny Villa",   "cargo": "Encargado de Empaque y Gestión de Guías",
             "area": "Empaque",        "whatsapp": "", "email": ""},
            {"nombre": "Simón Vera",     "cargo": "Apoyo en Guías y Envíos",
             "area": "Empaque",        "whatsapp": "", "email": ""}
        ],
        "Ventas al Por Mayor (Picking y Packing)": [
            {"nombre": "Jhonny Guadalupe","cargo": "Ventas al Por Mayor",
             "area": "Ventas Mayoristas","whatsapp": "","email": ""},
            {"nombre": "Rocío Cadena",   "cargo": "Ventas al Por Mayor",
             "area": "Ventas Mayoristas","whatsapp": "","email": ""}
        ],
        "Cuarentena y Calidad": [
            {"nombre": "Diana García",   "cargo": "Reprocesado de Prendas en Cuarentena",
             "area": "Cuarentena",    "whatsapp": "", "email": ""}
        ]
    }

    # Inicializar session state para contactos y chat
    if "equipo_contactos" not in st.session_state:
        st.session_state.equipo_contactos = {}
    if "chat_gemini" not in st.session_state:
        st.session_state.chat_gemini = []
    if "prompt_rapido" not in st.session_state:
        st.session_state.prompt_rapido = ""

    # ───────────── PESTAÑAS ─────────────
    tab1, tab2 = st.tabs(["👥 Directorio del Equipo", "🤖 Asistente IA — Gemini"])

    # =====================================================================
    # PESTAÑA 1 – DIRECTORIO DEL EQUIPO
    # =====================================================================
    with tab1:
        st.markdown("### 📋 Directorio de Contactos")
        # Calcular métricas
        total_personas = sum(len(miembros) for miembros in EQUIPO_LOGISTICO.values())
        total_areas = len(EQUIPO_LOGISTICO)
        jefe = EQUIPO_LOGISTICO["Liderazgo y Control"][0]["nombre"]

        col1, col2, col3 = st.columns(3)
        col1.metric("👥 Total colaboradores", total_personas)
        col2.metric("📂 Áreas funcionales", total_areas)
        col3.metric("👑 Jefe de Logística", jefe)

        st.markdown("---")

        # Emojis por área para hacerlo visual
        area_emojis = {
            "Liderazgo y Control": "👑",
            "Transferencias": "🔄",
            "Distribución": "📦",
            "Empaque, Guías y Envíos": "📮",
            "Ventas al Por Mayor (Picking y Packing)": "💰",
            "Cuarentena y Calidad": "🔍"
        }

        # Para cada área, un expander
        for area_nombre, miembros in EQUIPO_LOGISTICO.items():
            emoji = area_emojis.get(area_nombre, "👤")
            with st.expander(f"{emoji} {area_nombre} ({len(miembros)} personas)", expanded=True):
                # Distribuir tarjetas en columnas (máximo 3 por fila)
                n_cols = min(len(miembros), 3)
                if n_cols > 0:
                    cols = st.columns(n_cols)
                    for idx, persona in enumerate(miembros):
                        with cols[idx % n_cols]:
                            nombre = persona["nombre"]
                            cargo = persona["cargo"]
                            # Clave única para inputs
                            key_base = nombre.replace(" ", "_").lower()
                            wa_key = f"wa_{key_base}"
                            email_key = f"email_{key_base}"

                            # Obtener valores guardados previamente
                            saved = st.session_state.equipo_contactos.get(nombre, {})
                            wa_val = saved.get("whatsapp", "")
                            email_val = saved.get("email", "")

                            st.markdown(f"**{nombre}**")
                            st.caption(cargo)
                            whatsapp = st.text_input(
                                "📱 WhatsApp",
                                value=wa_val,
                                key=wa_key,
                                placeholder="+593 9XXXXXXXX"
                            )
                            correo = st.text_input(
                                "📧 Email",
                                value=email_val,
                                key=email_key,
                                placeholder="correo@aeropostale.com"
                            )
                            # Almacenar temporalmente en dict local (se guarda al pulsar el botón)
                            # No usamos session_state directamente para no perder datos entre renders
                            # pero sí se necesita un mecanismo de guardado.
                            # Usaremos on_change o guardado manual. Aquí simplemente se guardará
                            # al presionar el botón general, leyendo los valores de los inputs actuales.
                            # (Los inputs ya viven en session_state por defecto con Streamlit)

        # Botón para persistir todos los contactos
        if st.button("💾 Guardar contactos", use_container_width=True):
            nuevos_contactos = {}
            for area_nombre, miembros in EQUIPO_LOGISTICO.items():
                for persona in miembros:
                    nombre = persona["nombre"]
                    key_base = nombre.replace(" ", "_").lower()
                    wa_val = st.session_state.get(f"wa_{key_base}", "")
                    email_val = st.session_state.get(f"email_{key_base}", "")
                    # Guardar solo si no están vacíos
                    if wa_val or email_val:
                        nuevos_contactos[nombre] = {"whatsapp": wa_val, "email": email_val}
            st.session_state.equipo_contactos = nuevos_contactos
            st.success("✅ Contactos guardados correctamente.")
            st.rerun()

    # =====================================================================
    # PESTAÑA 2 – ASISTENTE IA GEMINI
    # =====================================================================
    with tab2:
        st.markdown("### 🤖 Asistente Inteligente — Gemini")

        # Mensaje de bienvenida automático al abrir el chat
        if len(st.session_state.chat_gemini) == 0:
            mensaje_bienvenida = (
                "¡Hola Wilson! 👋 Soy tu asistente Gemini para el Centro de Distribución.\n\n"
                "Puedo ayudarte a:\n"
                "• Redactar y enviar mensajes a tu equipo (WhatsApp o Email)\n"
                "• Enviar reportes a tu jefe Miguel\n"
                "• Consultar información del equipo\n"
                "• Coordinar comunicaciones de operaciones\n\n"
                "¿En qué te ayudo hoy?"
            )
            st.session_state.chat_gemini.append({"role": "assistant", "content": mensaje_bienvenida})

        # Layout de dos columnas: izquierda para acciones rápidas, derecha para el chat
        col_izq, col_der = st.columns([0.3, 0.7])

        # --- PANEL DE ACCIONES RÁPIDAS ---
        with col_izq:
            st.subheader("⚡ Acciones rápidas")
            if st.button("📋 Solicitar actividades diarias al equipo", use_container_width=True):
                st.session_state.prompt_rapido = (
                    "Envía un mensaje a todo mi equipo pidiéndoles que me envíen "
                    "un resumen de las actividades realizadas hoy antes de las 5pm."
                )
                st.rerun()
            if st.button("📦 Informar distribución del día", use_container_width=True):
                st.session_state.prompt_rapido = (
                    "Redacta un comunicado para el equipo de Distribución avisando "
                    "que la carga del día ya está lista para ser recolectada a las 10am."
                )
                st.rerun()
            if st.button("🔄 Reporte de transferencias a jefe", use_container_width=True):
                st.session_state.prompt_rapido = (
                    "Genera un breve reporte de transferencias del día de hoy "
                    "para enviárselo a mi jefe Miguel por correo."
                )
                st.rerun()
            if st.button("📣 Comunicado general al equipo", use_container_width=True):
                st.session_state.prompt_rapido = (
                    "Envía un mensaje a todo el equipo recordándoles la reunión "
                    "de mañana a las 8am en la sala principal."
                )
                st.rerun()

            st.markdown("---")
            if st.button("🗑️ Limpiar conversación", use_container_width=True):
                st.session_state.chat_gemini = []
                st.session_state.prompt_rapido = ""
                st.rerun()

        # --- ÁREA DE CHAT ---
        with col_der:
            st.caption("🤖 Asistente Gemini — Centro de Distribución Aeropostale")
            chat_container = st.container(height=500)

            # Mostrar historial acumulado
            with chat_container:
                for msg in st.session_state.chat_gemini:
                    with st.chat_message(msg["role"]):
                        st.markdown(msg["content"])

            # Procesar prompt rápido si existe (se ejecuta en el siguiente ciclo tras pulsar botón)
            if st.session_state.prompt_rapido:
                prompt = st.session_state.prompt_rapido
                st.session_state.prompt_rapido = ""  # Reset
                # Agregar mensaje del usuario al historial
                st.session_state.chat_gemini.append({"role": "user", "content": prompt})
                # Llamar a Gemini
                with st.spinner("🤖 Pensando..."):
                    respuesta = llamar_groq(prompt, EQUIPO_LOGISTICO, st.session_state.chat_gemini)
                st.session_state.chat_gemini.append({"role": "assistant", "content": respuesta})
                st.rerun()

            # Entrada de chat normal
            user_input = st.chat_input("Escribe tu mensaje o comando...")
            if user_input:
                st.session_state.chat_gemini.append({"role": "user", "content": user_input})
                with st.spinner("🤖 Pensando..."):
                    respuesta = llamar_groq(user_input, EQUIPO_LOGISTICO, st.session_state.chat_gemini)
                st.session_state.chat_gemini.append({"role": "assistant", "content": respuesta})
                st.rerun()

        # Información adicional sobre envío manual
        st.info("📋 **Nota:** Los mensajes generados son **texto listo para copiar y pegar** en "
                "tu WhatsApp o cliente de correo. Cuando haya número registrado, el asistente "
                "incluirá un enlace directo a WhatsApp (wa.me) para facilitar el envío.")

    st.markdown('</div>', unsafe_allow_html=True)

# ==============================================================================
# FUNCIONES AUXILIARES PARA GROQ (ASISTENTE IA)
# ==============================================================================
def llamar_groq(prompt_usuario: str, contexto_equipo: dict, historial: list) -> str:
    """
    Llama a la API de Groq usando el modelo Llama 3.3 70B (gratuito).
    Reintenta automáticamente si hay error de límite (429).
    """
    GROQ_API_KEY = "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"  # ← reemplaza esto
    GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"

    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json"
    }

    # Construir descripción del equipo
    equipo_txt = "EQUIPO DE TRABAJO DEL CENTRO DE DISTRIBUCIÓN AEROPOSTALE:\n"
    contactos = st.session_state.get("equipo_contactos", {})
    for area, miembros in contexto_equipo.items():
        equipo_txt += f"\nÁrea: {area}\n"
        for m in miembros:
            info = contactos.get(m["nombre"], {})
            wa = info.get("whatsapp", "no registrado")
            em = info.get("email", "no registrado")
            equipo_txt += (
                f"  - {m['nombre']} | Cargo: {m['cargo']} | "
                f"WhatsApp: {wa} | Email: {em}\n"
            )

    system_prompt = f"""Eres el asistente inteligente de Wilson Pérez, Jefe de Logística del Centro de Distribución de Aeropostale Ecuador. Su jefe superior es Miguel.

{equipo_txt}

Tu función principal es ayudar a Wilson con la comunicación hacia su equipo:
- Redactar mensajes para enviar por WhatsApp o Email a miembros concretos o grupos.
- Generar reportes breves que se puedan reenviar a Miguel.
- Cuando Wilson pida “enviar un mensaje a…” o “comunicar…”, debes:
    1. Identificar al destinatario exacto (nombre o grupo como “todo el equipo”, “transferencias”, etc.).
    2. Redactar el texto del mensaje de forma profesional, clara y directa.
    3. Indicar el canal recomendado (WhatsApp si hay número, Email si hay correo).
    4. Mostrar el texto listo para que Wilson lo copie y envíe manualmente.
    5. Si tienes un número de WhatsApp, incluir también un enlace wa.me con el mensaje codificado.
        Ejemplo: https://wa.me/593XXXXXXXXX?text=MENSAJE_CODIFICADO
- Si Wilson hace preguntas generales sobre logística, el equipo o las operaciones, responde con información útil y concisa.
- Responde siempre en español formal pero directo. Sé proactivo.
Fecha y hora actual: {datetime.now().strftime('%d/%m/%Y %H:%M')}"""

    # Convertir historial al formato esperado por Groq (OpenAI-like)
    messages = [{"role": "system", "content": system_prompt}]
    for msg in historial[-10:]:  # últimos 10 turnos
        role = "user" if msg["role"] == "user" else "assistant"
        messages.append({"role": role, "content": msg["content"]})
    messages.append({"role": "user", "content": prompt_usuario})

    payload = {
        "model": "llama-3.3-70b-versatile",
        "messages": messages,
        "temperature": 0.7,
        "max_tokens": 1024
    }

    max_reintentos = 3
    for intento in range(max_reintentos):
        try:
            resp = requests.post(
                GROQ_API_URL,
                headers=headers,
                json=payload,
                timeout=30
            )
            if resp.status_code == 200:
                data = resp.json()
                return data["choices"][0]["message"]["content"]
            elif resp.status_code == 429:
                if intento < max_reintentos - 1:
                    espera = 2 ** intento  # 1, 2, 4 segundos
                    st.warning(f"⏳ Límite de uso alcanzado. Reintentando en {espera} segundos...")
                    time.sleep(espera)
                else:
                    return "❌ Límite de solicitudes excedido. Espera un minuto y vuelve a intentarlo."
            else:
                resp.raise_for_status()
        except requests.exceptions.RequestException as e:
            if intento < max_reintentos - 1:
                time.sleep(1)
            else:
                return f"❌ Error al conectar con Groq: {str(e)}"
        except (KeyError, IndexError) as e:
            return "❌ Respuesta inesperada de la API. Intenta nuevamente."

    return "❌ No se pudo obtener respuesta después de varios intentos."
# ==============================================================================
# FUNCIONES AUXILIARES PARA GENERAR GUÍAS
# ==============================================================================
def descargar_logo(url):
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            return response.content
        return None
    except:
        return None

def extraer_datos_transferencia(url: str) -> dict:
    """
    Extrae información de la página de transferencia:
    - Número de transferencia (ej: '85035' de 'N.- TRANSFERENCIAS 00085035')
    - Código (el que viene en URL: codigo=1751139)
    - Otros datos que puedan ser útiles (fecha, proveedor, etc.)
    """
    datos = {
        "numero_transferencia": "",
        "codigo": "",
        "fecha": "",
        "proveedor": "",
        "observacion": "",
        "usuario": "",
        "items": []  # Lista de productos si se desea
    }
    try:
        # Extraer código de la URL
        match_codigo = re.search(r'codigo=(\d+)', url)
        if match_codigo:
            datos["codigo"] = match_codigo.group(1)
        
        # Hacer request a la URL
        response = requests.get(url, timeout=15)
        if response.status_code != 200:
            return datos
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Buscar el texto "N.- TRANSFERENCIAS" y extraer número
        texto_completo = soup.get_text()
        match_trans = re.search(r'N\.-\s*TRANSFERENCIAS\s*0*(\d+)', texto_completo, re.IGNORECASE)
        if match_trans:
            datos["numero_transferencia"] = match_trans.group(1)
        
        # Buscar fecha (patrón común)
        match_fecha = re.search(r'FECHA:\s*(\d{4}-\d{2}-\d{2})', texto_completo)
        if match_fecha:
            datos["fecha"] = match_fecha.group(1)
        
        # Buscar proveedor
        match_prov = re.search(r'PROVEEDOR:\s*([^\n]+)', texto_completo)
        if match_prov:
            datos["proveedor"] = match_prov.group(1).strip()
        
        # Observación y usuario
        match_obs = re.search(r'OBSERVACION:\s*([^\n]+)', texto_completo)
        if match_obs:
            datos["observacion"] = match_obs.group(1).strip()
        match_user = re.search(r'USUARIO:\s*([^\n]+)', texto_completo)
        if match_user:
            datos["usuario"] = match_user.group(1).strip()
        
        # Intentar extraer tabla de items (opcional)
        # Podríamos usar pandas para leer tablas HTML, pero por simplicidad omitimos.
        
    except Exception as e:
        st.warning(f"No se pudieron extraer todos los datos de la URL: {str(e)}")
    
    return datos

# ==============================================================================
# MÓDULO: GENERAR GUÍAS (MEJORADO)
# ==============================================================================
def generar_pdf_profesional(guia_data):
    """
    Genera una guía A4 vertical con diseño limpio, colores Aeropostale (negro),
    iconos, espacio para firmas y número de transferencia sobre el logo.
    """
    buffer = io.BytesIO()
    
    from reportlab.lib.pagesizes import A4
    page_width, page_height = A4
    margen = 1.5 * cm
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(page_width, page_height),
        leftMargin=margen,
        rightMargin=margen,
        topMargin=margen,
        bottomMargin=margen
    )
    
    styles = getSampleStyleSheet()
    
    # Colores (primario = negro)
    color_primario = black
    color_acento = HexColor("#E4002B")
    color_texto = HexColor("#1E293B")
    color_texto_suave = HexColor("#64748B")
    color_fondo = HexColor("#F8FAFC")
    
    # Estilos de texto
    titulo_principal = ParagraphStyle(
        'TituloPrincipal',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=26,
        textColor=color_primario,
        alignment=TA_CENTER,
        spaceAfter=2,
        leading=24
    )
    
    subtitulo_style = ParagraphStyle(
        'Subtitulo',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=color_texto_suave,
        alignment=TA_CENTER,
        spaceAfter=10,
        leading=18
    )
    
    tienda_style = ParagraphStyle(
        'Tienda',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=26,
        textColor=color_primario,
        alignment=TA_CENTER,
        spaceAfter=10,
        leading=18
    )
    
    seccion_title_style = ParagraphStyle(
        'SeccionTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=color_acento,
        alignment=TA_LEFT,
        spaceBefore=6,
        spaceAfter=2,
        leading=13
    )
    
    contenido_style = ParagraphStyle(
        'Contenido',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=color_texto,
        alignment=TA_LEFT,
        spaceAfter=2,
        leading=12
    )
    
    valor_destacado_style = ParagraphStyle(
        'ValorDestacado',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=color_primario,
        alignment=TA_RIGHT,
        spaceAfter=2
    )
    
    fecha_hora_style = ParagraphStyle(
        'FechaHora',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=color_texto_suave,
        alignment=TA_RIGHT,
        spaceAfter=0,
        leading=10
    )
    
    leyenda_qr_style = ParagraphStyle(
        'LeyendaQR',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=color_primario,
        alignment=TA_CENTER,
        spaceBefore=2,
        spaceAfter=0
    )
    
    firma_label_style = ParagraphStyle(
        'FirmaLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=color_primario,
        alignment=TA_CENTER,
        spaceBefore=10
    )
    
    firma_linea_style = ParagraphStyle(
        'FirmaLinea',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=color_texto_suave,
        alignment=TA_CENTER
    )

    contenido = []
    
    # Título principal
    contenido.append(Paragraph("🚚 GUÍA DE REMISIÓN", titulo_principal))
    contenido.append(Paragraph(f"Centro de Distribución {guia_data['marca']}", subtitulo_style))
    contenido.append(Spacer(1, 0.2*cm))
    
    # Nombre de la tienda destino
    tienda_nombre = guia_data.get('tienda_destino', 'Tienda no especificada')
    contenido.append(Paragraph(tienda_nombre, tienda_style))
    contenido.append(Spacer(1, 0.3*cm))
    
    # Número de transferencia (encima del logo)
    num_transferencia = guia_data.get('numero_transferencia', '')
    if num_transferencia:
        trans_style = ParagraphStyle(
            'Transferencia',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=color_primario,
            alignment=TA_CENTER,
            spaceAfter=6
        )
        contenido.append(Paragraph(f"TRANSFERENCIA N°: {num_transferencia}", trans_style))
        contenido.append(Spacer(1, 0.2*cm))
    
    # Datos de la guía (lado derecho)
    datos_tabla = Table(
        [[Paragraph(f"N° Guía: {guia_data['numero']}", valor_destacado_style)],
         [Paragraph(f"Fecha: {guia_data['fecha_emision']}", fecha_hora_style)],
         [Paragraph(f"Hora: {datetime.now().strftime('%H:%M')}", fecha_hora_style)]],
        colWidths=[6*cm]
    )
    datos_tabla.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    datos_fila = Table([[datos_tabla]], colWidths=[page_width - 2*margen])
    datos_fila.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
    ]))
    contenido.append(datos_fila)
    contenido.append(Spacer(1, 0.3*cm))
    
    # Bloque LOGO + QR
    logo_bytes = st.session_state.logos.get(guia_data['marca'])
    if not logo_bytes:
        # URLs de respaldo según marca
        if guia_data['marca'] == "Fashion Club":
            logo_url = "https://raw.githubusercontent.com/wilo3161/kpi-system/main/images/Fashion.jpg"
        elif guia_data['marca'] == "Tempo":
            logo_url = "https://raw.githubusercontent.com/wilo3161/kpi-system/main/images/Tempo.jpg"
        else:
            logo_url = "https://raw.githubusercontent.com/wilo3161/kpi-system/main/images/Aeropostale.jpg"
        logo_bytes = descargar_logo(logo_url)
        if logo_bytes:
            st.session_state.logos[guia_data['marca']] = logo_bytes
    
    if logo_bytes:
        try:
            logo_img = Image(io.BytesIO(logo_bytes), width=5.0*cm, height=5.0*cm)
            logo_cell = logo_img
        except:
            logo_cell = Paragraph(f"<b>{guia_data['marca']}</b>", contenido_style)
    else:
        logo_cell = Paragraph(f"<b>{guia_data['marca']}</b>", contenido_style)
    
    marca_texto = Paragraph(f"<b>{guia_data['marca']}</b>", contenido_style)
    
    bloque_izq = Table(
        [[logo_cell], [marca_texto]],
        colWidths=[5*cm]
    )
    bloque_izq.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    
    qr_bytes = guia_data.get("qr_bytes")
    if qr_bytes:
        try:
            qr_img = Image(io.BytesIO(qr_bytes), width=5.0*cm, height=5.0*cm)
        except:
            qr_img = Paragraph("[QR]", contenido_style)
        qr_leyenda = Paragraph("📱 Escanea aquí tu transferencia", leyenda_qr_style)
        bloque_qr = Table(
            [[qr_img], [qr_leyenda]],
            colWidths=[4*cm]
        )
        bloque_qr.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
    else:
        bloque_qr = Paragraph("", contenido_style)
    
    col_derecha = Table(
        [[bloque_qr]],
        colWidths=[4.5*cm]
    )
    col_derecha.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    
    fila_logo_qr = Table(
        [[bloque_izq, "", col_derecha]],
        colWidths=[5*cm, 9*cm, 4.5*cm]
    )
    fila_logo_qr.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (2,0), (2,0), 'RIGHT'),
    ]))
    contenido.append(fila_logo_qr)
    contenido.append(Spacer(1, 0.5*cm))
    
    # Sección Destinatario
    dest_data = [
        [Paragraph("👤 DESTINATARIO", seccion_title_style)],
        [Paragraph(f"{guia_data['destinatario']}", contenido_style)],
        [Paragraph(f"📞 {guia_data.get('telefono_destinatario', 'No especificado')}", contenido_style)],
        [Paragraph(f"📍 {guia_data['direccion_destinatario']}", contenido_style)]
    ]
    tabla_dest = Table(dest_data, colWidths=[8.5*cm])
    tabla_dest.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BACKGROUND', (0,0), (-1,-1), color_fondo),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (0,0), 4),
        ('BOTTOMPADDING', (-1,0), (-1,0), 4),
    ]))
    
    # Sección Remitente (con posible usuario generador)
    rem_data = [
        [Paragraph("🏢 REMITENTE", seccion_title_style)],
        [Paragraph(f"{guia_data['remitente']}", contenido_style)],
        [Paragraph(f"📍 {guia_data['direccion_remitente']}", contenido_style)]
    ]
    if guia_data.get("usuario_genera"):
        rem_data.append([Paragraph(f"👤 Generado por: {guia_data['usuario_genera']}", contenido_style)])
    
    tabla_rem = Table(rem_data, colWidths=[8.5*cm])
    tabla_rem.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BACKGROUND', (0,0), (-1,-1), color_fondo),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (0,0), 4),
        ('BOTTOMPADDING', (-1,0), (-1,0), 4),
    ]))
    
    contacto_fila = Table(
        [[tabla_dest, tabla_rem]],
        colWidths=[9*cm, 9*cm]
    )
    contacto_fila.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (0,0), 0),
        ('RIGHTPADDING', (0,0), (0,0), 4),
        ('LEFTPADDING', (1,0), (1,0), 4),
        ('RIGHTPADDING', (1,0), (1,0), 0),
    ]))
    contenido.append(contacto_fila)
    contenido.append(Spacer(1, 0.8*cm))
    
    # Firmas
    contenido.append(Spacer(1, 0.3*cm))
    firma_tabla = Table(
        [[Paragraph("_________________________", firma_linea_style),
          Paragraph("_________________________", firma_linea_style)],
         [Paragraph("Revisado por:", firma_label_style),
          Paragraph("Verificado por:", firma_label_style)]],
        colWidths=[9*cm, 9*cm]
    )
    firma_tabla.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
    ]))
    contenido.append(firma_tabla)
    
    # Pie de página
    contenido.append(Spacer(1, 0.3*cm))
    pie = Paragraph(
        "<font size=7 color='#94A3B8'>Documento generado electrónicamente — Válido sin firma</font>",
        ParagraphStyle('Pie', alignment=TA_CENTER)
    )
    contenido.append(pie)
    
    doc.build(contenido)
    buffer.seek(0)
    return buffer.getvalue()
def show_generar_guias():
    """Módulo para generar guías de remisión con código QR y número de transferencia"""
    add_back_button(key="back_guias")
    show_module_header("🚚 Generar Guías de Envío", "Sistema de guías con seguimiento QR")
    st.markdown('<div class="module-content">', unsafe_allow_html=True)

    # --- Constantes de marcas (logos y datos de remitente) ---
    MARCAS = {
        "Aeropostale": {
            "remitente": "AEROPOSTALE - Centro de Distribución",
            "direccion": "Av. Juan Tanca Marengo km 5.5, Guayaquil - Ecuador",
            "logo_url": "https://raw.githubusercontent.com/wilo3161/kpi-system/main/images/Aeropostale.jpg"
        },
        "Fashion Club": {
            "remitente": "FASHION CLUB - Centro de Distribución",
            "direccion": "Av. Principal 123, Guayaquil - Ecuador",
            "logo_url": "https://raw.githubusercontent.com/wilo3161/kpi-system/main/images/Fashion.jpg"
        },
        "Tempo": {
            "remitente": "TEMPO - Centro de Distribución",
            "direccion": "Calle Comercial 456, Quito - Ecuador",
            "logo_url": "https://raw.githubusercontent.com/wilo3161/kpi-system/main/images/Tempo.jpg"
        }
    }

    if "guias_registradas" not in st.session_state:
        st.session_state.guias_registradas = []
    if "contador_guias" not in st.session_state:
        st.session_state.contador_guias = 1000
    if "qr_images" not in st.session_state:
        st.session_state.qr_images = {}
    if "logos" not in st.session_state:
        st.session_state.logos = {}

    st.subheader("📝 Datos de la Guía")

    # --- Selección de marca ---
    marca_seleccionada = st.selectbox("🏷️ Marca", list(MARCAS.keys()))
    marca_info = MARCAS[marca_seleccionada]
    REMITENTE_NOMBRE = marca_info["remitente"]
    REMITENTE_DIRECCION = marca_info["direccion"]

    # --- Selección de tienda (datos predefinidos) ---
    tiendas_opciones = [tienda["Nombre de Tienda"] for tienda in TIENDAS_DATA]
    tienda_seleccionada = st.selectbox("🏪 Tienda Destino", tiendas_opciones)

    tienda_info = next((t for t in TIENDAS_DATA if t["Nombre de Tienda"] == tienda_seleccionada), None)
    if tienda_info:
        destinatario_nombre = tienda_info["Contacto"]
        destinatario_direccion = tienda_info["Dirección"]
        telefono_destino = tienda_info["Teléfono"]
        ciudad_destino = tienda_info["Destino"]
    else:
        destinatario_nombre = ""
        destinatario_direccion = ""
        telefono_destino = ""
        ciudad_destino = ""

    col1, col2 = st.columns(2)
    with col1:
        destinatario = st.text_input("👤 Destinatario (contacto)", value=destinatario_nombre)
        telefono = st.text_input("📞 Teléfono", value=telefono_destino)
    with col2:
        direccion = st.text_area("📍 Dirección completa", value=destinatario_direccion, height=100)
        ciudad = st.text_input("🌆 Ciudad", value=ciudad_destino)

    col3, col4 = st.columns(2)
    with col3:
        peso_kg = st.number_input("⚖️ Peso aproximado (kg)", min_value=0.0, step=0.5, format="%.1f")
    with col4:
        bultos = st.number_input("📦 Número de bultos", min_value=1, step=1, value=1)

    # --- Campo para URL de transferencia ---
    st.subheader("🔗 Datos de Transferencia")
    url_transferencia = st.text_input(
        "URL de la transferencia (ej: https://fashion.sisconti.com/...codigo=1751139...)",
        placeholder="Pega aquí la URL completa de la transferencia"
    )
    if url_transferencia and not url_transferencia.startswith(('http://', 'https://')):
        url_transferencia = 'https://' + url_transferencia

    numero_transferencia = ""
    if url_transferencia:
        # Intentar extraer el número usando la función definida anteriormente
        datos_extraidos = extraer_datos_transferencia(url_transferencia)
        numero_transferencia = datos_extraidos.get("numero_transferencia", "")
        if numero_transferencia:
            st.success(f"✅ Número de transferencia detectado: **{numero_transferencia}**")
        else:
            st.warning("⚠️ No se pudo extraer el número de transferencia automáticamente. Puedes ingresarlo manualmente.")
            numero_transferencia = st.text_input("Número de transferencia (manual)", key="num_transfer_manual")

    observaciones = st.text_area("📝 Observaciones (opcional)", placeholder="Ej: Fragilidad, temperatura controlada, etc.")

    nuevo_numero = st.session_state.contador_guias + 1
    st.info(f"📄 **Número de Guía asignado:** {nuevo_numero}")

    if st.button("🚀 Generar Guía y PDF", type="primary", use_container_width=True):
        if not destinatario or not direccion:
            st.error("❌ Debes completar al menos el destinatario y la dirección.")
        else:
            fecha_emision = datetime.now().strftime("%d/%m/%Y")
            guia_data = {
                "numero": nuevo_numero,
                "marca": marca_seleccionada,
                "tienda_destino": tienda_seleccionada,
                "destinatario": destinatario,
                "telefono_destinatario": telefono,
                "direccion_destinatario": direccion,
                "remitente": REMITENTE_NOMBRE,
                "direccion_remitente": REMITENTE_DIRECCION,
                "fecha_emision": fecha_emision,
                "peso": peso_kg,
                "bultos": bultos,
                "observaciones": observaciones,
                "ciudad": ciudad,
                "numero_transferencia": numero_transferencia,
                "url_transferencia": url_transferencia,
                "usuario_genera": st.session_state.get("user_name", "Usuario")
            }

            # Generar QR con la URL proporcionada por el usuario
            qr_url = url_transferencia if url_transferencia else f"https://aeropostale.ec/seguimiento?guia={nuevo_numero}"
            qr = qrcode.QRCode(box_size=5, border=2)
            qr.add_data(qr_url)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="#0033A0", back_color="white")
            qr_bytes = io.BytesIO()
            qr_img.save(qr_bytes, format="PNG")
            guia_data["qr_bytes"] = qr_bytes.getvalue()

            st.session_state.qr_images[str(nuevo_numero)] = qr_bytes.getvalue()

            try:
                pdf_bytes = generar_pdf_profesional(guia_data)
                st.session_state.guias_registradas.append({
                    "numero": nuevo_numero,
                    "fecha": fecha_emision,
                    "destinatario": destinatario,
                    "tienda": tienda_seleccionada,
                    "pdf": pdf_bytes
                })
                st.session_state.contador_guias = nuevo_numero
                st.success(f"✅ Guía N° {nuevo_numero} generada exitosamente")

                st.download_button(
                    label="📥 Descargar Guía en PDF",
                    data=pdf_bytes,
                    file_name=f"guia_{nuevo_numero}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

                st.markdown("---")
                st.subheader("📱 Código QR de seguimiento")
                st.image(qr_bytes, width=150, caption=f"Guía {nuevo_numero}")
                st.caption("Escanea para rastrear el envío")

            except Exception as e:
                st.error(f"❌ Error al generar el PDF: {str(e)}")

    if st.session_state.guias_registradas:
        st.markdown("---")
        st.subheader("📋 Últimas guías generadas")
        for guia in st.session_state.guias_registradas[-5:]:
            col_a, col_b = st.columns([3,1])
            with col_a:
                st.write(f"**N° {guia['numero']}** - {guia['tienda']} - {guia['fecha']}")
            with col_b:
                if st.button(f"Descargar PDF {guia['numero']}", key=f"re_dl_{guia['numero']}"):
                    st.download_button(
                        label="⬇️ Descargar",
                        data=guia["pdf"],
                        file_name=f"guia_{guia['numero']}.pdf",
                        mime="application/pdf",
                        key=f"dl_{guia['numero']}"
                    )

    st.markdown('</div>', unsafe_allow_html=True)
# ==============================================================================
# MÓDULO: CONTROL DE INVENTARIO (CON ARCHIVO JSON GLOBAL)

# ==============================================================================
def show_control_inventario():
    add_back_button(key="back_inventario")
    show_module_header("📋 Control de Inventario", "Gestión de stock en tiempo real")
    st.markdown('<div class="module-content">', unsafe_allow_html=True)
    if st.session_state.role not in ["Administrador","Bodega"]:
        st.error("⛔ Acceso denegado. Solo Administrador y Bodega pueden acceder a este módulo.")
        st.markdown("</div>", unsafe_allow_html=True)
        return
    INVENTARIO_FILE = "inventario_global.json"
    def cargar_inventario_global():
        try:
            if os.path.exists(INVENTARIO_FILE):
                with open(INVENTARIO_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                df = pd.DataFrame(data['data'])
                tiendas = data['tiendas']
                # Convertir columnas de fecha de vuelta a datetime si es necesario
                if 'FECHA_COMPRA_DT' in df.columns:
                    df['FECHA_COMPRA_DT'] = pd.to_datetime(df['FECHA_COMPRA_DT'])
                if 'FECHA COMPRA' in df.columns:
                    df['FECHA COMPRA'] = pd.to_datetime(df['FECHA COMPRA'])
                return df, tiendas
        except Exception as e:
            st.sidebar.warning(f"No se pudo cargar inventario global: {e}")
        return None, None

    def guardar_inventario_global(df, tiendas):
        try:
            # Crear una copia para no modificar el original
            df_copy = df.copy()
            # Convertir columnas de tipo datetime/Timestamp a string ISO
            for col in df_copy.columns:
                if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                    df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                # También manejar columnas con tipo 'object' que contengan Timestamp
                elif df_copy[col].dtype == 'object':
                    # Intentar convertir a datetime, si funciona, formatear
                    try:
                        if pd.to_datetime(df_copy[col], errors='coerce').notna().all():
                            df_copy[col] = pd.to_datetime(df_copy[col]).dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass
            data = {
                'data': df_copy.to_dict(orient='records'),
                'tiendas': tiendas,
                'fecha_actualizacion': datetime.now().isoformat()
            }
            with open(INVENTARIO_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            st.sidebar.error(f"Error al guardar inventario: {e}")
            return False

    if 'inventario_df' not in st.session_state:
        st.session_state.inventario_df = None
        st.session_state.inventario_tiendas = []
        df_global, tiendas_global = cargar_inventario_global()
        if df_global is not None:
            st.session_state.inventario_df = df_global
            st.session_state.inventario_tiendas = tiendas_global
            st.sidebar.success("📦 Inventario global cargado desde archivo.")
    es_admin = st.session_state.role == "Administrador"
    if es_admin:
        st.sidebar.header("📁 Carga / Actualización de Inventario")
        uploaded_file = st.sidebar.file_uploader("Selecciona archivo Excel de inventario", type=['xlsx','xls'], key="inv_upload_global")
        if uploaded_file is not None:
            if st.sidebar.button("📤 Cargar y Reemplazar Inventario Global"):
                with st.spinner("Procesando archivo..."):
                    try:
                        df = pd.read_excel(uploaded_file)
                        df.columns = df.columns.str.strip().str.upper()
                        cols = df.columns.tolist()
                        start_idx = 0
                        end_idx = len(cols)
                        if "MATRIZ" in cols:
                            start_idx = cols.index("MATRIZ") + 1
                        if "TOTAL" in cols:
                            end_idx = cols.index("TOTAL")
                        tiendas_cols = cols[start_idx:end_idx]
                        excluir = ['FECHA','CODIGO','PRODUCTO','COLECCION','DIVISION','DEPARTAMENTO','STOCK','TOTAL']
                        tiendas_cols = [c for c in tiendas_cols if not any(ex in c for ex in excluir)]
                        for t in ["MATRIZ","PRICE CLUB MATRIZ"]:
                            if t not in tiendas_cols and t in cols:
                                tiendas_cols.append(t)
                        if 'FECHA COMPRA' in df.columns:
                            df['FECHA_COMPRA_DT'] = pd.to_datetime(df['FECHA COMPRA'], errors='coerce')
                            df['DIAS_INVENTARIO'] = (datetime.now() - df['FECHA_COMPRA_DT']).dt.days
                        if guardar_inventario_global(df, tiendas_cols):
                            st.session_state.inventario_df = df
                            st.session_state.inventario_tiendas = tiendas_cols
                            st.sidebar.success(f"✅ Inventario global actualizado: {len(df)} SKUs, {len(tiendas_cols)} tiendas")
                            st.rerun()
                    except Exception as e:
                        st.sidebar.error(f"❌ Error: {e}")
    else:
        if st.session_state.inventario_df is None:
            st.sidebar.info("📌 El Administrador aún no ha cargado el inventario global.")
        else:
            st.sidebar.success("✅ Inventario global disponible (solo lectura)")
    if st.session_state.inventario_df is not None:
        df = st.session_state.inventario_df
        tiendas = st.session_state.inventario_tiendas
        total_skus = len(df)
        total_stock = df[tiendas].sum().sum() if tiendas else 0
        avg_dias = df['DIAS_INVENTARIO'].mean() if 'DIAS_INVENTARIO' in df.columns else 0
        col1, col2, col3 = st.columns(3)
        col1.metric("📦 Total SKUs", f"{total_skus:,}")
        col2.metric("📊 Stock Total", f"{total_stock:,.0f}")
        col3.metric("⏳ Días Promedio en Inventario", f"{avg_dias:.1f} días")
        st.markdown("---")
        st.subheader("🔍 Buscar SKU")
        sku_input = st.text_input("Ingrese el código SKU:", key="sku_search")
        col_b1, col_b2 = st.columns([1,4])
        with col_b1:
            buscar_btn = st.button("Buscar", type="primary")
        with col_b2:
            if st.button("🔄 Limpiar búsqueda"):
                sku_input = ""
                st.rerun()
        if buscar_btn and sku_input:
            sku = sku_input.strip()
            codigo_col = None
            for col in ['CODIGO','COD','SKU','CODIGO PRODUCTO']:
                if col in df.columns:
                    codigo_col = col
                    break
            if codigo_col is None:
                st.error("No se encontró columna de código (CODIGO, SKU, etc.)")
            else:
                mask = df[codigo_col].astype(str).str.strip() == sku
                if mask.any():
                    producto = df[mask].iloc[0]
                    st.success(f"✅ SKU encontrado: **{sku}**")
                    st.markdown("### 📄 Información del Producto")
                    info_cols = ['PRODUCTO','COLECCION','DIVISION','DEPARTAMENTO','FECHA COMPRA']
                    available_info = [c for c in info_cols if c in producto.index]
                    if available_info:
                        cols_info = st.columns(min(len(available_info),3))
                        for idx, col_name in enumerate(available_info):
                            with cols_info[idx%3]:
                                st.markdown(f"**{col_name.title()}**")
                                st.write(producto[col_name])
                    if 'DIAS_INVENTARIO' in producto.index and pd.notna(producto['DIAS_INVENTARIO']):
                        dias = int(producto['DIAS_INVENTARIO'])
                        color_dias = "red" if dias > 90 else "orange" if dias > 60 else "green"
                        st.markdown(f"**Días en inventario:** <span style='color:{color_dias};font-weight:bold;'>{dias}</span>", unsafe_allow_html=True)
                    st.markdown("### 🏪 Stock por Tienda (solo con stock > 0)")
                    stock_data = []
                    for tienda in tiendas:
                        stock_val = producto[tienda] if tienda in producto.index else 0
                        stock_num = int(stock_val) if pd.notna(stock_val) else 0
                        if stock_num > 0:
                            fundas_recomendadas = (stock_num // 100) * 100 if stock_num >= 300 else 0
                            stock_data.append({"Tienda":tienda,"Stock":stock_num,"Fundas Recomendadas":fundas_recomendadas if fundas_recomendadas>0 else "—"})
                    if stock_data:
                        stock_df = pd.DataFrame(stock_data)
                        total_unidades = stock_df['Stock'].sum()
                        col_m1, col_m2 = st.columns(2)
                        col_m1.metric("📦 Total Unidades (con stock)", f"{total_unidades:,}")
                        col_m2.metric("🏪 Tiendas con Stock", f"{len(stock_df)}")
                        st.dataframe(stock_df.style.format({"Stock":"{:,}"}), use_container_width=True, hide_index=True, height=400)
                        st.markdown("#### 📊 Distribución Visual")
                        fig_bar = px.bar(stock_df.sort_values('Stock', ascending=True), x='Stock', y='Tienda', orientation='h', title=f"Stock por tienda (solo con stock) — SKU: {sku}", color='Stock', color_continuous_scale='Blues', text='Stock')
                        fig_bar.update_traces(textposition='outside')
                        fig_bar.update_layout(height=max(300,len(stock_df)*30), showlegend=False, template="plotly_dark")
                        st.plotly_chart(fig_bar, use_container_width=True)
                        csv_sku = stock_df.to_csv(index=False).encode('utf-8')
                        st.download_button("📥 Descargar stock por tienda (CSV)", data=csv_sku, file_name=f"stock_sku_{sku}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", mime="text/csv", use_container_width=True)
                    else:
                        st.info("No hay stock en ninguna tienda para este SKU.")
                else:
                    st.warning(f"⚠️ No se encontró el SKU **'{sku}'**. Verifica el código e intenta de nuevo.")
        st.markdown("---")
        st.subheader("📊 Reportes")
        if st.button("Generar Reporte de Productos Lentos (>90 días)"):
            if 'DIAS_INVENTARIO' not in df.columns:
                st.warning("No hay datos de fecha de compra para calcular días en inventario.")
            else:
                slow = df[df['DIAS_INVENTARIO'] > 90].copy()
                if slow.empty:
                    st.success("No hay productos con más de 90 días en inventario.")
                else:
                    display_cols = ['CODIGO','PRODUCTO','DIAS_INVENTARIO'] + tiendas
                    available = [c for c in display_cols if c in slow.columns]
                    slow_display = slow[available].sort_values('DIAS_INVENTARIO', ascending=False)
                    st.dataframe(slow_display, use_container_width=True)
                    csv = slow_display.to_csv(index=False).encode('utf-8')
                    st.download_button("📥 Descargar CSV", csv, "productos_lentos.csv", "text/csv")
    else:
        st.info("👆 **El Administrador debe cargar el archivo de inventario global** usando el panel lateral.")
        with st.expander("📋 Estructura esperada del archivo"):
            st.markdown("""
            El archivo debe contener al menos las siguientes columnas:
            - **CODIGO** (o SKU): identificador único del producto.
            - **PRODUCTO**: descripción.
            - **FECHA COMPRA** (opcional): para calcular días en inventario.
            - Columnas con nombres de tiendas (ej: "MATRIZ", "MALL DEL SOL", etc.) con el stock.
            - **Importante**: Asegúrese de que existan las columnas "MATRIZ" y "PRICE CLUB MATRIZ" si desea verlas.
            """)
    st.markdown('</div>', unsafe_allow_html=True)
def show_reportes_avanzados():
    add_back_button(key="back_reportes")
    show_module_header("📈 Reportes Avanzados", "Análisis y estadísticas ejecutivas")
    st.markdown('<div class="module-content">', unsafe_allow_html=True)
    st.info("🚧 Módulo en desarrollo. Próximamente: reportes personalizados, análisis predictivo, dashboards ejecutivos.")
    st.markdown('</div>', unsafe_allow_html=True)

def show_configuracion():
    add_back_button(key="back_config")
    show_module_header("⚙️ Configuración", "Personalización del sistema ERP")
    st.markdown('<div class="module-content">', unsafe_allow_html=True)
    tab1, tab2, tab3 = st.tabs(["General", "Usuarios", "Seguridad"])
    
    with tab1:
        st.header("Configuración General")
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("🌐 Configuración Regional")
            st.selectbox("Zona Horaria", ["America/Guayaquil", "UTC"])
            st.selectbox("Moneda", ["USD", "EUR", "COP"])
            st.selectbox("Idioma", ["Español", "Inglés"])
        with col2:
            st.subheader("📊 Configuración de Reportes")
            st.selectbox("Formato de Fecha", ["DD/MM/YYYY", "MM/DD/YYYY", "YYYY-MM-DD"])
            st.slider("Decimales", 0, 4, 2)
            st.selectbox("Separador de Miles", [",", ".", " "])
        if st.button("💾 Guardar Configuración"):
            st.success("✅ Configuración guardada exitosamente")
    
    with tab2:
        st.header("Gestión de Usuarios")
        usuarios = local_db.query("users")
        df_usuarios = pd.DataFrame(usuarios)
        if not df_usuarios.empty:
            st.dataframe(df_usuarios[["username", "role"]], use_container_width=True)
        with st.form("form_usuario"):
            st.subheader("Agregar Nuevo Usuario")
            nuevo_usuario = st.text_input("Nombre de usuario")
            nueva_contrasena = st.text_input("Contraseña", type="password")
            rol = st.selectbox("Rol", ["admin", "user"])
            if st.form_submit_button("➕ Agregar Usuario"):
                if nuevo_usuario and nueva_contrasena:
                    try:
                        local_db.insert("users", {"username": nuevo_usuario, "role": rol, "password": hash_password(nueva_contrasena)})
                        st.success(f"✅ Usuario {nuevo_usuario} agregado exitosamente")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
    
    with tab3:
        st.header("Configuración de Seguridad")
        
        # Sección para cambiar contraseña personal
        with st.expander("🔐 Cambiar mi contraseña", expanded=False):
            st.markdown("### Actualiza tu contraseña personal")
            current_pwd = st.text_input("Contraseña actual", type="password", key="current_pwd")
            new_pwd = st.text_input("Nueva contraseña", type="password", key="new_pwd")
            confirm_pwd = st.text_input("Confirmar nueva contraseña", type="password", key="confirm_pwd")
            if st.button("Actualizar contraseña"):
                username = st.session_state.username
                # Verificar contraseña actual
                user_data = local_db.authenticate(username, current_pwd)
                if user_data:
                    if new_pwd == confirm_pwd and len(new_pwd) >= 4:
                        # Actualizar en la "base de datos"
                        new_hash = hash_password(new_pwd)
                        success = local_db.update_password(username, new_hash)
                        if success:
                            st.success("✅ Contraseña actualizada correctamente. Vuelve a iniciar sesión.")
                        else:
                            st.error("❌ No se pudo actualizar. Contacta al administrador.")
                    else:
                        st.error("❌ Las contraseñas nuevas no coinciden o son muy cortas.")
                else:
                    st.error("❌ Contraseña actual incorrecta.")
        
        st.subheader("🔐 Políticas de Contraseña")
        st.slider("Longitud mínima de contraseña", 6, 20, 8)
        st.checkbox("Requerir mayúsculas", True)
        st.checkbox("Requerir números", True)
        st.selectbox("Expiración de contraseña (días)", ["30", "60", "90", "Nunca"])
        
        st.subheader("🔒 Configuración de Sesión")
        st.slider("Tiempo de inactividad (minutos)", 5, 120, 30)
        st.slider("Máximo de intentos fallidos", 3, 10, 5)
        if st.button("🔒 Aplicar Configuración de Seguridad"):
            st.success("✅ Configuración de seguridad aplicada")
    
    st.markdown('</div>', unsafe_allow_html=True)

# ==============================================================================
# FUNCIÓN PRINCIPAL main()
# ==============================================================================
def main():
    initialize_session_state()
    if not check_password():
        return
    show_header()
    role = st.session_state.role
    allowed_modules = {
        "dashboard_kpis": ["Administrador"],
        "reconciliacion_v8": ["Administrador"],
        "auditoria_correos": ["Administrador"],
        "dashboard_logistico": ["Administrador"],
        "gestion_equipo": ["Administrador"],
        "generar_guias": ["Administrador","Bodega"],
        "control_inventario": ["Administrador","Bodega"],
        "reportes_avanzados": ["Administrador"],
        "configuracion": ["Administrador"],
    }
    current_page = st.session_state.current_page
    if current_page != "Inicio":
        if current_page in allowed_modules:
            if role not in allowed_modules[current_page]:
                st.error("⛔ Acceso denegado. No tienes permiso para ver este módulo.")
                st.session_state.current_page = "Inicio"
                st.rerun()
        else:
            if role != "Administrador":
                st.error("⛔ Acceso denegado.")
                st.session_state.current_page = "Inicio"
                st.rerun()
    page_mapping = {
        "Inicio": show_main_page,
        "dashboard_kpis": show_dashboard_kpis,
        "reconciliacion_v8": show_reconciliacion_v8,
        "auditoria_correos": show_auditoria_correos,
        "dashboard_logistico": show_logistica,
        "gestion_equipo": show_gestion_equipo,
        "generar_guias": show_generar_guias,
        "control_inventario": show_control_inventario,
        "reportes_avanzados": show_reportes_avanzados,
        "configuracion": show_configuracion,
    }
    if current_page in page_mapping:
        page_mapping[current_page]()
    else:
        st.session_state.current_page = "Inicio"
        st.rerun()

if __name__ == '__main__':
    main()
